"""
États — Carte de Stock par marketeur.

Accessible :
  - Marketeur  : /mon-espace/carte-stock/              → carte_stock
  - Admin/staff : /etat/carte-stock/<marketeur_pk>/    → carte_stock_admin

La carte affiche, pour un produit + régime douanier sélectionné :
  • REPORT  : stock cumulé avant le début de la période (ou avant date_debut)
  • Lignes  : mouvements chronologiques avec stock courant après chaque ligne
  • CUMUL   : totaux entrées / sorties / transferts sur la période
"""

from decimal import Decimal
from datetime import date as _date

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum

from .client import marketeur_required, _D


# ─────────────────────────────────────────────────────────────
#  Constantes
# ─────────────────────────────────────────────────────────────

REGIME_SD   = 'SOUS_DOUANE'
REGIME_AC   = 'ACQUITTE'
REGIME_TOUS = 'TOUS'


# ─────────────────────────────────────────────────────────────
#  Helpers internes
# ─────────────────────────────────────────────────────────────

def _delta_sd(mouvement, est_cession_recue=False):
    """
    Calcul du delta @15°C sur le stock SOUS DOUANE pour une ligne.

    SD = ENTREE(SD) + CESSION reçue(SD) − SORTIE(SD) − CESSION émise(SD) − ACQUITTEMENT
    NB : l'ACQUITTEMENT a regime_douanier='ACQUITTE' mais réduit le stock SD.
    """
    t = mouvement.type_mouvement
    if est_cession_recue:
        return _D(mouvement.cession_volume_15c)          # +
    if t == 'ENTREE':
        return _D(mouvement.volume_15c_recu)              # +
    if t == 'SORTIE':
        return -_D(mouvement.volume_15c_sortie)           # −
    if t == 'CESSION':
        return -_D(mouvement.cession_volume_15c)          # −  (émise)
    if t == 'ACQUITTEMENT':
        return -_D(mouvement.acquittement_volume_15c)     # −  (passage vers AC)
    return Decimal('0')


def _delta_sd_amb(mouvement, est_cession_recue=False):
    """Delta volume ambiant SD."""
    t = mouvement.type_mouvement
    if est_cession_recue:
        return _D(mouvement.cession_volume_ambiant)
    if t == 'ENTREE':
        return _D(mouvement.volume_ambiant_recu)
    if t == 'SORTIE':
        return -_D(mouvement.volume_ambiant_sortie)
    if t == 'CESSION':
        return -_D(mouvement.cession_volume_ambiant)
    if t == 'ACQUITTEMENT':
        return -_D(mouvement.acquittement_volume_ambiant)
    return Decimal('0')


def _delta_ac(mouvement, est_cession_recue=False):
    """
    Calcul du delta @15°C sur le stock ACQUITTÉ pour une ligne.

    AC = ENTREE(AC) + ACQUITTEMENT + CESSION reçue(AC) − SORTIE(AC) − CESSION émise(AC)
    """
    t = mouvement.type_mouvement
    if est_cession_recue:
        return _D(mouvement.cession_volume_15c)           # +
    if t == 'ENTREE':
        return _D(mouvement.volume_15c_recu)              # +
    if t == 'ACQUITTEMENT':
        return _D(mouvement.acquittement_volume_15c)      # +  (arrivée depuis SD)
    if t == 'SORTIE':
        return -_D(mouvement.volume_15c_sortie)           # −
    if t == 'CESSION':
        return -_D(mouvement.cession_volume_15c)          # −  (émise)
    return Decimal('0')


def _delta_ac_amb(mouvement, est_cession_recue=False):
    """Delta volume ambiant AC."""
    t = mouvement.type_mouvement
    if est_cession_recue:
        return _D(mouvement.cession_volume_ambiant)
    if t == 'ENTREE':
        return _D(mouvement.volume_ambiant_recu)
    if t == 'ACQUITTEMENT':
        return _D(mouvement.acquittement_volume_ambiant)
    if t == 'SORTIE':
        return -_D(mouvement.volume_ambiant_sortie)
    if t == 'CESSION':
        return -_D(mouvement.cession_volume_ambiant)
    return Decimal('0')


def _qs_sd(marketeur, produit):
    """
    QuerySet des mouvements SD d'un marketeur/produit.
    Inclut les ACQUITTEMENTS (regime=ACQUITTE mais impacte le SD).
    """
    from SGDS.models import Mouvement
    return (
        Mouvement.objects
        .filter(marketeur=marketeur, produit=produit)
        .filter(
            Q(regime_douanier=REGIME_SD) |
            Q(type_mouvement='ACQUITTEMENT')
        )
        .select_related('produit', 'camion', 'chauffeur',
                        'cession_marketeur_destinataire')
    )


def _qs_ac(marketeur, produit):
    """QuerySet des mouvements AC d'un marketeur/produit."""
    from SGDS.models import Mouvement
    return (
        Mouvement.objects
        .filter(marketeur=marketeur, produit=produit,
                regime_douanier=REGIME_AC)
        .select_related('produit', 'camion', 'chauffeur',
                        'cession_marketeur_destinataire')
    )


def _qs_cessions_recues(marketeur, produit, regime):
    """
    QuerySet des cessions reçues par ce marketeur (il est le destinataire).
    On filtre par le régime douanier de la cession.
    """
    from SGDS.models import Mouvement
    return (
        Mouvement.objects
        .filter(
            cession_marketeur_destinataire=marketeur,
            produit=produit,
            type_mouvement='CESSION',
            regime_douanier=regime,
        )
        .select_related('produit', 'camion', 'chauffeur', 'marketeur')
    )


# ─────────────────────────────────────────────────────────────
#  Calcul principal : stock REPORT + lignes enrichies + cumul
# ─────────────────────────────────────────────────────────────

def _calculer_carte_stock(marketeur, produit, regime, date_debut=None, date_fin=None):
    """
    Retourne un dict contenant :
      report_15c   : Decimal — stock @15°C avant date_debut
      report_amb   : Decimal — stock ambiant avant date_debut
      lignes       : list[dict] — mouvements enrichis, triés chronologiquement
      cumul_entrees_15c  : Decimal
      cumul_entrees_amb  : Decimal
      cumul_sorties_15c  : Decimal
      cumul_sorties_amb  : Decimal
      cumul_transferts_15c : Decimal  (ACQUITTEMENT ou CESSION)
      stock_final_15c    : Decimal
      stock_final_amb    : Decimal
    """
    is_sd = (regime == REGIME_SD)
    delta_fn   = _delta_sd   if is_sd else _delta_ac
    delta_fn_a = _delta_sd_amb if is_sd else _delta_ac_amb

    qs_base    = _qs_sd(marketeur, produit) if is_sd else _qs_ac(marketeur, produit)
    qs_recues  = _qs_cessions_recues(marketeur, produit, regime)

    # ── 1. REPORT : inventaire initial + mouvements AVANT date_debut ────────
    report_15c = Decimal('0')
    report_amb = Decimal('0')

    # 1a. Stock inventaire initial (saisi une fois au démarrage)
    from SGDS.models import InventaireInitialMarketeur
    inv = InventaireInitialMarketeur.objects.filter(
        marketeur=marketeur,
        produit=produit,
        regime_douanier=regime,
    ).first()
    if inv:
        # On intègre l'inventaire si sa date est <= date_debut
        # (il représente le stock à l'ouverture, donc inclus dans le REPORT même
        # si la date d'inventaire tombe exactement sur le début de période)
        if date_debut is None or inv.date_inventaire <= date_debut:
            report_15c += inv.volume_15c
            report_amb += inv.volume_ambiant

    # 1b. Cumul des mouvements antérieurs à date_debut
    if date_debut:
        hist_base   = qs_base.filter(date_mouvement__lt=date_debut)
        hist_recues = qs_recues.filter(date_mouvement__lt=date_debut)

        for m in hist_base:
            report_15c += delta_fn(m, est_cession_recue=False)
            report_amb += delta_fn_a(m, est_cession_recue=False)
        for m in hist_recues:
            report_15c += delta_fn(m, est_cession_recue=True)
            report_amb += delta_fn_a(m, est_cession_recue=True)

    # ── 2. Mouvements de la période ───────────────────────────────────────
    qs_periode   = qs_base
    qs_rec_prd   = qs_recues

    if date_debut:
        qs_periode = qs_periode.filter(date_mouvement__gte=date_debut)
        qs_rec_prd = qs_rec_prd.filter(date_mouvement__gte=date_debut)
    if date_fin:
        qs_periode = qs_periode.filter(date_mouvement__lte=date_fin)
        qs_rec_prd = qs_rec_prd.filter(date_mouvement__lte=date_fin)

    # Fusionner les deux QuerySets en une liste de (mouvement, est_cession_recue)
    paires = [(m, False) for m in qs_periode] + [(m, True) for m in qs_rec_prd]
    # Tri chronologique puis par pk pour stabilité
    paires.sort(key=lambda x: (x[0].date_mouvement, x[0].pk))

    # ── 3. Calcul des lignes enrichies + stock courant ─────────────────────
    stock_courant_15c = report_15c
    stock_courant_amb = report_amb

    cumul_entrees_15c    = Decimal('0')
    cumul_entrees_amb    = Decimal('0')
    cumul_sorties_15c    = Decimal('0')
    cumul_sorties_amb    = Decimal('0')
    cumul_transferts_15c = Decimal('0')

    lignes = []

    for mouv, est_recue in paires:
        t = mouv.type_mouvement

        d15 = delta_fn(mouv, est_cession_recue=est_recue)
        dab = delta_fn_a(mouv, est_cession_recue=est_recue)

        stock_courant_15c += d15
        stock_courant_amb += dab

        # ── Catégorisation pour le template ──────────────────────────────
        if est_recue:
            sens = 'cession_recue'
        elif t == 'ENTREE':
            sens = 'entree'
        elif t == 'SORTIE':
            sens = 'sortie'
        elif t == 'CESSION':
            sens = 'cession_emise'
        elif t == 'ACQUITTEMENT':
            sens = 'transfert'
        else:
            sens = 'autre'

        # ── Cumuls ────────────────────────────────────────────────────────
        if d15 > 0 and sens not in ('transfert',):
            cumul_entrees_15c += d15
            cumul_entrees_amb += dab
        elif d15 > 0 and sens == 'transfert':
            # acquittement entrant (côté AC)
            cumul_transferts_15c += d15
        elif d15 < 0 and sens == 'transfert':
            # acquittement sortant (côté SD)
            cumul_transferts_15c += d15  # valeur négative
        elif d15 < 0:
            cumul_sorties_15c += abs(d15)
            cumul_sorties_amb += abs(dab)

        # ── Volume principal affiché selon le sens ─────────────────────
        if sens == 'entree':
            vol_entree_15c = _D(mouv.volume_15c_recu)
            vol_entree_amb = _D(mouv.volume_ambiant_recu)
            vol_sortie_15c = None
            vol_sortie_amb = None
            vol_manquant   = _D(mouv.perte_gain_15c)
        elif sens == 'sortie':
            vol_entree_15c = None
            vol_entree_amb = None
            vol_sortie_15c = _D(mouv.volume_15c_sortie)
            vol_sortie_amb = _D(mouv.volume_ambiant_sortie)
            vol_manquant   = None
        elif sens in ('cession_emise',):
            vol_entree_15c = None
            vol_entree_amb = None
            vol_sortie_15c = _D(mouv.cession_volume_15c)
            vol_sortie_amb = _D(mouv.cession_volume_ambiant)
            vol_manquant   = None
        elif sens == 'cession_recue':
            vol_entree_15c = _D(mouv.cession_volume_15c)
            vol_entree_amb = _D(mouv.cession_volume_ambiant)
            vol_sortie_15c = None
            vol_sortie_amb = None
            vol_manquant   = None
        elif sens == 'transfert':
            # ACQUITTEMENT : côté SD = sortie, côté AC = entrée
            if is_sd:
                vol_entree_15c = None
                vol_entree_amb = None
                vol_sortie_15c = _D(mouv.acquittement_volume_15c)
                vol_sortie_amb = _D(mouv.acquittement_volume_ambiant)
            else:
                vol_entree_15c = _D(mouv.acquittement_volume_15c)
                vol_entree_amb = _D(mouv.acquittement_volume_ambiant)
                vol_sortie_15c = None
                vol_sortie_amb = None
            vol_manquant = None
        else:
            vol_entree_15c = None
            vol_entree_amb = None
            vol_sortie_15c = None
            vol_sortie_amb = None
            vol_manquant   = None

        # ── Origine / Destination affichée ─────────────────────────────
        if sens == 'entree':
            origine_dest = mouv.provenance or '—'
        elif sens == 'sortie':
            origine_dest = mouv.destination or mouv.code_destination or '—'
        elif sens == 'cession_emise':
            dest = mouv.cession_marketeur_destinataire
            origine_dest = str(dest.sigle or dest.raison_sociale) if dest else '—'
        elif sens == 'cession_recue':
            src = mouv.marketeur
            origine_dest = str(src.sigle or src.raison_sociale) if src else '—'
        elif sens == 'transfert':
            origine_dest = 'Acquittement douanier'
        else:
            origine_dest = '—'

        # ── Label régime pour affichage ────────────────────────────────
        t_m = mouv.type_mouvement
        if t_m == 'ACQUITTEMENT':
            regime_ligne = 'SD→AC'
        elif mouv.regime_douanier == REGIME_SD:
            regime_ligne = 'SD'
        elif mouv.regime_douanier == REGIME_AC:
            regime_ligne = 'AC'
        else:
            regime_ligne = mouv.regime_douanier or '—'

        lignes.append({
            'mouvement':         mouv,
            'est_cession_recue': est_recue,
            'sens':              sens,
            'origine_dest':      origine_dest,
            'regime_ligne':      regime_ligne,
            'vol_entree_15c':    vol_entree_15c,
            'vol_entree_amb':    vol_entree_amb,
            'vol_sortie_15c':    vol_sortie_15c,
            'vol_sortie_amb':    vol_sortie_amb,
            'vol_manquant':      vol_manquant,
            'delta_15c':         d15,
            'stock_apres_15c':   stock_courant_15c,
            'stock_apres_amb':   stock_courant_amb,
        })

    return {
        'report_15c':           report_15c,
        'report_amb':           report_amb,
        'lignes':               lignes,
        'cumul_entrees_15c':    cumul_entrees_15c,
        'cumul_entrees_amb':    cumul_entrees_amb,
        'cumul_sorties_15c':    cumul_sorties_15c,
        'cumul_sorties_amb':    cumul_sorties_amb,
        'cumul_transferts_15c': cumul_transferts_15c,
        'stock_final_15c':      stock_courant_15c,
        'stock_final_amb':      stock_courant_amb,
        'is_tous':              False,
    }


# ─────────────────────────────────────────────────────────────
#  Calcul combiné SD + AC (vue "Tous")
# ─────────────────────────────────────────────────────────────

def _calculer_carte_stock_tous(marketeur, produit, date_debut=None, date_fin=None):
    """
    Carte de stock combinée SD + AC.

    Règles :
      • Tous les mouvements (SD, AC, ACQUITTEMENT, cessions) sont affichés.
      • ACQUITTEMENT → delta global = 0  (transfert interne SD↔AC, ne change pas
        le stock total du marketeur).
      • Le stock affiché est uniquement en volume AMBIANT (volume réel).
      • Une colonne 'regime_ligne' identifie le régime de chaque ligne.
    """
    from SGDS.models import Mouvement

    # ── QuerySets ────────────────────────────────────────────────
    qs_all = (
        Mouvement.objects
        .filter(marketeur=marketeur, produit=produit)
        .select_related('produit', 'camion', 'chauffeur',
                        'cession_marketeur_destinataire')
    )
    qs_recues = (
        Mouvement.objects
        .filter(
            cession_marketeur_destinataire=marketeur,
            produit=produit,
            type_mouvement='CESSION',
        )
        .select_related('produit', 'camion', 'chauffeur', 'marketeur')
    )

    # ── Delta ambiant global (ACQUITTEMENT = 0) ───────────────────
    def _delta_global_amb(mouv, est_cession_recue=False):
        t = mouv.type_mouvement
        if est_cession_recue:
            return _D(mouv.cession_volume_ambiant)
        if t == 'ENTREE':
            return _D(mouv.volume_ambiant_recu)
        if t == 'SORTIE':
            return -_D(mouv.volume_ambiant_sortie)
        if t == 'CESSION':
            return -_D(mouv.cession_volume_ambiant)
        if t == 'ACQUITTEMENT':
            return Decimal('0')   # net 0 sur le stock global
        return Decimal('0')

    # ── 1. REPORT : inventaire initial (SD + AC) + mouvements antérieurs ────
    report_amb = Decimal('0')

    # 1a. Inventaire initial — somme SD + AC pour vue "tous régimes"
    from SGDS.models import InventaireInitialMarketeur
    for inv in InventaireInitialMarketeur.objects.filter(
        marketeur=marketeur, produit=produit
    ):
        if date_debut is None or inv.date_inventaire <= date_debut:
            report_amb += inv.volume_ambiant

    # 1b. Mouvements antérieurs à date_debut
    if date_debut:
        for m in qs_all.filter(date_mouvement__lt=date_debut):
            report_amb += _delta_global_amb(m)
        for m in qs_recues.filter(date_mouvement__lt=date_debut):
            report_amb += _delta_global_amb(m, est_cession_recue=True)

    # ── 2. Mouvements de la période ───────────────────────────────
    qs_p   = qs_all
    qs_r   = qs_recues
    if date_debut:
        qs_p = qs_p.filter(date_mouvement__gte=date_debut)
        qs_r = qs_r.filter(date_mouvement__gte=date_debut)
    if date_fin:
        qs_p = qs_p.filter(date_mouvement__lte=date_fin)
        qs_r = qs_r.filter(date_mouvement__lte=date_fin)

    paires = [(m, False) for m in qs_p] + [(m, True) for m in qs_r]
    paires.sort(key=lambda x: (x[0].date_mouvement, x[0].pk))

    # ── 3. Lignes enrichies ───────────────────────────────────────
    stock_courant_amb    = report_amb
    cumul_entrees_amb    = Decimal('0')
    cumul_sorties_amb    = Decimal('0')
    cumul_transferts_amb = Decimal('0')

    lignes = []

    for mouv, est_recue in paires:
        t  = mouv.type_mouvement
        da = _delta_global_amb(mouv, est_cession_recue=est_recue)
        stock_courant_amb += da

        # Sens
        if est_recue:
            sens = 'cession_recue'
        elif t == 'ENTREE':
            sens = 'entree'
        elif t == 'SORTIE':
            sens = 'sortie'
        elif t == 'CESSION':
            sens = 'cession_emise'
        elif t == 'ACQUITTEMENT':
            sens = 'transfert'
        else:
            sens = 'autre'

        # Cumuls (transfert exclu car delta = 0)
        if da > 0:
            cumul_entrees_amb += da
        elif da < 0:
            cumul_sorties_amb += abs(da)

        # Volumes affichés
        if sens == 'entree':
            vol_entree_amb   = _D(mouv.volume_ambiant_recu)
            vol_sortie_amb   = None
            vol_manquant     = _D(mouv.perte_gain_15c)
            vol_manquant_amb = _D(mouv.perte_gain_reception)
        elif sens == 'sortie':
            vol_entree_amb   = None
            vol_sortie_amb   = _D(mouv.volume_ambiant_sortie)
            vol_manquant     = None
            vol_manquant_amb = None
        elif sens == 'cession_emise':
            vol_entree_amb   = None
            vol_sortie_amb   = _D(mouv.cession_volume_ambiant)
            vol_manquant     = None
            vol_manquant_amb = None
        elif sens == 'cession_recue':
            vol_entree_amb   = _D(mouv.cession_volume_ambiant)
            vol_sortie_amb   = None
            vol_manquant     = None
            vol_manquant_amb = None
        elif sens == 'transfert':
            # ACQUITTEMENT : affiché pour info, delta = 0 sur stock global
            vol_entree_amb   = _D(mouv.acquittement_volume_ambiant)  # info
            vol_sortie_amb   = None
            vol_manquant     = None
            vol_manquant_amb = None
        else:
            vol_entree_amb   = None
            vol_sortie_amb   = None
            vol_manquant     = None
            vol_manquant_amb = None

        # Origine / Destination
        if sens == 'entree':
            origine_dest = mouv.provenance or '—'
        elif sens == 'sortie':
            origine_dest = mouv.destination or mouv.code_destination or '—'
        elif sens == 'cession_emise':
            dest = mouv.cession_marketeur_destinataire
            origine_dest = str(dest.sigle or dest.raison_sociale) if dest else '—'
        elif sens == 'cession_recue':
            src = mouv.marketeur
            origine_dest = str(src.sigle or src.raison_sociale) if src else '—'
        elif sens == 'transfert':
            origine_dest = 'Acquittement douanier'
        else:
            origine_dest = '—'

        # Régime de la ligne
        if t == 'ACQUITTEMENT':
            regime_ligne = 'SD→AC'
        elif mouv.regime_douanier == REGIME_SD:
            regime_ligne = 'SD'
        elif mouv.regime_douanier == REGIME_AC:
            regime_ligne = 'AC'
        else:
            regime_ligne = mouv.regime_douanier or '—'

        lignes.append({
            'mouvement':         mouv,
            'est_cession_recue': est_recue,
            'sens':              sens,
            'origine_dest':      origine_dest,
            'regime_ligne':      regime_ligne,
            'vol_entree_amb':    vol_entree_amb,
            'vol_sortie_amb':    vol_sortie_amb,
            'vol_manquant':      vol_manquant,
            'vol_manquant_amb':  vol_manquant_amb,
            'delta_amb':         da,
            'stock_apres_amb':   stock_courant_amb,
            # Champs @15°C fictifs pour compatibilité template SD/AC
            'vol_entree_15c':    None,
            'vol_sortie_15c':    None,
            'stock_apres_15c':   stock_courant_amb,
        })

    return {
        'report_amb':           report_amb,
        'report_15c':           report_amb,
        'lignes':               lignes,
        'cumul_entrees_amb':    cumul_entrees_amb,
        'cumul_entrees_15c':    cumul_entrees_amb,
        'cumul_sorties_amb':    cumul_sorties_amb,
        'cumul_sorties_15c':    cumul_sorties_amb,
        'cumul_transferts_15c': Decimal('0'),
        'stock_final_amb':      stock_courant_amb,
        'stock_final_15c':      stock_courant_amb,
        'is_tous':              True,
    }


# ─────────────────────────────────────────────────────────────
#  Helpers UI
# ─────────────────────────────────────────────────────────────

def _get_filtres(request):
    """Extrait et valide les paramètres GET communs (produit, regime, dates)."""
    return {
        'produit_id':  request.GET.get('produit', '').strip(),
        'regime':      request.GET.get('regime', REGIME_SD).strip(),
        'date_debut':  request.GET.get('date_debut', '').strip(),
        'date_fin':    request.GET.get('date_fin', '').strip(),
    }


def _parse_date(s):
    """Convertit une chaîne YYYY-MM-DD en date ou None."""
    if not s:
        return None
    try:
        from datetime import datetime
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


def _produits_avec_activite(marketeur):
    """Retourne les produits ayant au moins un mouvement pour ce marketeur."""
    from SGDS.models import Produit, Mouvement, InventaireInitialMarketeur
    ids_emis = (
        Mouvement.objects
        .filter(marketeur=marketeur)
        .values_list('produit_id', flat=True)
        .distinct()
    )
    ids_recus = (
        Mouvement.objects
        .filter(cession_marketeur_destinataire=marketeur)
        .values_list('produit_id', flat=True)
        .distinct()
    )
    # Inclure aussi les produits ayant un inventaire initial (même sans mouvement)
    ids_inventaire = (
        InventaireInitialMarketeur.objects
        .filter(marketeur=marketeur)
        .values_list('produit_id', flat=True)
        .distinct()
    )
    all_ids = set(ids_emis) | set(ids_recus) | set(ids_inventaire)
    return (
        Produit.objects
        .filter(id__in=all_ids, statut='ACTIF')
        .select_related('famille')
        .order_by('famille__nom', 'nom')
    )


# ─────────────────────────────────────────────────────────────
#  VUE 1 : Carte de stock — espace marketeur
# ─────────────────────────────────────────────────────────────

@marketeur_required
def carte_stock(request):
    """
    Carte de stock personnelle du marketeur connecté.
    URL : /mon-espace/carte-stock/
    """
    from SGDS.models import Produit

    mkt      = request.user.marketeur
    produits = _produits_avec_activite(mkt)

    filtres    = _get_filtres(request)
    produit_id = filtres['produit_id']
    regime     = filtres['regime'] if filtres['regime'] in (REGIME_SD, REGIME_AC, REGIME_TOUS) else REGIME_SD
    date_debut = _parse_date(filtres['date_debut'])
    date_fin   = _parse_date(filtres['date_fin'])

    produit_sel = None
    carte       = None

    if produit_id:
        try:
            produit_sel = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
        except (Produit.DoesNotExist, ValueError):
            messages.warning(request, "Produit introuvable.")

    if produit_sel:
        if regime == REGIME_TOUS:
            carte = _calculer_carte_stock_tous(mkt, produit_sel, date_debut, date_fin)
        else:
            carte = _calculer_carte_stock(mkt, produit_sel, regime, date_debut, date_fin)

    if regime == REGIME_SD:
        regime_label = 'Sous Douane'
    elif regime == REGIME_AC:
        regime_label = 'Acquitté'
    else:
        regime_label = 'Tous (SD + AC)'

    ctx = {
        'mkt':           mkt,
        'produits':      produits,
        'produit_sel':   produit_sel,
        'regime':        regime,
        'regime_label':  regime_label,
        'date_debut':    filtres['date_debut'],
        'date_fin':      filtres['date_fin'],
        'carte':         carte,
        'REGIME_SD':     REGIME_SD,
        'REGIME_AC':     REGIME_AC,
        'REGIME_TOUS':   REGIME_TOUS,
        'is_admin_view': False,
    }
    return render(request, 'Espace_Marketeur/carte_stock.html', ctx)


# ─────────────────────────────────────────────────────────────
#  VUE 2b : Redirection vers le premier marketeur actif
# ─────────────────────────────────────────────────────────────

@login_required
def etat_carte_stock_redirect(request):
    """Redirige vers la carte de stock du premier marketeur actif."""
    from SGDS.models import Marketeur
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_carte_stock')
    mkt = Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale').first()
    if mkt:
        return redirect('etat_carte_stock', marketeur_pk=mkt.pk)
    messages.warning(request, "Aucun marketeur actif trouvé.")
    return redirect('marketeur_list')


# ─────────────────────────────────────────────────────────────
#  VUE 2 : Carte de stock — accès admin
# ─────────────────────────────────────────────────────────────

@login_required
def carte_stock_admin(request, marketeur_pk):
    """
    Carte de stock d'un marketeur spécifique, vue depuis l'interface admin.
    URL : /etat/carte-stock/<marketeur_pk>/
    Accessible uniquement aux utilisateurs non-marketeur (admin/opérateur/chef).
    """
    from SGDS.models import Produit, Marketeur

    # Vérification : les marketeurs ne peuvent pas consulter les comptes des autres
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_carte_stock')

    mkt      = get_object_or_404(Marketeur, pk=marketeur_pk)
    produits = _produits_avec_activite(mkt)

    # Liste de tous les marketeurs pour la navigation (dropdown)
    tous_marketeurs = Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale')

    filtres    = _get_filtres(request)
    produit_id = filtres['produit_id']
    regime     = filtres['regime'] if filtres['regime'] in (REGIME_SD, REGIME_AC, REGIME_TOUS) else REGIME_SD
    date_debut = _parse_date(filtres['date_debut'])
    date_fin   = _parse_date(filtres['date_fin'])

    produit_sel = None
    carte       = None

    if produit_id:
        try:
            produit_sel = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
        except (Produit.DoesNotExist, ValueError):
            messages.warning(request, "Produit introuvable.")

    if produit_sel:
        if regime == REGIME_TOUS:
            carte = _calculer_carte_stock_tous(mkt, produit_sel, date_debut, date_fin)
        else:
            carte = _calculer_carte_stock(mkt, produit_sel, regime, date_debut, date_fin)

    if regime == REGIME_SD:
        regime_label = 'Sous Douane'
    elif regime == REGIME_AC:
        regime_label = 'Acquitté'
    else:
        regime_label = 'Tous (SD + AC)'

    ctx = {
        'mkt':             mkt,
        'tous_marketeurs': tous_marketeurs,
        'produits':        produits,
        'produit_sel':     produit_sel,
        'regime':          regime,
        'regime_label':    regime_label,
        'date_debut':      filtres['date_debut'],
        'date_fin':        filtres['date_fin'],
        'carte':           carte,
        'REGIME_SD':       REGIME_SD,
        'REGIME_AC':       REGIME_AC,
        'REGIME_TOUS':     REGIME_TOUS,
        'is_admin_view':   True,
    }
    return render(request, 'Etat/carte_stock_admin.html', ctx)


# ─────────────────────────────────────────────────────────────
#  Calcul : stock global dépôt (tous marketeurs ou un seul)
# ─────────────────────────────────────────────────────────────

def _calculer_stock_global(produit, date_debut=None, date_fin=None, marketeur=None):
    """
    Calcul du stock global ambiant pour un produit donné.

    Si `marketeur` est fourni  → stock global de CE marketeur (TOUS régimes).
    Sinon                       → stock global de TOUS les marketeurs.

    Règles de delta global :
      ENTREE      : +volume_ambiant_recu
      SORTIE      : −volume_ambiant_sortie
      ACQUITTEMENT: 0  (transfert SD↔AC interne, affiché pour traçabilité)
      CESSION     : 0  (transfert M1→M2, bilan global nul)

    Retourne un dict :
      report_amb, lignes, cumul_entrees_amb, cumul_sorties_amb,
      cumul_manquants_amb, cumul_manquants_15c, stock_final_amb
    """
    from SGDS.models import Mouvement

    qs = (
        Mouvement.objects
        .filter(produit=produit)
        .select_related('produit', 'camion', 'chauffeur',
                        'marketeur', 'cession_marketeur_destinataire')
    )
    if marketeur:
        qs = qs.filter(marketeur=marketeur)

    def _delta(mouv):
        t = mouv.type_mouvement
        if t == 'ENTREE':
            return _D(mouv.volume_ambiant_recu)
        if t == 'SORTIE':
            return -_D(mouv.volume_ambiant_sortie)
        return Decimal('0')   # ACQUITTEMENT, CESSION → 0

    # ── 1. REPORT : inventaire initial + mouvements antérieurs ──
    from SGDS.models import InventaireInitialMarketeur
    report_amb = Decimal('0')

    # Inventaire initial (SD + AC cumulés pour la vue globale)
    inv_qs = InventaireInitialMarketeur.objects.filter(produit=produit)
    if marketeur:
        inv_qs = inv_qs.filter(marketeur=marketeur)
    for inv in inv_qs:
        if date_debut is None or inv.date_inventaire <= date_debut:
            report_amb += _D(inv.volume_ambiant)

    # Mouvements antérieurs à date_debut
    if date_debut:
        for m in qs.filter(date_mouvement__lt=date_debut):
            report_amb += _delta(m)

    # ── 2. Période ────────────────────────────────────────────
    qs_p = qs
    if date_debut:
        qs_p = qs_p.filter(date_mouvement__gte=date_debut)
    if date_fin:
        qs_p = qs_p.filter(date_mouvement__lte=date_fin)
    mouvements = list(qs_p.order_by('date_mouvement', 'pk'))

    # ── 3. Lignes enrichies ───────────────────────────────────
    stock_courant       = report_amb
    cumul_entrees_amb   = Decimal('0')
    cumul_sorties_amb   = Decimal('0')
    cumul_manquants_amb = Decimal('0')
    cumul_manquants_15c = Decimal('0')
    lignes = []

    for mouv in mouvements:
        t  = mouv.type_mouvement
        da = _delta(mouv)
        stock_courant += da

        if t == 'ENTREE':
            sens = 'entree'
        elif t == 'SORTIE':
            sens = 'sortie'
        elif t == 'CESSION':
            sens = 'cession'
        elif t == 'ACQUITTEMENT':
            sens = 'acquittement'
        else:
            sens = 'autre'

        # Cumuls
        if da > 0:
            cumul_entrees_amb += da
        elif da < 0:
            cumul_sorties_amb += abs(da)

        # Volumes
        if sens == 'entree':
            vol_entree_amb   = _D(mouv.volume_ambiant_recu)
            vol_sortie_amb   = None
            vol_manquant_amb = _D(mouv.perte_gain_reception)
            vol_manquant_15c = _D(mouv.perte_gain_15c)
            if vol_manquant_amb:
                cumul_manquants_amb += abs(vol_manquant_amb)
            if vol_manquant_15c:
                cumul_manquants_15c += abs(vol_manquant_15c)
        elif sens == 'sortie':
            vol_entree_amb   = None
            vol_sortie_amb   = _D(mouv.volume_ambiant_sortie)
            vol_manquant_amb = None
            vol_manquant_15c = None
        else:
            vol_entree_amb   = None
            vol_sortie_amb   = None
            vol_manquant_amb = None
            vol_manquant_15c = None

        # Libellé (provenance / destination réelle)
        if sens == 'entree':
            libelle = mouv.provenance or '—'
        elif sens == 'sortie':
            dest = mouv.destination or mouv.code_destination
            libelle = dest if dest else 'BON DE SORTIE'
        elif sens == 'cession':
            dest_mkt = mouv.cession_marketeur_destinataire
            libelle = (
                f"Cession → {dest_mkt.sigle or dest_mkt.raison_sociale}"
                if dest_mkt else 'Cession'
            )
        elif sens == 'acquittement':
            libelle = 'Acquittement douanier'
        else:
            libelle = '—'

        # Régime de la ligne
        if t == 'ACQUITTEMENT':
            regime_ligne = 'SD→AC'
        elif mouv.regime_douanier == REGIME_SD:
            regime_ligne = 'SD'
        elif mouv.regime_douanier == REGIME_AC:
            regime_ligne = 'AC'
        else:
            regime_ligne = mouv.regime_douanier or '—'

        lignes.append({
            'mouvement':         mouv,
            'sens':              sens,
            'libelle':           libelle,
            'regime_ligne':      regime_ligne,
            'vol_entree_amb':    vol_entree_amb,
            'vol_sortie_amb':    vol_sortie_amb,
            'vol_manquant_amb':  vol_manquant_amb,
            'vol_manquant_15c':  vol_manquant_15c,
            'delta_global':      da,
            'stock_apres_amb':   stock_courant,
        })

    return {
        'report_amb':           report_amb,
        'lignes':               lignes,
        'cumul_entrees_amb':    cumul_entrees_amb,
        'cumul_sorties_amb':    cumul_sorties_amb,
        'cumul_manquants_amb':  cumul_manquants_amb,
        'cumul_manquants_15c':  cumul_manquants_15c,
        'stock_final_amb':      stock_courant,
        'pour_marketeur':       marketeur,
    }


def _produits_avec_mouvements_global():
    """Retourne tous les produits actifs (avec ou sans mouvement)."""
    from SGDS.models import Produit
    return (
        Produit.objects
        .filter(statut='ACTIF')
        .select_related('famille')
        .order_by('famille__nom', 'nom')
    )


# ─────────────────────────────────────────────────────────────
#  VUE 5 : Stock Global — espace marketeur (Mon Stock Global)
# ─────────────────────────────────────────────────────────────

@marketeur_required
def stock_global_marketeur(request):
    """
    Stock global du marketeur connecté (tous régimes SD+AC).
    URL : /mon-espace/stock-global/
    """
    from SGDS.models import Produit

    mkt      = request.user.marketeur
    produits = _produits_avec_activite(mkt)

    produit_id = request.GET.get('produit', '').strip()
    date_debut = _parse_date(request.GET.get('date_debut', '').strip())
    date_fin   = _parse_date(request.GET.get('date_fin', '').strip())

    produit_sel = None
    stock       = None

    if produit_id:
        try:
            produit_sel = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
        except (Produit.DoesNotExist, ValueError):
            messages.warning(request, "Produit introuvable.")

    if produit_sel:
        stock = _calculer_stock_global(
            produit_sel, date_debut, date_fin, marketeur=mkt
        )

    ctx = {
        'mkt':          mkt,
        'produits':     produits,
        'produit_sel':  produit_sel,
        'date_debut':   request.GET.get('date_debut', ''),
        'date_fin':     request.GET.get('date_fin', ''),
        'stock':        stock,
    }
    return render(request, 'Espace_Marketeur/stock_global.html', ctx)


# ─────────────────────────────────────────────────────────────
#  VUE 6 : Stock Global Tout Marketeur — admin
# ─────────────────────────────────────────────────────────────

@login_required
def stock_global_admin(request):
    """
    Stock global de tous les marketeurs pour un produit.
    URL : /etat/stock-global/
    Réservé aux non-marketeurs (admin/opérateur/chef).
    """
    from SGDS.models import Produit, Marketeur

    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_stock_global')

    produits        = _produits_avec_mouvements_global()
    tous_marketeurs = Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale')

    produit_id   = request.GET.get('produit', '').strip()
    marketeur_id = request.GET.get('marketeur', '').strip()
    date_debut   = _parse_date(request.GET.get('date_debut', '').strip())
    date_fin     = _parse_date(request.GET.get('date_fin', '').strip())

    produit_sel   = None
    marketeur_sel = None
    stock         = None

    if produit_id:
        try:
            produit_sel = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
        except (Produit.DoesNotExist, ValueError):
            messages.warning(request, "Produit introuvable.")

    if marketeur_id:
        try:
            marketeur_sel = Marketeur.objects.get(pk=int(marketeur_id))
        except (Marketeur.DoesNotExist, ValueError):
            pass

    if produit_sel:
        stock = _calculer_stock_global(
            produit_sel, date_debut, date_fin,
            marketeur=marketeur_sel  # None = tous les marketeurs
        )

    from SGDS.models import Societe
    ctx = {
        'produits':        produits,
        'tous_marketeurs': tous_marketeurs,
        'produit_sel':     produit_sel,
        'marketeur_sel':   marketeur_sel,
        'date_debut':      request.GET.get('date_debut', ''),
        'date_fin':        request.GET.get('date_fin', ''),
        'stock':           stock,
        'societe':         Societe.get_instance(),
    }
    return render(request, 'Etat/stock_global.html', ctx)


# ─────────────────────────────────────────────────────────────
#  Helpers export Excel — Stock Global
# ─────────────────────────────────────────────────────────────

def _generer_xlsx_stock_global(produit, stock, societe, is_admin=True,
                                marketeur=None, marketeur_sel=None,
                                date_debut=None, date_fin=None):
    """Génère les bytes d'un fichier xlsx Stock Global."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl est requis : pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Stock Global {produit.nom[:18]}"

    couleur_argb = _couleur_hex(getattr(societe, 'couleur_principale', '#1e3a5f'))

    police_base   = Font(name='Arial', size=9)
    police_titre  = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    police_sous   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    police_entete = Font(name='Arial', size=9,  bold=True, color='FFFFFF')
    police_bold   = Font(name='Arial', size=9,  bold=True)
    police_report = Font(name='Arial', size=9,  bold=True, color='0C447C')
    police_cumul  = Font(name='Arial', size=9,  bold=True)

    fill_titre    = PatternFill('solid', fgColor=couleur_argb)
    fill_entete   = PatternFill('solid', fgColor='2F5496')
    fill_report   = PatternFill('solid', fgColor='DBEAFE')
    fill_cumul    = PatternFill('solid', fgColor='E2E8F0')
    fill_entree   = PatternFill('solid', fgColor='F0FDF4')
    fill_sortie   = PatternFill('solid', fgColor='FEF2F2')
    fill_neutre   = PatternFill('solid', fgColor='FFFBEB')

    al_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_l = Alignment(horizontal='left',   vertical='center')
    al_r = Alignment(horizontal='right',  vertical='center')
    bord = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    fmt_vol = '#,##0.00'

    if is_admin:
        # Colonnes admin : Date | N°Doc | N°C | Marketeur | Libellé | Régime | Type
        #                  | Entrées Amb | Mq Amb | Mq @15°C | Sorties Amb | Solde
        COLS = {'A':12,'B':16,'C':12,'D':18,'E':22,'F':8,'G':14,
                'H':14,'I':12,'J':12,'K':14,'L':16}
        NB   = 12
        MH   = 'L'
    else:
        # Colonnes marketeur : Date | N°Doc | N°C | Libellé | Régime | Type
        #                      | Entrées Amb | Mq Amb | Mq @15°C | Sorties Amb | Solde
        COLS = {'A':12,'B':16,'C':12,'D':22,'E':8,'F':14,
                'G':14,'H':12,'I':12,'J':14,'K':16}
        NB   = 11
        MH   = 'K'

    for col, w in COLS.items():
        ws.column_dimensions[col].width = w

    row = 1

    def _merge_full(r, val, font, fill, align=al_c):
        ws.merge_cells(f'A{r}:{MH}{r}')
        c = ws.cell(r, 1)
        c.value = val; c.font = font; c.fill = fill
        c.alignment = align; c.border = bord

    # Ligne 1 : société
    _merge_full(row, f"{societe.raison_sociale}  —  {societe.nom_depot}",
                police_titre, fill_titre)
    ws.row_dimensions[row].height = 22
    row += 1

    # Ligne 2 : infos société
    infos = []
    if societe.adresse:   infos.append(societe.adresse.replace('\n', ' '))
    if societe.ville:     infos.append(societe.ville)
    if societe.telephone: infos.append(f"Tél: {societe.telephone}")
    if societe.email:     infos.append(f"Email: {societe.email}")
    _merge_full(row, '  |  '.join(infos), police_sous, fill_titre)
    ws.row_dimensions[row].height = 16
    row += 1

    ws.row_dimensions[row].height = 6
    row += 1  # séparateur

    # Fiches
    titre_etat = "Stock Global Tout Marketeur" if is_admin else "Mon Stock Global"
    mkt_str = (
        (marketeur_sel.raison_sociale if marketeur_sel else "Tous les marketeurs")
        if is_admin
        else (marketeur.raison_sociale if marketeur else "—")
    )
    infos_fiche = [
        ('État',      titre_etat),
        ('Produit',   str(produit.nom)),
        ('Marketeur', mkt_str),
        ('Période',   (str(date_debut) if date_debut else 'Début')
                     + " → " + (str(date_fin) if date_fin else "Aujourd'hui")),
    ]
    for label, val in infos_fiche:
        ws.merge_cells(f'A{row}:B{row}')
        ws.merge_cells(f'C{row}:{MH}{row}')
        cl = ws.cell(row, 1)
        cl.value = label; cl.font = Font(name='Arial', size=9, bold=True)
        cl.alignment = al_l; cl.border = bord
        cv = ws.cell(row, 3)
        cv.value = val; cv.font = police_base
        cv.alignment = al_l; cv.border = bord
        ws.row_dimensions[row].height = 14
        row += 1

    row += 1  # séparateur

    # En-têtes
    if is_admin:
        hdrs = ['Date', 'N° Document', 'N° C', 'Marketeur',
                'Libellé', 'Rég.', 'Type',
                'Entrées Amb. (L)', 'Mq Amb. (L)', 'Mq @15°C (L)',
                'Sorties Amb. (L)', 'Solde Amb. (L)']
    else:
        hdrs = ['Date', 'N° Document', 'N° C',
                'Libellé', 'Rég.', 'Type',
                'Entrées Amb. (L)', 'Mq Amb. (L)', 'Mq @15°C (L)',
                'Sorties Amb. (L)', 'Solde Amb. (L)']
    ws.row_dimensions[row].height = 28
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row, ci)
        c.value = h; c.font = police_entete; c.fill = fill_entete
        c.alignment = al_c; c.border = bord
    row += 1

    # Ligne REPORT
    ws.merge_cells(f'A{row}:{chr(64 + NB - 1)}{row}')
    c = ws.cell(row, 1)
    c.value = 'REPORT'; c.font = police_report; c.fill = fill_report
    c.alignment = al_c; c.border = bord
    c2 = ws.cell(row, NB)
    c2.value = float(stock['report_amb']); c2.font = police_report
    c2.fill = fill_report; c2.number_format = fmt_vol
    c2.alignment = al_r; c2.border = bord
    ws.row_dimensions[row].height = 14
    row += 1

    SENS_FILLS  = {'entree': fill_entree, 'sortie': fill_sortie}
    SENS_LABELS = {
        'entree': 'Entrée', 'sortie': 'Sortie',
        'cession': 'Cession', 'acquittement': 'Acquittement',
    }

    num_start = 5 if is_admin else 4   # colonne à partir de laquelle appliquer fmt_vol

    for ligne in stock['lignes']:
        m  = ligne['mouvement']
        fl = SENS_FILLS.get(ligne['sens'], fill_neutre)
        ws.row_dimensions[row].height = 13

        date_str = m.date_mouvement.strftime('%d/%m/%Y') if m.date_mouvement else ''
        num_c    = getattr(m, 'numero_c', '') or ''
        mkt_nom  = (m.marketeur.sigle or m.marketeur.raison_sociale) if m.marketeur else '—'
        type_lbl = SENS_LABELS.get(ligne['sens'], ligne['sens'])

        if is_admin:
            data = [
                date_str,
                m.numero_enregistrement or '',
                num_c,
                mkt_nom,
                ligne['libelle'],
                ligne['regime_ligne'],
                type_lbl,
                float(ligne['vol_entree_amb'])   if ligne['vol_entree_amb']   else None,
                float(ligne['vol_manquant_amb'])  if ligne['vol_manquant_amb'] else None,
                float(ligne['vol_manquant_15c'])  if ligne['vol_manquant_15c'] else None,
                float(ligne['vol_sortie_amb'])    if ligne['vol_sortie_amb']   else None,
                float(ligne['stock_apres_amb']),
            ]
            aligns = [al_l,al_l,al_l,al_l,al_l,al_c,al_c,
                      al_r,al_r,al_r,al_r,al_r]
            num_start_row = 8
        else:
            data = [
                date_str,
                m.numero_enregistrement or '',
                num_c,
                ligne['libelle'],
                ligne['regime_ligne'],
                type_lbl,
                float(ligne['vol_entree_amb'])   if ligne['vol_entree_amb']   else None,
                float(ligne['vol_manquant_amb'])  if ligne['vol_manquant_amb'] else None,
                float(ligne['vol_manquant_15c'])  if ligne['vol_manquant_15c'] else None,
                float(ligne['vol_sortie_amb'])    if ligne['vol_sortie_amb']   else None,
                float(ligne['stock_apres_amb']),
            ]
            aligns = [al_l,al_l,al_l,al_l,al_c,al_c,
                      al_r,al_r,al_r,al_r,al_r]
            num_start_row = 7

        for ci, (val, align) in enumerate(zip(data, aligns), 1):
            c = ws.cell(row, ci)
            c.value = val; c.font = police_base
            if fl: c.fill = fl
            c.alignment = align; c.border = bord
            if ci >= num_start_row and val is not None:
                c.number_format = fmt_vol
        row += 1

    # CUMUL
    if stock['lignes']:
        ws.row_dimensions[row].height = 15
        merge_end_col = NB - 4   # 4 colonnes numériques à la fin (Entr, Mq1, Mq2, Sort + Solde = 5)
        ws.merge_cells(f'A{row}:{chr(64 + NB - 5)}{row}')
        c = ws.cell(row, 1)
        c.value = 'CUMUL PÉRIODE'; c.font = police_cumul; c.fill = fill_cumul
        c.alignment = al_c; c.border = bord
        cumul_vals = [
            float(stock['cumul_entrees_amb']),
            float(stock['cumul_manquants_amb']),
            float(stock['cumul_manquants_15c']),
            float(stock['cumul_sorties_amb']),
            float(stock['stock_final_amb']),
        ]
        for ci, val in enumerate(cumul_vals, NB - 4):
            c = ws.cell(row, ci)
            c.value = val; c.font = police_cumul; c.fill = fill_cumul
            c.alignment = al_r; c.border = bord
            c.number_format = fmt_vol
        row += 1

    # Pied de page
    row += 1
    ws.merge_cells(f'A{row}:{MH}{row}')
    c = ws.cell(row, 1)
    pied = getattr(societe, 'pied_de_page', '') or f"Document officiel {societe.raison_sociale}"
    from django.utils import timezone
    c.value = f"{pied}  —  Imprimé le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    c.font  = Font(name='Arial', size=8, italic=True, color='888888')
    c.alignment = al_c

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────
#  VUE 7 : Export Excel — Stock Global marketeur
# ─────────────────────────────────────────────────────────────

@marketeur_required
def stock_global_marketeur_export(request):
    """Export Excel Stock Global — marketeur connecté."""
    from SGDS.models import Produit, Societe
    from django.http import HttpResponse

    mkt        = request.user.marketeur
    produit_id = request.GET.get('produit', '').strip()
    date_debut = _parse_date(request.GET.get('date_debut', '').strip())
    date_fin   = _parse_date(request.GET.get('date_fin', '').strip())

    if not produit_id:
        messages.warning(request, "Sélectionnez un produit avant d'exporter.")
        return redirect('client_stock_global')

    try:
        produit = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
    except (Produit.DoesNotExist, ValueError):
        messages.error(request, "Produit introuvable.")
        return redirect('client_stock_global')

    stock   = _calculer_stock_global(produit, date_debut, date_fin, marketeur=mkt)
    societe = Societe.get_instance()

    try:
        contenu = _generer_xlsx_stock_global(
            produit, stock, societe,
            is_admin=False, marketeur=mkt,
            date_debut=date_debut, date_fin=date_fin,
        )
    except ImportError as e:
        messages.error(request, str(e))
        return redirect('client_stock_global')

    nom = f"stock_global_{produit.nom.replace(' ', '_')}{'_' + str(date_debut) if date_debut else ''}.xlsx"
    resp = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nom}"'
    return resp


# ─────────────────────────────────────────────────────────────
#  VUE 8 : Export Excel — Stock Global admin
# ─────────────────────────────────────────────────────────────

@login_required
def stock_global_admin_export(request):
    """Export Excel Stock Global Tout Marketeur — vue admin."""
    from SGDS.models import Produit, Marketeur, Societe
    from django.http import HttpResponse

    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_stock_global')

    produit_id   = request.GET.get('produit', '').strip()
    marketeur_id = request.GET.get('marketeur', '').strip()
    date_debut   = _parse_date(request.GET.get('date_debut', '').strip())
    date_fin     = _parse_date(request.GET.get('date_fin', '').strip())

    if not produit_id:
        messages.warning(request, "Sélectionnez un produit avant d'exporter.")
        return redirect('etat_stock_global')

    try:
        produit = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
    except (Produit.DoesNotExist, ValueError):
        messages.error(request, "Produit introuvable.")
        return redirect('etat_stock_global')

    marketeur_sel = None
    if marketeur_id:
        try:
            marketeur_sel = Marketeur.objects.get(pk=int(marketeur_id))
        except (Marketeur.DoesNotExist, ValueError):
            pass

    stock   = _calculer_stock_global(produit, date_debut, date_fin, marketeur=marketeur_sel)
    societe = Societe.get_instance()

    try:
        contenu = _generer_xlsx_stock_global(
            produit, stock, societe,
            is_admin=True, marketeur_sel=marketeur_sel,
            date_debut=date_debut, date_fin=date_fin,
        )
    except ImportError as e:
        messages.error(request, str(e))
        return redirect('etat_stock_global')

    mkt_str = f"_{marketeur_sel.sigle or 'mkt'}" if marketeur_sel else "_tous"
    nom = f"stock_global{mkt_str}_{produit.nom.replace(' ', '_')}{'_' + str(date_debut) if date_debut else ''}.xlsx"
    resp = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{nom}"'
    return resp


# ─────────────────────────────────────────────────────────────
#  Helpers export Excel
# ─────────────────────────────────────────────────────────────

def _couleur_hex(hex_str, defaut='1E3A5F'):
    """Retourne un code hex ARGB 8 chars pour openpyxl (ex: 'FF1E3A5F')."""
    h = (hex_str or defaut).lstrip('#').upper()
    if len(h) == 6:
        return 'FF' + h
    return 'FF' + defaut


def _generer_xlsx_carte(marketeur, produit, regime, carte, societe, date_debut=None, date_fin=None):
    """
    Génère un fichier xlsx carte de stock et retourne les bytes.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl est requis : pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    is_tous = carte.get('is_tous', False)
    if regime == REGIME_SD:
        regime_label = 'Sous Douane (SD)'
    elif regime == REGIME_AC:
        regime_label = 'Acquitté (AC)'
    else:
        regime_label = 'Tous SD+AC'
    ws.title = f"Carte Stock {produit.nom[:12]} {regime_label[:2]}"

    # ── Couleur principale de la société ─────────────────────
    couleur_argb = _couleur_hex(getattr(societe, 'couleur_principale', '#1e3a5f'))

    # ── Styles ────────────────────────────────────────────────
    police_base   = Font(name='Arial', size=9)
    police_titre  = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    police_sous   = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    police_entete = Font(name='Arial', size=9,  bold=True, color='FFFFFF')
    police_bold   = Font(name='Arial', size=9,  bold=True)
    police_report = Font(name='Arial', size=9,  bold=True, color='0C447C')
    police_cumul  = Font(name='Arial', size=9,  bold=True)

    fill_titre    = PatternFill('solid', fgColor=couleur_argb)
    fill_entete   = PatternFill('solid', fgColor='2F5496')
    fill_report   = PatternFill('solid', fgColor='DBEAFE')
    fill_cumul    = PatternFill('solid', fgColor='E2E8F0')
    fill_entree   = PatternFill('solid', fgColor='F0FDF4')
    fill_sortie   = PatternFill('solid', fgColor='FEF2F2')
    fill_cession  = PatternFill('solid', fgColor='EFF6FF')
    fill_transfert= PatternFill('solid', fgColor='FFFBEB')

    al_centre = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_gauche = Alignment(horizontal='left',   vertical='center')
    al_droite = Alignment(horizontal='right',  vertical='center')

    bord = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    fmt_vol  = '#,##0.00'

    # ── Colonnes ──────────────────────────────────────────────
    if is_tous:
        # A:Date B:N°Doc C:N°C D:Régime E:Origine/Dest F:Type
        # G:Entrées Amb H:Manquant I:Sorties Amb J:Stock Amb
        COLS      = {'A':12, 'B':16, 'C':12, 'D':10, 'E':22, 'F':14,
                     'G':14, 'H':12, 'I':14, 'J':16}
        NB_COLS   = 10
        MERGE_HDR = 'J'
    else:
        # A:Date B:N°Doc C:N°C D:Origine/Dest E:Type
        # F:Entrées@15° G:Entrées Amb H:Manquant I:Sorties@15° J:Sorties Amb K:Stock Amb
        COLS      = {'A':12, 'B':16, 'C':12, 'D':22, 'E':14,
                     'F':14, 'G':14, 'H':12, 'I':14, 'J':14, 'K':16}
        NB_COLS   = 11
        MERGE_HDR = 'K'
    for col_letter, width in COLS.items():
        ws.column_dimensions[col_letter].width = width

    row = 1

    # ── Ligne 1 : Raison sociale + dépôt ─────────────────────
    ws.merge_cells(f'A{row}:{MERGE_HDR}{row}')
    c = ws.cell(row, 1)
    c.value = f"{societe.raison_sociale}  —  {societe.nom_depot}"
    c.font  = police_titre
    c.fill  = fill_titre
    c.alignment = al_centre
    c.border = bord
    ws.row_dimensions[row].height = 22
    row += 1

    # ── Ligne 2 : Adresse + contacts ─────────────────────────
    ws.merge_cells(f'A{row}:{MERGE_HDR}{row}')
    c = ws.cell(row, 1)
    infos = []
    if societe.adresse:  infos.append(societe.adresse.replace('\n', ' '))
    if societe.ville:    infos.append(societe.ville)
    if societe.telephone: infos.append(f"Tél: {societe.telephone}")
    if societe.email:    infos.append(f"Email: {societe.email}")
    c.value = '  |  '.join(infos) if infos else ''
    c.font  = police_sous
    c.fill  = fill_titre
    c.alignment = al_centre
    c.border = bord
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Ligne 3 : vide séparateur ─────────────────────────────
    ws.row_dimensions[row].height = 6
    row += 1

    # ── Lignes 4-7 : Fiche carte ─────────────────────────────
    infos_carte = [
        ('Marketeur',  str(marketeur.raison_sociale)),
        ('Produit',    str(produit.nom)),
        ('Régime',     regime_label),
        ('Période',    (str(date_debut) if date_debut else 'Début') + " → " + (str(date_fin) if date_fin else "Aujourd'hui")),
    ]
    for label, val in infos_carte:
        ws.merge_cells(f'A{row}:B{row}')
        ws.merge_cells(f'C{row}:{MERGE_HDR}{row}')
        cl = ws.cell(row, 1)
        cl.value = label; cl.font = Font(name='Arial', size=9, bold=True)
        cl.alignment = al_gauche; cl.border = bord
        cv = ws.cell(row, 3)
        cv.value = val; cv.font = police_base
        cv.alignment = al_gauche; cv.border = bord
        ws.row_dimensions[row].height = 14
        row += 1

    row += 1  # séparateur

    # ── En-têtes colonnes ─────────────────────────────────────
    if is_tous:
        headers = ['Date', 'N° Document', 'N° C', 'Régime',
                   'Origine / Destination', 'Type',
                   'Entrées Amb. (L)', 'Manquant (L)',
                   'Sorties Amb. (L)', 'Stock Amb. (L)']
    else:
        headers = ['Date', 'N° Document', 'N° C', 'Origine / Destination', 'Type',
                   'Entrées @15°C (L)', 'Entrées Amb. (L)', 'Manquant (L)',
                   'Sorties @15°C (L)', 'Sorties Amb. (L)', 'Stock Amb. (L)']
    ws.row_dimensions[row].height = 28
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row, col_idx)
        c.value = h; c.font = police_entete; c.fill = fill_entete
        c.alignment = al_centre; c.border = bord
    row += 1

    # ── Ligne REPORT ──────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    nb_merge_report = NB_COLS - 1
    ws.merge_cells(f'A{row}:{chr(64 + nb_merge_report)}{row}')
    c = ws.cell(row, 1)
    c.value = 'REPORT'; c.font = police_report; c.fill = fill_report
    c.alignment = al_centre; c.border = bord
    c_stock = ws.cell(row, NB_COLS)
    c_stock.value = float(carte['report_amb']); c_stock.font = police_report
    c_stock.fill = fill_report; c_stock.number_format = fmt_vol
    c_stock.alignment = al_droite; c_stock.border = bord
    row += 1

    # ── Lignes mouvements ─────────────────────────────────────
    SENS_FILLS = {
        'entree':         fill_entree,
        'sortie':         fill_sortie,
        'cession_emise':  fill_cession,
        'cession_recue':  fill_cession,
        'transfert':      fill_transfert,
    }
    SENS_LABELS = {
        'entree':        'Entrée',
        'sortie':        'Sortie',
        'cession_emise': 'Cession émise',
        'cession_recue': 'Cession reçue',
        'transfert':     'Acquittement',
    }

    for ligne in carte['lignes']:
        m  = ligne['mouvement']
        fl = SENS_FILLS.get(ligne['sens'])
        ws.row_dimensions[row].height = 13

        if is_tous:
            data = [
                m.date_mouvement.strftime('%d/%m/%Y') if m.date_mouvement else '',
                m.numero_enregistrement or '',
                getattr(m, 'numero_c', '') or '',
                ligne.get('regime_ligne', '—'),
                ligne['origine_dest'],
                SENS_LABELS.get(ligne['sens'], ''),
                float(ligne['vol_entree_amb']) if ligne['vol_entree_amb'] else None,
                float(ligne['vol_manquant'])   if ligne['vol_manquant']   else None,
                float(ligne['vol_sortie_amb']) if ligne['vol_sortie_amb'] else None,
                float(ligne['stock_apres_amb']),
            ]
            aligns = [al_gauche, al_gauche, al_gauche, al_centre, al_gauche, al_centre,
                      al_droite, al_droite, al_droite, al_droite]
            num_start = 7
        else:
            data = [
                m.date_mouvement.strftime('%d/%m/%Y') if m.date_mouvement else '',
                m.numero_enregistrement or '',
                getattr(m, 'numero_c', '') or '',
                ligne['origine_dest'],
                SENS_LABELS.get(ligne['sens'], ''),
                float(ligne['vol_entree_15c']) if ligne['vol_entree_15c'] else None,
                float(ligne['vol_entree_amb']) if ligne['vol_entree_amb'] else None,
                float(ligne['vol_manquant'])   if ligne['vol_manquant']   else None,
                float(ligne['vol_sortie_15c']) if ligne['vol_sortie_15c'] else None,
                float(ligne['vol_sortie_amb']) if ligne['vol_sortie_amb'] else None,
                float(ligne['stock_apres_amb']),
            ]
            aligns = [al_gauche, al_gauche, al_gauche, al_gauche, al_centre,
                      al_droite, al_droite, al_droite, al_droite, al_droite, al_droite]
            num_start = 6

        for col_idx, (val, align) in enumerate(zip(data, aligns), start=1):
            c = ws.cell(row, col_idx)
            c.value = val; c.font = police_base
            if fl: c.fill = fl
            c.alignment = align; c.border = bord
            if col_idx >= num_start and val is not None:
                c.number_format = fmt_vol
        row += 1

    # ── Ligne CUMUL ───────────────────────────────────────────
    if carte['lignes']:
        ws.row_dimensions[row].height = 15
        if is_tous:
            ws.merge_cells(f'A{row}:F{row}')
            c = ws.cell(row, 1)
            c.value = 'CUMUL PÉRIODE'; c.font = police_cumul; c.fill = fill_cumul
            c.alignment = al_centre; c.border = bord
            cumul_vals = [
                float(carte['cumul_entrees_amb']),
                None,
                float(carte['cumul_sorties_amb']),
                float(carte['stock_final_amb']),
            ]
            for col_idx, val in enumerate(cumul_vals, start=7):
                c = ws.cell(row, col_idx)
                c.value = val; c.font = police_cumul; c.fill = fill_cumul
                c.alignment = al_droite; c.border = bord
                if val is not None: c.number_format = fmt_vol
        else:
            ws.merge_cells(f'A{row}:E{row}')
            c = ws.cell(row, 1)
            c.value = 'CUMUL PÉRIODE'; c.font = police_cumul; c.fill = fill_cumul
            c.alignment = al_centre; c.border = bord
            cumul_vals = [
                float(carte['cumul_entrees_15c']),
                float(carte['cumul_entrees_amb']),
                None,
                float(carte['cumul_sorties_15c']),
                float(carte['cumul_sorties_amb']),
                float(carte['stock_final_amb']),
            ]
            for col_idx, val in enumerate(cumul_vals, start=6):
                c = ws.cell(row, col_idx)
                c.value = val; c.font = police_cumul; c.fill = fill_cumul
                c.alignment = al_droite; c.border = bord
                if val is not None: c.number_format = fmt_vol
        row += 1

    # ── Pied de page ─────────────────────────────────────────
    row += 1
    ws.merge_cells(f'A{row}:{MERGE_HDR}{row}')
    c = ws.cell(row, 1)
    pied = getattr(societe, 'pied_de_page', '') or f"Document officiel {societe.raison_sociale}"
    from django.utils import timezone
    c.value = f"{pied}  —  Imprimé le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    c.font  = Font(name='Arial', size=8, italic=True, color='888888')
    c.alignment = al_centre

    # ── Sauvegarder en bytes ──────────────────────────────────
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────
#  VUE 3 : Export Excel — espace marketeur
# ─────────────────────────────────────────────────────────────

@marketeur_required
def carte_stock_export(request):
    """Export Excel carte de stock — marketeur connecté."""
    from SGDS.models import Produit, Societe
    from django.http import HttpResponse

    mkt        = request.user.marketeur
    filtres    = _get_filtres(request)
    produit_id = filtres['produit_id']
    regime     = filtres['regime'] if filtres['regime'] in (REGIME_SD, REGIME_AC, REGIME_TOUS) else REGIME_SD
    date_debut = _parse_date(filtres['date_debut'])
    date_fin   = _parse_date(filtres['date_fin'])

    if not produit_id:
        messages.warning(request, "Sélectionnez un produit avant d'exporter.")
        return redirect('client_carte_stock')

    try:
        produit = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
    except (Produit.DoesNotExist, ValueError):
        messages.error(request, "Produit introuvable.")
        return redirect('client_carte_stock')

    if regime == REGIME_TOUS:
        carte = _calculer_carte_stock_tous(mkt, produit, date_debut, date_fin)
    else:
        carte = _calculer_carte_stock(mkt, produit, regime, date_debut, date_fin)
    societe = Societe.get_instance()

    try:
        contenu = _generer_xlsx_carte(mkt, produit, regime, carte, societe, date_debut, date_fin)
    except ImportError as e:
        messages.error(request, str(e))
        return redirect('client_carte_stock')

    nom_fichier = (
        f"carte_stock_{produit.nom.replace(' ', '_')}"
        f"_{'SD' if regime == REGIME_SD else 'AC'}"
        f"{'_' + str(date_debut) if date_debut else ''}.xlsx"
    )
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response


# ─────────────────────────────────────────────────────────────
#  VUE 4 : Export Excel — accès admin
# ─────────────────────────────────────────────────────────────

@login_required
def carte_stock_export_admin(request, marketeur_pk):
    """Export Excel carte de stock — vue admin."""
    from SGDS.models import Produit, Marketeur, Societe
    from django.http import HttpResponse

    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_carte_stock')

    mkt        = get_object_or_404(Marketeur, pk=marketeur_pk)
    filtres    = _get_filtres(request)
    produit_id = filtres['produit_id']
    regime     = filtres['regime'] if filtres['regime'] in (REGIME_SD, REGIME_AC, REGIME_TOUS) else REGIME_SD
    date_debut = _parse_date(filtres['date_debut'])
    date_fin   = _parse_date(filtres['date_fin'])

    if not produit_id:
        messages.warning(request, "Sélectionnez un produit avant d'exporter.")
        return redirect('etat_carte_stock', marketeur_pk=marketeur_pk)

    try:
        produit = Produit.objects.get(pk=int(produit_id), statut='ACTIF')
    except (Produit.DoesNotExist, ValueError):
        messages.error(request, "Produit introuvable.")
        return redirect('etat_carte_stock', marketeur_pk=marketeur_pk)

    if regime == REGIME_TOUS:
        carte = _calculer_carte_stock_tous(mkt, produit, date_debut, date_fin)
    else:
        carte = _calculer_carte_stock(mkt, produit, regime, date_debut, date_fin)
    societe = Societe.get_instance()

    try:
        contenu = _generer_xlsx_carte(mkt, produit, regime, carte, societe, date_debut, date_fin)
    except ImportError as e:
        messages.error(request, str(e))
        return redirect('etat_carte_stock', marketeur_pk=marketeur_pk)

    nom_fichier = (
        f"carte_stock_{mkt.sigle or 'mkt'}_{produit.nom.replace(' ', '_')}"
        f"_{'SD' if regime == REGIME_SD else 'AC'}"
        f"{'_' + str(date_debut) if date_debut else ''}.xlsx"
    )
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response
