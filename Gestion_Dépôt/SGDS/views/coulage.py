"""
Vues du module Répartition du Coulage par Marketeur (COULA REPAR).
"""
import io
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, TemplateView


# ─────────────────────────────────────────────────────────────
#  LISTE DES PÉRIODES
# ─────────────────────────────────────────────────────────────
class ListePeriodesCoulageView(LoginRequiredMixin, ListView):
    template_name = 'coulage/liste_periodes.html'
    paginate_by   = 24

    def get_queryset(self):
        from SGDS.models import PeriodeComptable
        return (
            PeriodeComptable.objects
            .select_related('cloture_coulage')
            .order_by('-annee', '-mois')
        )

    def get_context_data(self, **kwargs):
        from SGDS.models import Produit
        ctx = super().get_context_data(**kwargs)
        ctx['produits_actifs_list'] = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
        return ctx


# ─────────────────────────────────────────────────────────────
#  DÉTAIL / RAPPORT D'UNE PÉRIODE
# ─────────────────────────────────────────────────────────────
class RepartitionCoulageView(LoginRequiredMixin, TemplateView):
    template_name = 'coulage/repartition.html'

    def get_context_data(self, **kwargs):
        from SGDS.models import PeriodeComptable
        from SGDS.services.coulage_repartition import calculer_repartition_coulage

        ctx  = super().get_context_data(**kwargs)
        periode = get_object_or_404(PeriodeComptable, pk=self.kwargs['periode_id'])

        # Utiliser le snapshot si période clôturée et snapshot disponible
        if periode.statut == 'CLOTUREE' and hasattr(periode, 'cloture_coulage'):
            rapport = self._rapport_depuis_snapshot(periode.cloture_coulage)
            source  = 'SNAPSHOT'
        else:
            rapport = calculer_repartition_coulage(periode)
            source  = 'TEMPS_REEL'

        ctx['periode']       = periode
        ctx['rapport']       = rapport
        ctx['source']        = source
        ctx['peut_cloturer'] = (periode.est_ouverte and self.request.user.is_staff)
        return ctx

    def _rapport_depuis_snapshot(self, cloture):
        """Reconstruit un dict rapport identique à calculer_repartition_coulage() depuis la DB."""
        from decimal import Decimal

        # Produits du snapshot (ordonnés par nom)
        produits_qs = list(
            cloture.produits_coulage
            .select_related('produit')
            .order_by('produit__nom')
        )
        produits = [pc.produit for pc in produits_qs]

        coefficients = {pc.produit_id: pc.coefficient for pc in produits_qs}
        pertes_gains  = {pc.produit_id: pc.pertes_gains for pc in produits_qs}
        cumuls = {
            pc.produit_id: {
                'brut_entree': pc.cumul_entree,
                'coul_entree': Decimal('0'),
                'entree':      pc.cumul_entree,
                'sortie':      pc.cumul_sortie,
            }
            for pc in produits_qs
        }

        # Lignes marketeur × produit
        lignes_db = list(
            cloture.lignes
            .select_related('marketeur', 'produit')
            .order_by('marketeur__raison_sociale', 'produit__nom')
        )

        # Regrouper par marketeur
        from collections import defaultdict
        par_mkt = defaultdict(dict)
        vol_mkt = defaultdict(Decimal)
        mont_mkt = defaultdict(Decimal)
        mkt_obj = {}
        pu_mkt = {}
        motif_mkt = {}

        for l in lignes_db:
            mkt_id = l.marketeur_id
            mkt_obj[mkt_id] = l.marketeur
            pu_mkt[mkt_id]  = l.prix_unitaire
            motif_mkt[mkt_id] = l.motif
            par_mkt[mkt_id][l.produit_id] = {
                'brut_entree':  l.brut_entree,
                'coul_entree':  l.coul_entree,
                'entree_nette': l.entree_nette,
                'sortie':       l.sortie,
                'base_qp_coul': l.base_qp_coul,
                'coef_qp_coul': l.coef_qp_coul,
                'qp_coul':      l.qp_coul,
                'volume_sorti': l.volume_sorti,
            }
            vol_mkt[mkt_id]  += l.volume_sorti
            mont_mkt[mkt_id] += l.montant

        lignes = [
            {
                'marketeur':           mkt_obj[mkt_id],
                'par_produit':         par_mkt[mkt_id],
                'volume_global_sorti': vol_mkt[mkt_id],
                'motif':               motif_mkt[mkt_id],
                'prix_unitaire':       pu_mkt[mkt_id],
                'montant':             mont_mkt[mkt_id],
            }
            for mkt_id in mkt_obj
        ]

        # Totaux
        def _sum_pp(pk, field):
            return sum(
                (l['par_produit'].get(pk, {}).get(field, Decimal('0')) for l in lignes),
                Decimal('0')
            )

        totaux_par_produit = {
            p.pk: {
                'brut_entree':  _sum_pp(p.pk, 'brut_entree'),
                'coul_entree':  _sum_pp(p.pk, 'coul_entree'),
                'entree_nette': _sum_pp(p.pk, 'entree_nette'),
                'sortie':       _sum_pp(p.pk, 'sortie'),
                'base_qp_coul': _sum_pp(p.pk, 'base_qp_coul'),
                'coef_qp_coul': coefficients[p.pk],
                'qp_coul':      _sum_pp(p.pk, 'qp_coul'),
                'volume_sorti': _sum_pp(p.pk, 'volume_sorti'),
            }
            for p in produits
        }

        totaux = {
            'par_produit':         totaux_par_produit,
            'volume_global_sorti': sum((l['volume_global_sorti'] for l in lignes), Decimal('0')),
            'montant':             sum((l['montant'] for l in lignes), Decimal('0')),
            'motif':               cloture.motif,
            'prix_unitaire':       cloture.prix_unitaire_passage,
        }

        return {
            'periode':      cloture.periode,
            'produits':     produits,
            'coefficients': coefficients,
            'pertes_gains': pertes_gains,
            'cumuls':       cumuls,
            'lignes':       lignes,
            'totaux':       totaux,
            'parametres': {
                'prix_unitaire_passage': cloture.prix_unitaire_passage,
                'motif': cloture.motif,
            },
        }


# ─────────────────────────────────────────────────────────────
#  CLÔTURE D'UNE PÉRIODE
# ─────────────────────────────────────────────────────────────
class ClotureCoulageView(LoginRequiredMixin, UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_staff

    def post(self, request, periode_id):
        from SGDS.models import PeriodeComptable, JaugeageJour

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)

        if periode.statut != 'OUVERTE':
            messages.warning(request, f"La période {periode} n'est pas ouverte.")
            return redirect('coulage_detail', periode_id=periode_id)

        if not JaugeageJour.objects.filter(
            date_jaugeage__gte=periode.date_debut,
            date_jaugeage__lte=periode.date_fin,
        ).exists():
            messages.error(
                request,
                f"Impossible de clôturer : aucun jaugeage trouvé pour {periode}."
            )
            return redirect('coulage_detail', periode_id=periode_id)

        notes = request.POST.get('notes', '').strip() or None

        from SGDS.services.periode_comptable import cloturer_periode
        from django.core.exceptions import ValidationError

        try:
            cloture = cloturer_periode(periode, user=request.user, notes=notes)
        except ValidationError as e:
            messages.error(request, e.message)
            return redirect('coulage_detail', periode_id=periode_id)

        from SGDS.services.periode_comptable import mois_suivant as _ms
        m_s, a_s = _ms(periode.mois, periode.annee)
        messages.success(
            request,
            f"Période {periode} clôturée. "
            f"Montant coulage total : {int(cloture.total_montant):,} FCFA. "
            f"Pour commencer le mois suivant, ouvrez la période {m_s}/{a_s} "
            "depuis la liste des périodes."
        )
        return redirect('coulage_detail', periode_id=periode_id)


# ─────────────────────────────────────────────────────────────
#  EXPORT EXCEL
# ─────────────────────────────────────────────────────────────
class ExportCoulageExcelView(LoginRequiredMixin, View):
    def get(self, request, periode_id):
        from SGDS.models import PeriodeComptable
        from SGDS.services.coulage_repartition import calculer_repartition_coulage

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)

        if periode.statut == 'CLOTUREE' and hasattr(periode, 'cloture_coulage'):
            vue = RepartitionCoulageView()
            vue.kwargs = {'periode_id': periode_id}
            vue.request = request
            rapport = vue._rapport_depuis_snapshot(periode.cloture_coulage)
        else:
            rapport = calculer_repartition_coulage(periode)

        xlsx_bytes = _generer_xlsx_coulage(rapport)
        nom_fichier = f"COULA_REPAR_{periode.mois:02d}_{periode.annee}.xlsx"

        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
        return response


# ─────────────────────────────────────────────────────────────
#  GÉNÉRATION DU FICHIER XLSX
# ─────────────────────────────────────────────────────────────
def _generer_xlsx_coulage(rapport) -> bytes:
    """Reproduit fidèlement la feuille COULA REPAR du RJJ — générique par produit."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl est requis pour l'export Excel. pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COULA REPAR"

    periode  = rapport['periode']
    produits = rapport['produits']
    noms_mois = ['', 'JANVIER', 'FÉVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN',
                 'JUILLET', 'AOÛT', 'SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DÉCEMBRE']

    # ── Styles ───────────────────────────────────────────────
    police_base   = Font(name='Arial', size=10)
    police_titre  = Font(name='Arial', size=12, bold=True)
    police_entete = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    police_total  = Font(name='Arial', size=10, bold=True)

    remplissage_bleu  = PatternFill('solid', fgColor='2F5496')
    remplissage_gris  = PatternFill('solid', fgColor='D9E1F2')

    alignement_centre = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignement_gauche = Alignment(horizontal='left',   vertical='center')
    alignement_droite = Alignment(horizontal='right',  vertical='center')

    bordure_fine = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    fmt_vol  = '#,##0;(#,##0);-'
    fmt_coef = '0.00000000;(0.00000000);-'
    fmt_pu   = '0.0000'
    fmt_mont = '#,##0.00;(#,##0.00);-'

    def _ent(cell, val):
        cell.value = val; cell.font = police_entete
        cell.fill = remplissage_bleu; cell.alignment = alignement_centre
        cell.border = bordure_fine

    def _tot(cell, val=None, fmt=None):
        if val is not None: cell.value = val
        cell.font = police_total; cell.fill = remplissage_gris
        cell.border = bordure_fine; cell.alignment = alignement_droite
        if fmt: cell.number_format = fmt

    def _dat(cell, val=None, fmt=None, align=None):
        if val is not None: cell.value = val
        cell.font = police_base; cell.border = bordure_fine
        cell.alignment = align or alignement_droite
        if fmt: cell.number_format = fmt

    # ── Titre ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    nb_produits = len(produits)
    # 2 (code+mkt) + nb_produits*3 (entrée) + nb_produits (sortie) + nb_produits*3 (QP) + 1 vide + 4 = 2+7*nb_p+5
    last_col = 2 + nb_produits * 7 + 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    t = ws.cell(1, 1)
    t.value = (f"TABLEAU REPARTITION DES COULAGE INSTALLATIONS "
               f"{noms_mois[periode.mois]} {periode.annee}")
    t.font = police_titre; t.alignment = alignement_centre; t.border = bordure_fine

    # ── En-têtes ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 40

    # Colonnes fixes A-B
    for col, lab in [(1, 'CODE\nINTERNE\nCLIENT'), (2, 'LIBELLÉ MARKETEUR')]:
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        _ent(ws.cell(3, col), lab)

    col_idx = 3
    # Entrée par produit (3 colonnes chacun)
    for p in produits:
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+2)
        _ent(ws.cell(3, col_idx), f'ENTRÉE {p.nom.upper()}')
        for i, sl in enumerate([f'BRUT {p.code}', f'COUL {p.code}', f'NETTE {p.code}']):
            _ent(ws.cell(4, col_idx + i), sl)
        col_idx += 3

    # Sortie (1 colonne par produit)
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+nb_produits-1)
    _ent(ws.cell(3, col_idx), 'SORTIE')
    for i, p in enumerate(produits):
        _ent(ws.cell(4, col_idx + i), f'SORTIE {p.code}')
    col_idx += nb_produits

    # QP coulage (3 colonnes par produit)
    for p in produits:
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=3, end_column=col_idx+2)
        _ent(ws.cell(3, col_idx), f'RÉPRT. COUL. {p.nom.upper()}')
        for i, sl in enumerate(['BASE QP', 'COEF QP', 'QP COUL']):
            _ent(ws.cell(4, col_idx + i), sl)
        col_idx += 3

    # Colonne vide
    ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
    ws.cell(3, col_idx).border = bordure_fine
    col_vide = col_idx
    col_idx += 1

    # Colonnes finales
    for lab in ['VOLUME\nGLOBAL SORTI', 'MOTIF', 'PU/L', 'MONTANT']:
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        _ent(ws.cell(3, col_idx), lab)
        col_idx += 1

    col_vol_global = col_vide + 1
    col_motif      = col_vide + 2
    col_pu         = col_vide + 3
    col_montant    = col_vide + 4

    # ── Données ──────────────────────────────────────────────
    debut_data = 5
    for i, ligne in enumerate(rapport['lignes']):
        row = debut_data + i
        mkt = ligne['marketeur']
        _dat(ws.cell(row, 1), f"MARK{mkt.pk:03d}", align=alignement_centre)
        _dat(ws.cell(row, 2), mkt.sigle or mkt.raison_sociale[:30], align=alignement_gauche)

        c = 3
        for p in produits:
            pp = ligne['par_produit'].get(p.pk, {})
            _dat(ws.cell(row, c),   float(pp.get('brut_entree',  0)), fmt_vol); c+=1
            _dat(ws.cell(row, c),   float(pp.get('coul_entree',  0)), fmt_vol); c+=1
            _dat(ws.cell(row, c),   float(pp.get('entree_nette', 0)), fmt_vol); c+=1

        for p in produits:
            pp = ligne['par_produit'].get(p.pk, {})
            _dat(ws.cell(row, c),   float(pp.get('sortie', 0)), fmt_vol); c+=1

        for p in produits:
            pp = ligne['par_produit'].get(p.pk, {})
            _dat(ws.cell(row, c),   float(pp.get('base_qp_coul', 0)), fmt_vol); c+=1
            _dat(ws.cell(row, c),   float(pp.get('coef_qp_coul', 0)), fmt_coef); c+=1
            _dat(ws.cell(row, c),   float(pp.get('qp_coul',      0)), fmt_vol); c+=1

        ws.cell(row, col_vide).border = bordure_fine
        _dat(ws.cell(row, col_vol_global), float(ligne['volume_global_sorti']), fmt_vol)
        _dat(ws.cell(row, col_motif),      ligne['motif'], align=alignement_gauche)
        _dat(ws.cell(row, col_pu),         float(ligne['prix_unitaire']), fmt_pu)
        _dat(ws.cell(row, col_montant),    float(ligne['montant']), fmt_mont)

    # ── Totaux ───────────────────────────────────────────────
    row_tot = debut_data + len(rapport['lignes'])
    ws.merge_cells(start_row=row_tot, start_column=1, end_row=row_tot, end_column=2)
    c1 = ws.cell(row_tot, 1)
    c1.value = 'CUMULS'; c1.font = police_total; c1.fill = remplissage_gris
    c1.border = bordure_fine; c1.alignment = alignement_centre

    tot = rapport['totaux']
    c = 3
    for p in produits:
        pp = tot['par_produit'].get(p.pk, {})
        _tot(ws.cell(row_tot, c), float(pp.get('brut_entree',  0)), fmt_vol); c+=1
        _tot(ws.cell(row_tot, c), float(pp.get('coul_entree',  0)), fmt_vol); c+=1
        _tot(ws.cell(row_tot, c), float(pp.get('entree_nette', 0)), fmt_vol); c+=1

    for p in produits:
        pp = tot['par_produit'].get(p.pk, {})
        _tot(ws.cell(row_tot, c), float(pp.get('sortie', 0)), fmt_vol); c+=1

    for p in produits:
        pp = tot['par_produit'].get(p.pk, {})
        _tot(ws.cell(row_tot, c), float(pp.get('base_qp_coul', 0)), fmt_vol); c+=1
        _tot(ws.cell(row_tot, c), float(pp.get('coef_qp_coul', 0)), fmt_coef); c+=1
        _tot(ws.cell(row_tot, c), float(pp.get('qp_coul',      0)), fmt_vol); c+=1

    ws.cell(row_tot, col_vide).border = bordure_fine
    _tot(ws.cell(row_tot, col_vol_global), float(tot['volume_global_sorti']), fmt_vol)
    _tot(ws.cell(row_tot, col_motif),      tot['motif'])
    _tot(ws.cell(row_tot, col_pu),         float(tot['prix_unitaire']), fmt_pu)
    _tot(ws.cell(row_tot, col_montant),    float(tot['montant']), fmt_mont)

    # ── Largeurs ─────────────────────────────────────────────
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions[get_column_letter(col_vide)].width = 3
    ws.column_dimensions[get_column_letter(col_pu)].width   = 10
    ws.column_dimensions[get_column_letter(col_montant)].width = 16
    for ci in range(3, col_vide):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.column_dimensions[get_column_letter(col_vol_global)].width = 14
    ws.column_dimensions[get_column_letter(col_motif)].width      = 16

    for row in range(debut_data, row_tot + 1):
        ws.row_dimensions[row].height = 16

    ws.freeze_panes = ws.cell(debut_data, 3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
#  SUIVI ÉVOLUTION JOURNALIER
# ─────────────────────────────────────────────────────────────

class SuiviEvolutionView(LoginRequiredMixin, View):
    """Tableau journalier stock/mouvements pour un produit sur une période."""

    def get(self, request, periode_id, produit_id):
        from SGDS.models import PeriodeComptable, Produit
        from SGDS.services.suivi_evolution import calculer_suivi_evolution

        periode  = get_object_or_404(PeriodeComptable, pk=periode_id)
        produit  = get_object_or_404(Produit, pk=produit_id)
        produits = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
        rapport  = calculer_suivi_evolution(periode, produit)

        from django.shortcuts import render
        return render(request, 'coulage/suivi_evolution.html', {
            'periode':  periode,
            'produit':  produit,
            'produits': produits,
            'rapport':  rapport,
        })


class ExportSuiviExcelView(LoginRequiredMixin, View):
    """Export Excel du suivi évolution journalier."""

    def get(self, request, periode_id, produit_id):
        from SGDS.models import PeriodeComptable, Produit
        from SGDS.services.suivi_evolution import calculer_suivi_evolution
        from SGDS.services.export_excel import exporter_suivi_xlsx

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)
        produit = get_object_or_404(Produit, pk=produit_id)
        rapport = calculer_suivi_evolution(periode, produit)

        contenu = exporter_suivi_xlsx(rapport)
        nom = f"suivi_{produit.code}_{periode.libelle.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            contenu,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nom}"'
        return response


# ─────────────────────────────────────────────────────────────
#  FRAIS DE PASSAGE
# ─────────────────────────────────────────────────────────────

class FraisPassageView(LoginRequiredMixin, View):
    """Document de facturation mensuel — frais de passage par marketeur."""

    def get(self, request, periode_id):
        from SGDS.models import PeriodeComptable
        from SGDS.services.frais_passage import calculer_frais_passage

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)
        rapport = calculer_frais_passage(periode)

        from django.shortcuts import render
        return render(request, 'coulage/frais_passage.html', {
            'periode': periode,
            'rapport': rapport,
        })


class ExportFraisPassageExcelView(LoginRequiredMixin, View):
    """Export Excel des frais de passage."""

    def get(self, request, periode_id):
        from SGDS.models import PeriodeComptable
        from SGDS.services.frais_passage import calculer_frais_passage
        from SGDS.services.export_excel import exporter_frais_passage_xlsx

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)
        rapport = calculer_frais_passage(periode)

        contenu = exporter_frais_passage_xlsx(rapport)
        nom = f"frais_passage_{periode.libelle.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            contenu,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nom}"'
        return response
