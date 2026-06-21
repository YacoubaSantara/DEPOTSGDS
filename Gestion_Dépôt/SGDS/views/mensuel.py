"""
États mensuels :
  1. Stock Ouverture / Fermeture          → chef_depot_required
  2. Global Mensuel Dépôt                 → login_required (staff)
  3. Global Mensuel RJJ                   → login_required (staff)
  4. Coulage Répartition (admin)          → login_required (staff)
  4b. Coulage Répartition (marketeur)     → marketeur_required
  5. Stock Mensuel Format A (admin + mkt) → login_required / marketeur_required
  6. Stock Mensuel Format B (admin + mkt) → login_required / marketeur_required
  7. Frais de Passage (admin + mkt)       → login_required / marketeur_required
"""

from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError

from SGDS.users.decorators import chef_depot_required
from .client import marketeur_required, _D


_Z  = Decimal('0')
_Q2 = Decimal('0.01')


# ─────────────────────────────────────────────────────────────
#  HELPERS COMMUNS
# ─────────────────────────────────────────────────────────────

def _periodes_disponibles():
    from SGDS.models import PeriodeComptable
    return PeriodeComptable.objects.order_by('-annee', '-mois')


def _get_periode(request):
    """Période sélectionnée via GET ?periode_id= (UUID), sinon la plus récente."""
    from SGDS.models import PeriodeComptable
    pid = request.GET.get('periode_id')
    if pid:
        try:
            return PeriodeComptable.objects.get(uuid=pid)
        except (PeriodeComptable.DoesNotExist, ValueError, ValidationError):
            pass
    return PeriodeComptable.objects.order_by('-annee', '-mois').first()


def _get_marketeur_optionnel(request):
    """Retourne le Marketeur sélectionné via GET ?marketeur= (UUID), ou None."""
    from SGDS.models import Marketeur
    mkt_id = request.GET.get('marketeur')
    if mkt_id:
        try:
            return Marketeur.objects.get(uuid=mkt_id)
        except (Marketeur.DoesNotExist, ValueError, ValidationError):
            pass
    return None


# ═════════════════════════════════════════════════════════════
#  ÉTAT 1 — STOCK OUVERTURE / FERMETURE  (format fichier 06)
# ═════════════════════════════════════════════════════════════

def _calculer_stock_ouverture_fermeture(periode, date_fin_override=None, date_jaugeage_override=None):
    """
    Pour chaque produit actif, calcule tous les agrégats du tableau mensuel
    conformément au format Excel (fichier 06) :
      Stock Ouverture (SD + Acquittée) /
      Entrées détaillées (Réception CC + P/G Réception + Reclassement) /
      Sorties détaillées (CC Acquittée + Cession + Reclassement + Coulage) /
      Stock Comptable (SD + Acquittée) /
      P/G Installation / Ratio /
      Stock Clôture (SD + Acquittée + Total)
    Tous les anciens champs sont conservés pour la rétrocompatibilité.

    date_fin_override     : restreint les mouvements à cette date (vue journalière).
    date_jaugeage_override: date à utiliser pour chercher le jaugeage en vue journalière.
                            Peut différer de date_fin_override (ex : ouverture du jour J
                            utilise date_fin_override=J-1 pour les mouvements mais cherche
                            le jaugeage de J).
    """
    from SGDS.models import (Produit, StockOuverture, Mouvement,
                              InventaireInitialMarketeur, JaugeageJour, MesureCuve)

    date_fin_calc = date_fin_override if date_fin_override is not None else periode.date_fin
    filtre_journalier = date_fin_override is not None

    produits   = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
    mouvements = list(
        Mouvement.objects
        .filter(date_mouvement__range=(periode.date_debut, date_fin_calc))
        .select_related('produit', 'marketeur', 'cession_marketeur_destinataire')
    )

    # ── Pertes coulage (si clôture disponible, ignoré en vue journalière) ──
    coulage_pertes = {}
    if not filtre_journalier:
        cloture = getattr(periode, 'cloture_coulage', None)
        if cloture:
            for cp in cloture.produits_coulage.select_related('produit'):
                coulage_pertes[cp.produit_id] = _D(cp.pertes_gains)

    # ── Stock physique (jaugeage) ──
    sf_physique  = {}
    if not filtre_journalier:
        # Période entière : StockOuverture de la période suivante en priorité
        periode_suiv = periode.periode_suivante()
        if periode_suiv:
            for so in StockOuverture.objects.filter(periode=periode_suiv).select_related('produit'):
                pid = so.produit_id
                if pid not in sf_physique:
                    sf_physique[pid] = {'amb': _Z, '15c': _Z}
                sf_physique[pid]['amb'] += _D(so.volume_ambiant)
                sf_physique[pid]['15c'] += _D(so.volume_15c)

        # Fallback : dernier jaugeage validé de la période
        if not sf_physique:
            dernier_j = (
                JaugeageJour.objects
                .filter(
                    date_jaugeage__range=(periode.date_debut, periode.date_fin),
                    est_valide=True,
                )
                .order_by('-date_jaugeage', '-heure_jaugeage', '-date_creation')
                .first()
            )
            if dernier_j:
                for mesure in dernier_j.mesures.select_related(
                        'cuve__produit__famille', 'cuve__parametre_jaugeage').all():
                    cuve = mesure.cuve
                    if cuve.produit is None:
                        continue
                    pid   = cuve.produit_id
                    v_amb = mesure.volume_ambiant_depot
                    v_15c = mesure.volume_standard_15c_calcule
                    if v_amb is not None:
                        if pid not in sf_physique:
                            sf_physique[pid] = {'amb': _Z, '15c': _Z}
                        sf_physique[pid]['amb'] += _D(v_amb)
                    if v_15c is not None:
                        sf_physique.setdefault(pid, {'amb': _Z, '15c': _Z})
                        sf_physique[pid]['15c'] += _D(v_15c)
    else:
        # Vue journalière : chercher le jaugeage du jour sélectionné.
        # Ouverture (date_fin_override < date_jaugeage_override) → premier jaugeage du jour (matin).
        # Fermeture (date_fin_override == date_jaugeage_override) → dernier jaugeage du jour (soir).
        date_j = date_jaugeage_override or date_fin_override
        is_ouverture_mode = (date_fin_override is not None
                             and date_jaugeage_override is not None
                             and date_fin_override < date_jaugeage_override)
        heure_order = 'heure_jaugeage' if is_ouverture_mode else '-heure_jaugeage'

        dernier_j = (
            JaugeageJour.objects
            .filter(date_jaugeage=date_j, est_valide=True)
            .order_by(heure_order, '-date_creation')
            .first()
        )
        if not dernier_j and date_j >= periode.date_debut:
            dernier_j = (
                JaugeageJour.objects
                .filter(
                    date_jaugeage__range=(periode.date_debut, date_j),
                    est_valide=True,
                )
                .order_by('-date_jaugeage', heure_order, '-date_creation')
                .first()
            )
        if dernier_j:
            for mesure in dernier_j.mesures.select_related(
                    'cuve__produit__famille', 'cuve__parametre_jaugeage').all():
                cuve = mesure.cuve
                if cuve.produit is None:
                    continue
                pid   = cuve.produit_id
                v_amb = mesure.volume_ambiant_depot
                v_15c = mesure.volume_standard_15c_calcule
                if v_amb is not None:
                    if pid not in sf_physique:
                        sf_physique[pid] = {'amb': _Z, '15c': _Z}
                    sf_physique[pid]['amb'] += _D(v_amb)
                if v_15c is not None:
                    sf_physique.setdefault(pid, {'amb': _Z, '15c': _Z})
                    sf_physique[pid]['15c'] += _D(v_15c)

    # ── Fallback inventaires initiaux (1ère période sans StockOuverture) ──
    _inv_initiaux = {}
    if periode.periode_precedente() is None:
        for inv in InventaireInitialMarketeur.objects.filter(
            date_inventaire__lte=periode.date_fin,
        ).select_related('produit'):
            pid = inv.produit_id
            if pid not in _inv_initiaux:
                _inv_initiaux[pid] = {'amb': _Z, '15c': _Z}
            _inv_initiaux[pid]['amb'] += _D(inv.volume_ambiant)
            _inv_initiaux[pid]['15c'] += _D(inv.volume_15c)

    # ── Perte/Gain Installation par produit (ignoré en vue journalière) ──
    pg_install = {}
    if not filtre_journalier:
        try:
            from SGDS.models import PerteGainInstallation
            for pg in PerteGainInstallation.objects.filter(periode=periode).select_related('produit'):
                pid = pg.produit_id
                if pid not in pg_install:
                    pg_install[pid] = {'amb': _Z, '15c': _Z}
                pg_install[pid]['amb'] += _D(pg.volume_ambiant)
                pg_install[pid]['15c'] += _D(pg.volume_15c)
        except Exception:
            pass  # table pas encore migrée

    lignes = []
    for produit in produits:
        mvts_p = [m for m in mouvements if m.produit_id == produit.pk]

        # ── Stock Ouverture par régime ──
        so_qs  = list(StockOuverture.objects.filter(periode=periode, produit=produit))
        so_map = {so.regime_douanier: so for so in so_qs}
        if so_qs:
            so_sd      = so_map.get('SOUS_DOUANE')
            so_ac      = so_map.get('ACQUITTE')
            ouv_sd_amb = _D(so_sd.volume_ambiant) if so_sd else _Z
            ouv_sd_15c = _D(so_sd.volume_15c)     if so_sd else _Z
            ouv_ac_amb = _D(so_ac.volume_ambiant)  if so_ac else _Z
            ouv_ac_15c = _D(so_ac.volume_15c)      if so_ac else _Z
        else:
            ouv_sd_amb = ouv_sd_15c = _Z
            _inv       = _inv_initiaux.get(produit.pk, {})
            ouv_ac_amb = _inv.get('amb', _Z)
            ouv_ac_15c = _inv.get('15c', _Z)

        stock_debut_amb = ouv_sd_amb + ouv_ac_amb
        stock_debut_15c = ouv_sd_15c + ouv_ac_15c

        # ── Entrées détaillées ──
        rec_sd_amb         = rec_sd_15c         = _Z
        pg_rec_sd_amb      = pg_rec_sd_15c      = _Z
        rec_ac_amb         = rec_ac_15c         = _Z
        pg_rec_ac_amb      = pg_rec_ac_15c      = _Z
        recl_sd_entree_amb = recl_sd_entree_15c = _Z  # Acquittée→SD
        recl_ac_entree_amb = recl_ac_entree_15c = _Z  # SD→Acquittée

        for m in mvts_p:
            if m.type_mouvement == 'ENTREE':
                v_amb    = _D(m.volume_ambiant_recu)
                v_15c    = _D(m.volume_15c_recu)
                pg_r_amb = _D(m.perte_gain_reception) if m.perte_gain_reception is not None else _Z
                pg_r_15c = _D(m.perte_gain_15c)       if m.perte_gain_15c is not None else _Z
                if m.regime_douanier == 'SOUS_DOUANE':
                    rec_sd_amb += v_amb; rec_sd_15c += v_15c
                    # pg_rec_sd exclu de la vue dépôt global (affiché uniquement côté marketeur)
                else:
                    rec_ac_amb += v_amb; rec_ac_15c += v_15c
                    # pg_rec_ac exclu de la vue dépôt global (affiché uniquement côté marketeur)
            elif m.type_mouvement == 'RECLASSEMENT':
                v_amb = _D(m.volume_ambiant_recu) if m.volume_ambiant_recu else _Z
                v_15c = _D(m.volume_15c_recu)     if m.volume_15c_recu     else _Z
                # regime_douanier = régime SOURCE du produit reclassé
                if m.regime_douanier == 'SOUS_DOUANE':
                    recl_ac_entree_amb += v_amb   # entre en Acquittée
                    recl_ac_entree_15c += v_15c
                else:
                    recl_sd_entree_amb += v_amb   # entre en SD
                    recl_sd_entree_15c += v_15c

        # Cessions reçues (dépôt unique = 0, évite double comptage)
        cess_recues_amb = cess_recues_15c = _Z

        # Totaux entrées par régime
        total_entrees_sd_amb = rec_sd_amb + pg_rec_sd_amb + recl_sd_entree_amb
        total_entrees_sd_15c = rec_sd_15c + pg_rec_sd_15c + recl_sd_entree_15c
        total_entrees_ac_amb = rec_ac_amb + pg_rec_ac_amb + cess_recues_amb + recl_ac_entree_amb
        total_entrees_ac_15c = rec_ac_15c + pg_rec_ac_15c + cess_recues_15c + recl_ac_entree_15c
        total_entrees_amb    = total_entrees_sd_amb + total_entrees_ac_amb
        total_entrees_15c    = total_entrees_sd_15c + total_entrees_ac_15c

        # ── Sorties détaillées ──
        livr_ac_amb        = livr_ac_15c        = _Z
        livr_sd_amb        = livr_sd_15c        = _Z
        cess_emises_amb    = cess_emises_15c    = _Z
        recl_sd_sortie_amb = recl_sd_sortie_15c = _Z  # quitte SD (SD→Acquittée)
        recl_ac_sortie_amb = recl_ac_sortie_15c = _Z  # quitte Acquittée (Acquittée→SD)

        for m in mvts_p:
            if m.type_mouvement == 'SORTIE':
                v_amb = _D(m.volume_ambiant_sortie)
                v_15c = _D(m.volume_15c_sortie)
                if m.regime_douanier == 'ACQUITTE':
                    livr_ac_amb += v_amb; livr_ac_15c += v_15c
                else:
                    livr_sd_amb += v_amb; livr_sd_15c += v_15c
            elif m.type_mouvement == 'CESSION':
                pass  # intra-dépôt : émises et reçues se compensent, exclu du stock global
            elif m.type_mouvement == 'RECLASSEMENT':
                v_amb = _D(m.volume_ambiant_recu) if m.volume_ambiant_recu else _Z
                v_15c = _D(m.volume_15c_recu)     if m.volume_15c_recu     else _Z
                if m.regime_douanier == 'SOUS_DOUANE':
                    recl_sd_sortie_amb += v_amb   # quitte SD
                    recl_sd_sortie_15c += v_15c
                else:
                    recl_ac_sortie_amb += v_amb   # quitte Acquittée
                    recl_ac_sortie_15c += v_15c

        # Coulage installation = |perte| sur ce produit (depuis audit coulage)
        coul_pg          = coulage_pertes.get(produit.pk, _Z)
        coul_install_amb = abs(coul_pg) if coul_pg < _Z else _Z

        # Totaux sorties par régime
        total_sorties_sd_amb = recl_sd_sortie_amb
        total_sorties_sd_15c = recl_sd_sortie_15c
        total_sorties_ac_amb = livr_ac_amb + cess_emises_amb + recl_ac_sortie_amb + coul_install_amb
        total_sorties_ac_15c = livr_ac_15c + cess_emises_15c + recl_ac_sortie_15c
        total_sorties_amb    = total_sorties_sd_amb + total_sorties_ac_amb
        total_sorties_15c    = total_sorties_sd_15c + total_sorties_ac_15c

        # ── Stock Comptable par régime ──
        stk_c_sd_amb    = ouv_sd_amb + total_entrees_sd_amb - total_sorties_sd_amb
        stk_c_sd_15c    = ouv_sd_15c + total_entrees_sd_15c - total_sorties_sd_15c
        stk_c_ac_amb    = ouv_ac_amb + total_entrees_ac_amb - total_sorties_ac_amb
        stk_c_ac_15c    = ouv_ac_15c + total_entrees_ac_15c - total_sorties_ac_15c
        stock_fin_c_amb = stk_c_sd_amb + stk_c_ac_amb
        stock_fin_c_15c = stk_c_sd_15c + stk_c_ac_15c

        # ── P/G Installation (Acquittée uniquement) ──
        pg_inst     = pg_install.get(produit.pk, {})
        pg_inst_amb = pg_inst.get('amb', _Z)
        pg_inst_15c = pg_inst.get('15c', _Z)

        # ── Ratio = P/G Installation / Total Sorties Acquittées ──
        ratio = None
        if total_sorties_ac_amb and total_sorties_ac_amb != _Z:
            try:
                ratio = round(float(pg_inst_amb) / float(total_sorties_ac_amb), 6)
            except (ZeroDivisionError, Exception):
                pass

        # ── Stock physique (jaugeage) + P/G global ──
        sf     = sf_physique.get(produit.pk, {})
        sf_amb = sf.get('amb')
        sf_15c = sf.get('15c')
        pg_amb = (sf_amb - stock_fin_c_amb) if sf_amb is not None else None

        # ── Stock Clôture = Stock Comptable + P/G Installation (Acquittée) ──
        cloture_sd_amb    = stk_c_sd_amb
        cloture_sd_15c    = stk_c_sd_15c
        cloture_ac_amb    = stk_c_ac_amb + pg_inst_amb
        cloture_ac_15c    = stk_c_ac_15c + pg_inst_15c
        cloture_total_amb = cloture_sd_amb + cloture_ac_amb
        cloture_total_15c = cloture_sd_15c + cloture_ac_15c

        lignes.append({
            'produit':                  produit,
            # Ouverture par régime
            'ouv_sd_amb':               ouv_sd_amb,
            'ouv_sd_15c':               ouv_sd_15c,
            'ouv_ac_amb':               ouv_ac_amb,
            'ouv_ac_15c':               ouv_ac_15c,
            'stock_debut_amb':          stock_debut_amb,
            'stock_debut_15c':          stock_debut_15c,
            # Entrées détaillées
            'rec_sd_amb':               rec_sd_amb,
            'rec_sd_15c':               rec_sd_15c,
            'pg_rec_sd_amb':            pg_rec_sd_amb,
            'pg_rec_sd_15c':            pg_rec_sd_15c,
            'rec_ac_amb':               rec_ac_amb,
            'rec_ac_15c':               rec_ac_15c,
            'pg_rec_ac_amb':            pg_rec_ac_amb,
            'pg_rec_ac_15c':            pg_rec_ac_15c,
            'cess_recues_amb':          cess_recues_amb,
            'cess_recues_15c':          cess_recues_15c,
            'recl_sd_entree_amb':       recl_sd_entree_amb,
            'recl_sd_entree_15c':       recl_sd_entree_15c,
            'recl_ac_entree_amb':       recl_ac_entree_amb,
            'recl_ac_entree_15c':       recl_ac_entree_15c,
            'total_entrees_sd_amb':     total_entrees_sd_amb,
            'total_entrees_sd_15c':     total_entrees_sd_15c,
            'total_entrees_ac_amb':     total_entrees_ac_amb,
            'total_entrees_ac_15c':     total_entrees_ac_15c,
            'total_entrees_amb':        total_entrees_amb,
            'total_entrees_15c':        total_entrees_15c,
            # Sorties détaillées
            'livr_ac_amb':              livr_ac_amb,
            'livr_ac_15c':              livr_ac_15c,
            'livr_sd_amb':              livr_sd_amb,
            'livr_sd_15c':              livr_sd_15c,
            'cess_emises_amb':          cess_emises_amb,
            'cess_emises_15c':          cess_emises_15c,
            'coul_install_amb':         coul_install_amb,
            'recl_sd_sortie_amb':       recl_sd_sortie_amb,
            'recl_sd_sortie_15c':       recl_sd_sortie_15c,
            'recl_ac_sortie_amb':       recl_ac_sortie_amb,
            'recl_ac_sortie_15c':       recl_ac_sortie_15c,
            'total_sorties_sd_amb':     total_sorties_sd_amb,
            'total_sorties_sd_15c':     total_sorties_sd_15c,
            'total_sorties_ac_amb':     total_sorties_ac_amb,
            'total_sorties_ac_15c':     total_sorties_ac_15c,
            'total_sorties_amb':        total_sorties_amb,
            'total_sorties_15c':        total_sorties_15c,
            # Stock Comptable par régime
            'stk_c_sd_amb':             stk_c_sd_amb,
            'stk_c_sd_15c':             stk_c_sd_15c,
            'stk_c_ac_amb':             stk_c_ac_amb,
            'stk_c_ac_15c':             stk_c_ac_15c,
            'stock_fin_comptable_amb':  stock_fin_c_amb,
            'stock_fin_comptable_15c':  stock_fin_c_15c,
            # P/G Installation & Ratio
            'pg_inst_amb':              pg_inst_amb,
            'pg_inst_15c':              pg_inst_15c,
            'ratio':                    ratio,
            # Stock Clôture
            'cloture_sd_amb':           cloture_sd_amb,
            'cloture_sd_15c':           cloture_sd_15c,
            'cloture_ac_amb':           cloture_ac_amb,
            'cloture_ac_15c':           cloture_ac_15c,
            'cloture_total_amb':        cloture_total_amb,
            'cloture_total_15c':        cloture_total_15c,
            # Stock physique (jaugeage) + P/G global
            'stock_fin_physique_amb':   sf_amb,
            'stock_fin_physique_15c':   sf_15c,
            'perte_gain_amb':           pg_amb,
        })

    return lignes


@chef_depot_required
def etat_stock_ouverture_fermeture(request):
    from SGDS.models import Societe, Produit
    from datetime import datetime, timedelta
    periodes       = _periodes_disponibles()
    periode        = _get_periode(request)
    produits       = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
    produit_id     = request.GET.get('produit_id')
    date_str       = request.GET.get('date', '').strip()
    produit_filtre = None
    date_filtre    = None
    date_fin_override = None

    if produit_id:
        try:
            produit_filtre = Produit.objects.get(uuid=produit_id, statut='ACTIF')
        except (Produit.DoesNotExist, ValidationError, ValueError):
            pass

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                # Ouverture du jour J = mouvements jusqu'à J-1 (exclu J)
                date_fin_override = d - timedelta(days=1)
        except ValueError:
            pass

    lignes_all = _calculer_stock_ouverture_fermeture(periode, date_fin_override, date_filtre) if periode else []
    lignes = [l for l in lignes_all
              if produit_filtre is None or l['produit'].pk == produit_filtre.pk]
    societe = Societe.get_instance()
    return render(request, 'Etat/mensuel/stock_ouverture.html', {
        'periodes':       periodes,
        'periode':        periode,
        'lignes':         lignes,
        'societe':        societe,
        'produits':       produits,
        'produit_filtre': produit_filtre,
        'date_filtre':    date_filtre,
    })


@chef_depot_required
def etat_stock_fermeture(request):
    from SGDS.models import Societe, Produit
    from datetime import datetime
    periodes       = _periodes_disponibles()
    periode        = _get_periode(request)
    produits       = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
    produit_id     = request.GET.get('produit_id')
    date_str       = request.GET.get('date', '').strip()
    produit_filtre = None
    date_filtre    = None
    date_fin_override = None

    if produit_id:
        try:
            produit_filtre = Produit.objects.get(uuid=produit_id, statut='ACTIF')
        except (Produit.DoesNotExist, ValidationError, ValueError):
            pass

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                # Fermeture du jour J = mouvements jusqu'à J inclus
                date_fin_override = d
        except ValueError:
            pass

    lignes_all = _calculer_stock_ouverture_fermeture(periode, date_fin_override, date_filtre) if periode else []
    lignes = [l for l in lignes_all
              if produit_filtre is None or l['produit'].pk == produit_filtre.pk]
    societe = Societe.get_instance()
    return render(request, 'Etat/mensuel/stock_fermeture.html', {
        'periodes':       periodes,
        'periode':        periode,
        'lignes':         lignes,
        'societe':        societe,
        'produits':       produits,
        'produit_filtre': produit_filtre,
        'date_filtre':    date_filtre,
    })


@chef_depot_required
def etat_stock_fermeture_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    periode    = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune periode disponible.")
        return redirect('etat_mensuel_stock_fermeture')
    lignes_all = _calculer_stock_ouverture_fermeture(periode)
    lignes = [
        {
            'produit':           l['produit'],
            'stk_comptable_amb': l['stock_fin_comptable_amb'],
            'stk_comptable_15c': l['stock_fin_comptable_15c'],
            'stk_physique_amb':  l['stock_fin_physique_amb'],
            'stk_physique_15c':  l['stock_fin_physique_15c'],
            'perte_gain_amb':    l['perte_gain_amb'],
            'total_entrees_amb': l['total_entrees_amb'],
        }
        for l in lignes_all
    ]
    societe = Societe.get_instance()
    contenu = _xlsx_stock_fermeture(periode, lignes, societe)
    nom = f"stock_fermeture_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_stock_fermeture(periode, lignes, societe):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Stock Fermeture"
    navy  = PatternFill("solid", fgColor="1B3A6B")
    green = PatternFill("solid", fgColor="16A34A")
    red   = PatternFill("solid", fgColor="DC2626")
    bold  = Font(bold=True)
    white = Font(bold=True, color="FFFFFF")
    row = 1
    if societe:
        ws.cell(row, 1, societe.raison_sociale).font = bold; row += 1
    ws.cell(row, 1, f"Stock Fermeture — {periode}").font = bold; row += 2
    headers = ["Produit", "Stk Comptable AMB (L)", "Stk Comptable 15°C (L)",
               "Stk Physique AMB (L)", "Stk Physique 15°C (L)", "P/G AMB (L)"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = navy; cell.font = white; cell.alignment = Alignment(horizontal='center')
    row += 1
    for l in lignes:
        ws.cell(row, 1, l['produit'].nom)
        ws.cell(row, 2, float(l['stk_comptable_amb'] or 0))
        ws.cell(row, 3, float(l['stk_comptable_15c'] or 0))
        ws.cell(row, 4, float(l['stk_physique_amb']) if l['stk_physique_amb'] is not None else "")
        ws.cell(row, 5, float(l['stk_physique_15c']) if l['stk_physique_15c'] is not None else "")
        pg = l['perte_gain_amb']
        cell = ws.cell(row, 6, float(pg) if pg is not None else "")
        if pg is not None:
            cell.fill = green if pg >= 0 else red
            cell.font = Font(color="FFFFFF")
        row += 1
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


@chef_depot_required
def etat_stock_ouverture_fermeture_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    periode = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_stock_ouverture')
    lignes  = _calculer_stock_ouverture_fermeture(periode)
    societe = Societe.get_instance()
    contenu = _xlsx_stock_ouverture(periode, lignes, societe)
    nom = f"stock_ouverture_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_stock_ouverture(periode, lignes, societe):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    BLEU_ENT  = PatternFill(fill_type='solid', fgColor='1F4E79')
    BLEU_LG   = PatternFill(fill_type='solid', fgColor='DEEAF1')
    VERT_LG   = PatternFill(fill_type='solid', fgColor='E2EFDA')
    ROUGE_LG  = PatternFill(fill_type='solid', fgColor='FCE4D6')
    JAUNE_LG  = PatternFill(fill_type='solid', fgColor='FFF2CC')
    TITRE_LG  = PatternFill(fill_type='solid', fgColor='BDD7EE')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Stock {periode.mois:02d}-{periode.annee}"

    ws.merge_cells('A1:C1')
    ws['A1'] = societe.raison_sociale if societe else 'SGDS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = f"ÉTAT STOCK OUVERTURE / FERMETURE — {periode}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    for ci, (h, w) in enumerate(
        [('DÉSIGNATION', 35), ('V AMB (L)', 15), ('V @15°C (L)', 15)], 1
    ):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = BLEU_ENT
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    row += 1

    def _r(label, v_amb, v_15c, fill=None, bold=False):
        nonlocal row
        for ci, val in enumerate([label, v_amb, v_15c], 1):
            c = ws.cell(row=row, column=ci,
                        value=float(val) if isinstance(val, Decimal) else val)
            if fill:
                c.fill = fill
            if bold:
                c.font = Font(bold=True)
            if ci > 1 and isinstance(val, Decimal):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
        row += 1

    for ligne in lignes:
        ws.merge_cells(f'A{row}:C{row}')
        t = ws.cell(row=row, column=1, value=f"═══ {ligne['produit'].nom.upper()} ═══")
        t.font = Font(bold=True, size=11, color='1F4E79')
        t.fill = TITRE_LG
        row += 1

        _r('Stock de départ',        ligne['stock_debut_amb'],         ligne['stock_debut_15c'],         BLEU_LG, True)
        _r('  Réceptions SD CC',     ligne['rec_sd_amb'],              ligne['rec_sd_15c'])
        _r('  Réceptions AC CC',     ligne['rec_ac_amb'],              ligne['rec_ac_15c'])
        _r('  Cessions reçues',      ligne['cess_recues_amb'],         ligne['cess_recues_15c'])
        _r('▶ Total Entrées',        ligne['total_entrees_amb'],       ligne['total_entrees_15c'],        VERT_LG, True)
        _r('  Livraisons AC',        ligne['livr_ac_amb'],             ligne['livr_ac_15c'])
        _r('  Livraisons SD',        ligne['livr_sd_amb'],             ligne['livr_sd_15c'])
        _r('  Cessions émises',      ligne['cess_emises_amb'],         ligne['cess_emises_15c'])
        _r('  Coulage installation', ligne['coul_install_amb'],        _Z)
        _r('▶ Total Sorties',        ligne['total_sorties_amb'],       ligne['total_sorties_15c'],        ROUGE_LG, True)
        _r('▶ Stock comptable (fin)',ligne['stock_fin_comptable_amb'], ligne['stock_fin_comptable_15c'],  JAUNE_LG, True)

        if ligne['stock_fin_physique_amb'] is not None:
            _r('▶ Stock physique (jaugeage)', ligne['stock_fin_physique_amb'], ligne['stock_fin_physique_15c'], JAUNE_LG, True)
            pg = ligne['perte_gain_amb']
            _r('  Perte / Gain', pg, _Z,
               ROUGE_LG if (pg is not None and pg < _Z) else VERT_LG, True)
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════
#  ÉTAT 2 — GLOBAL MENSUEL DÉPÔT
# ═════════════════════════════════════════════════════════════

def _calculer_global_depot(periode, marketeur=None):
    from SGDS.models import Mouvement
    from collections import defaultdict

    qs = (
        Mouvement.objects
        .filter(date_mouvement__range=(periode.date_debut, periode.date_fin))
        .select_related('produit', 'marketeur', 'camion', 'cession_marketeur_destinataire')
        .order_by('type_mouvement', 'date_mouvement', 'pk')
    )
    if marketeur:
        qs = qs.filter(marketeur=marketeur)

    mouvements = list(qs)
    par_type   = defaultdict(list)
    for m in mouvements:
        par_type[m.type_mouvement].append(m)

    # Totaux globaux par produit
    totaux = defaultdict(lambda: {
        'entrees_amb': _Z, 'entrees_15c': _Z,
        'sorties_amb': _Z, 'sorties_15c': _Z,
        'cessions_amb': _Z, 'cessions_15c': _Z,
    })
    for m in mouvements:
        t = totaux[m.produit.nom]
        if m.type_mouvement == 'ENTREE':
            t['entrees_amb'] += _D(m.volume_ambiant_recu)
            t['entrees_15c'] += _D(m.volume_15c_recu)
        elif m.type_mouvement == 'SORTIE':
            t['sorties_amb'] += _D(m.volume_ambiant_sortie)
            t['sorties_15c'] += _D(m.volume_15c_sortie)
        elif m.type_mouvement == 'CESSION':
            t['cessions_amb'] += _D(m.cession_volume_ambiant)
            t['cessions_15c'] += _D(m.cession_volume_15c)

    return {
        'periode':    periode,
        'mouvements': mouvements,
        'par_type':   dict(par_type),
        'totaux':     dict(totaux),
    }


# ─────────────────────────────────────────────────────────────
#  STOCK OUVERTURE / FERMETURE — ESPACE MARKETEUR
# ─────────────────────────────────────────────────────────────

def _calculer_stock_ouverture_fermeture_marketeur(periode, marketeur, date_fin_override=None):
    """
    Version filtrée par marketeur de _calculer_stock_ouverture_fermeture.
    Utilise InventaireInitialMarketeur pour le stock de départ et filtre
    les mouvements par marketeur.

    date_fin_override : si fourni, restreint les mouvements à cette date au lieu
    de periode.date_fin (état journalier).
    """
    from SGDS.models import Produit, InventaireInitialMarketeur, Mouvement, StockOuvertureMarketeur, ClotureCoulageLigne

    date_fin_calc     = date_fin_override if date_fin_override is not None else periode.date_fin
    filtre_journalier = date_fin_override is not None

    # ── Quote-part du marketeur dans le P/G Installation du dépôt ──
    # (même logique que la répartition du coulage / frais de passage,
    # ignorée en vue journalière car calculée sur la période entière)
    # Période déjà clôturée → lire la quote-part figée (ne bouge plus jamais).
    # Période encore ouverte → estimation live (provisoire jusqu'à la clôture).
    qp_coul_par_produit = {}
    if not filtre_journalier:
        if periode.statut == 'CLOTUREE':
            for ligne in ClotureCoulageLigne.objects.filter(
                cloture__periode=periode, marketeur=marketeur,
            ).select_related('produit'):
                if ligne.produit_id:
                    qp_coul_par_produit[ligne.produit_id] = {
                        'qp_coul': ligne.qp_coul, 'coef_qp_coul': ligne.coef_qp_coul,
                    }
        else:
            try:
                from SGDS.services.coulage_repartition import calculer_repartition_coulage
                rapport_coul = calculer_repartition_coulage(periode, marketeurs=[marketeur])
                if rapport_coul['lignes']:
                    qp_coul_par_produit = rapport_coul['lignes'][0]['par_produit']
            except Exception:
                pass

    produits   = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
    mouvements = list(
        Mouvement.objects
        .filter(
            date_mouvement__range=(periode.date_debut, date_fin_calc),
            marketeur=marketeur,
        )
        .select_related('produit', 'cession_marketeur_destinataire')
    )

    # Cessions du marketeur sur la période (reçues + émises), pour répartition SD/AC
    cessions_recues = list(
        Mouvement.objects.filter(
            cession_marketeur_destinataire=marketeur,
            type_mouvement='CESSION',
            date_mouvement__range=(periode.date_debut, date_fin_calc),
        ).select_related('produit')
    )

    # Stock d'ouverture de ce marketeur pour la période — reporté depuis la
    # fermeture du mois précédent (cf. services/stock_ouverture_marketeur.py).
    # Repli sur InventaireInitialMarketeur seulement si rien n'a encore été
    # résolu (1ère période du système, avant toute clôture).
    report_par_produit = {}
    for som in StockOuvertureMarketeur.objects.filter(periode=periode, marketeur=marketeur).select_related('produit'):
        pid = som.produit_id
        bucket = report_par_produit.setdefault(pid, {'SOUS_DOUANE': {'amb': _Z, '15c': _Z}, 'ACQUITTE': {'amb': _Z, '15c': _Z}})
        bucket[som.regime_douanier] = {'amb': _D(som.volume_ambiant), '15c': _D(som.volume_15c)}

    if not report_par_produit and periode.periode_precedente() is None:
        for inv in InventaireInitialMarketeur.objects.filter(
            marketeur=marketeur,
            date_inventaire__lte=periode.date_fin,
        ).select_related('produit'):
            pid = inv.produit_id
            bucket = report_par_produit.setdefault(pid, {'SOUS_DOUANE': {'amb': _Z, '15c': _Z}, 'ACQUITTE': {'amb': _Z, '15c': _Z}})
            bucket[inv.regime_douanier]['amb'] += _D(inv.volume_ambiant)
            bucket[inv.regime_douanier]['15c'] += _D(inv.volume_15c)

    inv_mkt = report_par_produit

    lignes = []
    for produit in produits:
        mvts_p = [m for m in mouvements if m.produit_id == produit.pk]
        if not mvts_p and produit.pk not in report_par_produit and produit.pk not in inv_mkt:
            continue  # pas d'activité pour ce marketeur sur ce produit

        # ── Stock ouverture par régime ──
        rp = report_par_produit.get(produit.pk, inv_mkt.get(
            produit.pk, {'SOUS_DOUANE': {'amb': _Z, '15c': _Z}, 'ACQUITTE': {'amb': _Z, '15c': _Z}}
        ))
        ouv_sd_amb = rp['SOUS_DOUANE']['amb']; ouv_sd_15c = rp['SOUS_DOUANE']['15c']
        ouv_ac_amb = rp['ACQUITTE']['amb'];    ouv_ac_15c = rp['ACQUITTE']['15c']
        stock_debut_amb = ouv_sd_amb + ouv_ac_amb
        stock_debut_15c = ouv_sd_15c + ouv_ac_15c

        # ── Entrées détaillées ──
        rec_sd_amb = rec_sd_15c = rec_ac_amb = rec_ac_15c = _Z
        recl_sd_entree_amb = recl_sd_entree_15c = _Z  # Acquittée→SD
        recl_ac_entree_amb = recl_ac_entree_15c = _Z  # SD→Acquittée
        for m in mvts_p:
            if m.type_mouvement == 'ENTREE':
                v_amb = _D(m.volume_ambiant_recu)
                v_15c = _D(m.volume_15c_recu)
                if m.regime_douanier == 'SOUS_DOUANE':
                    rec_sd_amb += v_amb; rec_sd_15c += v_15c
                else:
                    rec_ac_amb += v_amb; rec_ac_15c += v_15c
            elif m.type_mouvement == 'RECLASSEMENT':
                v_amb = _D(m.volume_ambiant_recu) if m.volume_ambiant_recu else _Z
                v_15c = _D(m.volume_15c_recu)     if m.volume_15c_recu     else _Z
                if m.regime_douanier == 'SOUS_DOUANE':
                    recl_ac_entree_amb += v_amb   # entre en Acquittée
                    recl_ac_entree_15c += v_15c
                else:
                    recl_sd_entree_amb += v_amb   # entre en SD
                    recl_sd_entree_15c += v_15c

        cess_recues_sd_amb = cess_recues_sd_15c = _Z
        cess_recues_ac_amb = cess_recues_ac_15c = _Z
        for m in cessions_recues:
            if m.produit_id != produit.pk:
                continue
            if m.regime_douanier == 'SOUS_DOUANE':
                cess_recues_sd_amb += _D(m.cession_volume_ambiant)
                cess_recues_sd_15c += _D(m.cession_volume_15c)
            else:
                cess_recues_ac_amb += _D(m.cession_volume_ambiant)
                cess_recues_ac_15c += _D(m.cession_volume_15c)
        cess_recues_amb = cess_recues_sd_amb + cess_recues_ac_amb
        cess_recues_15c = cess_recues_sd_15c + cess_recues_ac_15c

        total_entrees_sd_amb = rec_sd_amb + cess_recues_sd_amb + recl_sd_entree_amb
        total_entrees_sd_15c = rec_sd_15c + cess_recues_sd_15c + recl_sd_entree_15c
        total_entrees_ac_amb = rec_ac_amb + cess_recues_ac_amb + recl_ac_entree_amb
        total_entrees_ac_15c = rec_ac_15c + cess_recues_ac_15c + recl_ac_entree_15c
        total_entrees_amb    = total_entrees_sd_amb + total_entrees_ac_amb
        total_entrees_15c    = total_entrees_sd_15c + total_entrees_ac_15c

        # ── Sorties détaillées ──
        livr_ac_amb = livr_ac_15c = livr_sd_amb = livr_sd_15c = _Z
        recl_sd_sortie_amb = recl_sd_sortie_15c = _Z  # quitte SD (SD→Acquittée)
        recl_ac_sortie_amb = recl_ac_sortie_15c = _Z  # quitte Acquittée (Acquittée→SD)
        cess_emises_sd_amb = cess_emises_sd_15c = _Z
        cess_emises_ac_amb = cess_emises_ac_15c = _Z
        for m in mvts_p:
            if m.type_mouvement == 'SORTIE':
                v_amb = _D(m.volume_ambiant_sortie)
                v_15c = _D(m.volume_15c_sortie)
                if m.regime_douanier == 'ACQUITTE':
                    livr_ac_amb += v_amb; livr_ac_15c += v_15c
                else:
                    livr_sd_amb += v_amb; livr_sd_15c += v_15c
            elif m.type_mouvement == 'CESSION':
                if m.regime_douanier == 'SOUS_DOUANE':
                    cess_emises_sd_amb += _D(m.cession_volume_ambiant)
                    cess_emises_sd_15c += _D(m.cession_volume_15c)
                else:
                    cess_emises_ac_amb += _D(m.cession_volume_ambiant)
                    cess_emises_ac_15c += _D(m.cession_volume_15c)
            elif m.type_mouvement == 'RECLASSEMENT':
                v_amb = _D(m.volume_ambiant_recu) if m.volume_ambiant_recu else _Z
                v_15c = _D(m.volume_15c_recu)     if m.volume_15c_recu     else _Z
                if m.regime_douanier == 'SOUS_DOUANE':
                    recl_sd_sortie_amb += v_amb   # quitte SD
                    recl_sd_sortie_15c += v_15c
                else:
                    recl_ac_sortie_amb += v_amb   # quitte Acquittée
                    recl_ac_sortie_15c += v_15c

        cess_emises_amb = cess_emises_sd_amb + cess_emises_ac_amb
        cess_emises_15c = cess_emises_sd_15c + cess_emises_ac_15c

        total_sorties_sd_amb = livr_sd_amb + cess_emises_sd_amb + recl_sd_sortie_amb
        total_sorties_sd_15c = livr_sd_15c + cess_emises_sd_15c + recl_sd_sortie_15c
        total_sorties_ac_amb = livr_ac_amb + cess_emises_ac_amb + recl_ac_sortie_amb
        total_sorties_ac_15c = livr_ac_15c + cess_emises_ac_15c + recl_ac_sortie_15c
        total_sorties_amb    = total_sorties_sd_amb + total_sorties_ac_amb
        total_sorties_15c    = total_sorties_sd_15c + total_sorties_ac_15c

        # ── Stock Comptable par régime ──
        stk_c_sd_amb    = ouv_sd_amb + total_entrees_sd_amb - total_sorties_sd_amb
        stk_c_sd_15c    = ouv_sd_15c + total_entrees_sd_15c - total_sorties_sd_15c
        stk_c_ac_amb    = ouv_ac_amb + total_entrees_ac_amb - total_sorties_ac_amb
        stk_c_ac_15c    = ouv_ac_15c + total_entrees_ac_15c - total_sorties_ac_15c
        stock_fin_c_amb = stk_c_sd_amb + stk_c_ac_amb
        stock_fin_c_15c = stk_c_sd_15c + stk_c_ac_15c

        # ── Quote-part P/G Installation (Acquittée uniquement) ──
        qp        = qp_coul_par_produit.get(produit.pk, {})
        pg_inst_amb = qp.get('qp_coul', _Z)
        coef_qp     = qp.get('coef_qp_coul')
        ratio       = float(coef_qp) if coef_qp is not None else None

        # ── Stock Clôture = Stock Comptable + quote-part P/G Installation ──
        cloture_sd_amb    = stk_c_sd_amb
        cloture_sd_15c    = stk_c_sd_15c
        cloture_ac_amb    = stk_c_ac_amb + pg_inst_amb
        cloture_ac_15c    = stk_c_ac_15c
        cloture_total_amb = cloture_sd_amb + cloture_ac_amb
        cloture_total_15c = cloture_sd_15c + cloture_ac_15c

        lignes.append({
            'produit':                  produit,
            # Ouverture par régime
            'ouv_sd_amb':               ouv_sd_amb,
            'ouv_sd_15c':               ouv_sd_15c,
            'ouv_ac_amb':               ouv_ac_amb,
            'ouv_ac_15c':               ouv_ac_15c,
            'stock_debut_amb':          stock_debut_amb,
            'stock_debut_15c':          stock_debut_15c,
            # Entrées détaillées
            'rec_sd_amb':               rec_sd_amb,
            'rec_sd_15c':               rec_sd_15c,
            'rec_ac_amb':               rec_ac_amb,
            'rec_ac_15c':               rec_ac_15c,
            'cess_recues_sd_amb':       cess_recues_sd_amb,
            'cess_recues_sd_15c':       cess_recues_sd_15c,
            'cess_recues_ac_amb':       cess_recues_ac_amb,
            'cess_recues_ac_15c':       cess_recues_ac_15c,
            'cess_recues_amb':          cess_recues_amb,
            'cess_recues_15c':          cess_recues_15c,
            'recl_sd_entree_amb':       recl_sd_entree_amb,
            'recl_sd_entree_15c':       recl_sd_entree_15c,
            'recl_ac_entree_amb':       recl_ac_entree_amb,
            'recl_ac_entree_15c':       recl_ac_entree_15c,
            'total_entrees_sd_amb':     total_entrees_sd_amb,
            'total_entrees_sd_15c':     total_entrees_sd_15c,
            'total_entrees_ac_amb':     total_entrees_ac_amb,
            'total_entrees_ac_15c':     total_entrees_ac_15c,
            'total_entrees_amb':        total_entrees_amb,
            'total_entrees_15c':        total_entrees_15c,
            # Sorties détaillées
            'livr_ac_amb':              livr_ac_amb,
            'livr_ac_15c':              livr_ac_15c,
            'livr_sd_amb':              livr_sd_amb,
            'livr_sd_15c':              livr_sd_15c,
            'cess_emises_sd_amb':       cess_emises_sd_amb,
            'cess_emises_sd_15c':       cess_emises_sd_15c,
            'cess_emises_ac_amb':       cess_emises_ac_amb,
            'cess_emises_ac_15c':       cess_emises_ac_15c,
            'cess_emises_amb':          cess_emises_amb,
            'cess_emises_15c':          cess_emises_15c,
            'recl_sd_sortie_amb':       recl_sd_sortie_amb,
            'recl_sd_sortie_15c':       recl_sd_sortie_15c,
            'recl_ac_sortie_amb':       recl_ac_sortie_amb,
            'recl_ac_sortie_15c':       recl_ac_sortie_15c,
            'total_sorties_sd_amb':     total_sorties_sd_amb,
            'total_sorties_sd_15c':     total_sorties_sd_15c,
            'total_sorties_ac_amb':     total_sorties_ac_amb,
            'total_sorties_ac_15c':     total_sorties_ac_15c,
            'total_sorties_amb':        total_sorties_amb,
            'total_sorties_15c':        total_sorties_15c,
            # Stock Comptable par régime
            'stk_c_sd_amb':             stk_c_sd_amb,
            'stk_c_sd_15c':             stk_c_sd_15c,
            'stk_c_ac_amb':             stk_c_ac_amb,
            'stk_c_ac_15c':             stk_c_ac_15c,
            'stock_fin_comptable_amb':  stock_fin_c_amb,
            'stock_fin_comptable_15c':  stock_fin_c_15c,
            # Quote-part P/G Installation & Ratio
            'pg_inst_amb':              pg_inst_amb,
            'ratio':                    ratio,
            # Stock Clôture
            'cloture_sd_amb':           cloture_sd_amb,
            'cloture_sd_15c':           cloture_sd_15c,
            'cloture_ac_amb':           cloture_ac_amb,
            'cloture_ac_15c':           cloture_ac_15c,
            'cloture_total_amb':        cloture_total_amb,
            'cloture_total_15c':        cloture_total_15c,
        })

    return lignes


@marketeur_required
def etat_stock_ouverture_marketeur(request):
    from SGDS.models import Produit
    from datetime import datetime, timedelta
    periodes       = _periodes_disponibles()
    periode        = _get_periode(request)
    mkt            = request.user.marketeur
    produits       = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
    produit_id     = request.GET.get('produit_id')
    date_str       = request.GET.get('date', '').strip()
    produit_filtre = None
    date_filtre    = None
    date_fin_override = None

    if produit_id:
        try:
            produit_filtre = Produit.objects.get(uuid=produit_id, statut='ACTIF')
        except (Produit.DoesNotExist, ValueError):
            pass

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                # Ouverture du jour J = mouvements jusqu'à J-1 (exclu J)
                date_fin_override = d - timedelta(days=1)
        except ValueError:
            pass

    lignes_all = _calculer_stock_ouverture_fermeture_marketeur(periode, mkt, date_fin_override) if periode else []
    lignes = [l for l in lignes_all
              if produit_filtre is None or l['produit'].pk == produit_filtre.pk]
    return render(request, 'Espace_Marketeur/mensuel/stock_ouverture.html', {
        'periodes':       periodes,
        'periode':        periode,
        'lignes':         lignes,
        'produits':       produits,
        'produit_filtre': produit_filtre,
        'marketeur':      mkt,
        'date_filtre':    date_filtre,
    })


@marketeur_required
def etat_stock_fermeture_marketeur(request):
    from SGDS.models import Produit
    from datetime import datetime
    periodes       = _periodes_disponibles()
    periode        = _get_periode(request)
    mkt            = request.user.marketeur
    produits       = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
    produit_id     = request.GET.get('produit_id')
    date_str       = request.GET.get('date', '').strip()
    produit_filtre = None
    date_filtre    = None
    date_fin_override = None

    if produit_id:
        try:
            produit_filtre = Produit.objects.get(uuid=produit_id, statut='ACTIF')
        except (Produit.DoesNotExist, ValueError):
            pass

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                # Fermeture du jour J = mouvements jusqu'à J inclus
                date_fin_override = d
        except ValueError:
            pass

    lignes_all = _calculer_stock_ouverture_fermeture_marketeur(periode, mkt, date_fin_override) if periode else []
    lignes = [l for l in lignes_all
              if produit_filtre is None or l['produit'].pk == produit_filtre.pk]
    return render(request, 'Espace_Marketeur/mensuel/stock_fermeture.html', {
        'periodes':       periodes,
        'periode':        periode,
        'lignes':         lignes,
        'produits':       produits,
        'produit_filtre': produit_filtre,
        'marketeur':      mkt,
        'date_filtre':    date_filtre,
    })


@login_required
def etat_global_mensuel_depot(request):
    from SGDS.models import Societe, Marketeur
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_dashboard')

    periodes    = _periodes_disponibles()
    periode     = _get_periode(request)
    marketeurs  = list(Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale'))
    marketeur   = _get_marketeur_optionnel(request)
    rapport     = _calculer_global_depot(periode, marketeur) if periode else {}
    societe     = Societe.get_instance()

    return render(request, 'Etat/mensuel/global_depot.html', {
        'periodes':      periodes,
        'periode':       periode,
        'rapport':       rapport,
        'marketeurs':    marketeurs,
        'marketeur_sel': marketeur,
        'societe':       societe,
    })


@login_required
def etat_global_mensuel_depot_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    if request.user.is_marketeur_role:
        return redirect('client_dashboard')

    periode  = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_global_depot')

    marketeur = _get_marketeur_optionnel(request)
    rapport   = _calculer_global_depot(periode, marketeur)
    societe   = Societe.get_instance()
    contenu   = _xlsx_global_depot(rapport, societe)

    suffix = f"_{marketeur.sigle or 'mkt'}" if marketeur else "_TOUS"
    nom = f"global_depot_{periode.mois:02d}_{periode.annee}{suffix}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_global_depot(rapport, societe):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    periode = rapport.get('periode')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Global Dépôt"

    ws.merge_cells('A1:J1')
    ws['A1'] = societe.raison_sociale if societe else 'SGDS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"ÉTAT GLOBAL MENSUEL DÉPÔT — {periode}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    BLEU  = PatternFill(fill_type='solid', fgColor='1F4E79')
    BLEU2 = PatternFill(fill_type='solid', fgColor='2E75B6')
    COLS  = ['Date', 'N°Enreg.', 'Marketeur', 'Régime', 'Produit',
             'Libellé', 'V Amb (L)', 'V @15°C (L)', 'Mode règl.', 'Notes']
    WIDTHS = [12, 18, 22, 8, 14, 28, 14, 14, 14, 25]

    TYPES_LABELS = {
        'ENTREE': 'ENTRÉES',
        'SORTIE': 'SORTIES',
        'CESSION': 'CESSIONS',
        'ACQUITTEMENT': 'ACQUITTEMENTS',
    }

    row = 4
    par_type = rapport.get('par_type', {})

    for type_key, label in TYPES_LABELS.items():
        mvts = par_type.get(type_key, [])
        if not mvts:
            continue

        ws.merge_cells(f'A{row}:J{row}')
        c = ws.cell(row=row, column=1, value=f"── {label} ({len(mvts)}) ──")
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = BLEU
        row += 1

        for ci, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = BLEU2
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(ci)].width = w
        row += 1

        for i, m in enumerate(mvts):
            # Libellé
            if type_key == 'ENTREE':
                libelle = m.provenance or ''
            elif type_key == 'SORTIE':
                libelle = m.destination or m.code_destination or 'BON DE SORTIE'
            elif type_key == 'CESSION':
                dest = m.cession_marketeur_destinataire
                libelle = f"→ {dest.sigle or dest.raison_sociale}" if dest else ''
            else:
                libelle = 'Acquittement douanier'

            if type_key == 'ENTREE':
                v_amb, v_15c = m.volume_ambiant_recu, m.volume_15c_recu
            elif type_key == 'SORTIE':
                v_amb, v_15c = m.volume_ambiant_sortie, m.volume_15c_sortie
            elif type_key == 'CESSION':
                v_amb, v_15c = m.cession_volume_ambiant, m.cession_volume_15c
            else:
                v_amb, v_15c = None, None

            vals = [
                str(m.date_mouvement),
                m.numero_enregistrement or '',
                str(m.marketeur),
                'SD' if m.regime_douanier == 'SOUS_DOUANE' else 'AC',
                m.produit.nom,
                libelle,
                float(_D(v_amb)) if v_amb else '',
                float(_D(v_15c)) if v_15c else '',
                m.mode_reglement or '',
                m.notes or '',
            ]
            fill = PatternFill(fill_type='solid', fgColor='EBF3FB') if i % 2 == 0 else None
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                if fill:
                    c.fill = fill
                if isinstance(v, float):
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right')
            row += 1

        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════
#  ÉTAT 3 — GLOBAL MENSUEL RJJ
# ═════════════════════════════════════════════════════════════

def _calculer_rjj(periode, cuve_id=None):
    from SGDS.models import JaugeageJour, Cuve

    qs = (
        JaugeageJour.objects
        .filter(date_jaugeage__range=(periode.date_debut, periode.date_fin))
        .prefetch_related('mesures__cuve__parametre_jaugeage')
        .order_by('date_jaugeage', 'heure_jaugeage')
    )
    jaugeages = list(qs)
    cuves     = list(Cuve.objects.filter(parametre_jaugeage__isnull=False)
                     .select_related('produit', 'parametre_jaugeage')
                     .order_by('numero'))
    if cuve_id:
        cuves = [c for c in cuves if str(c.uuid) == str(cuve_id)]

    lignes = []
    for j in jaugeages:
        mesures_dict = {m.cuve_id: m for m in j.mesures.all()}
        for cuve in cuves:
            m = mesures_dict.get(cuve.pk)
            if not m:
                continue
            lignes.append({
                'jaugeage_pk':     j.pk,
                'jaugeage_uuid':   j.uuid,
                'jaugeage_slug':   j.slug,
                'date':            j.date_jaugeage,
                'type_jaugeage':   j.get_type_jaugeage_display(),
                'heure':           j.heure_jaugeage,
                'est_valide':      j.est_valide,
                'cuve_numero':     cuve.numero,
                'produit_nom':     cuve.produit.nom if cuve.produit else '—',
                'creux':           m.creux_mesure,
                'creux_corrige':   m.creux_corrige,
                'hauteur_produit': m.hauteur_produit,
                't1':              m.t1,
                't2':              m.t2,
                't3':              m.t3,
                'temp_moy':        m.temperature_moyenne,
                'temp_obs':        m.temperature_obs,
                'densite_moy':     m.densite_moyenne,
                'densite_15c':     m.densite_15c,
                'facteur_vcf':     m.facteur_vcf,
                'v_amb_bac':       m.volume_ambiant_bac,
                'v_amb_depot':     m.volume_ambiant_depot,
                'v_15c':           m.volume_standard_15c,
                'v_disponible':    m.volume_disponible,
            })
    return {
        'periode':   periode,
        'jaugeages': jaugeages,
        'cuves':     cuves,
        'lignes':    lignes,
    }


@login_required
def etat_global_mensuel_rjj(request):
    from SGDS.models import Societe, Cuve
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_dashboard')

    periodes  = _periodes_disponibles()
    periode   = _get_periode(request)
    cuves     = list(Cuve.objects.filter(parametre_jaugeage__isnull=False).order_by('numero'))
    cuve_id   = request.GET.get('cuve')
    rapport   = _calculer_rjj(periode, cuve_id) if periode else {'lignes': [], 'cuves': []}
    societe   = Societe.get_instance()

    return render(request, 'Etat/mensuel/rjj.html', {
        'periodes': periodes,
        'periode':  periode,
        'rapport':  rapport,
        'cuves':    cuves,
        'cuve_sel': cuve_id,
        'societe':  societe,
    })


@login_required
def etat_global_mensuel_rjj_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    if request.user.is_marketeur_role:
        return redirect('client_dashboard')

    periode = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_rjj')

    cuve_id = request.GET.get('cuve')
    rapport = _calculer_rjj(periode, cuve_id)
    societe = Societe.get_instance()
    contenu = _xlsx_rjj(rapport, societe)

    nom = f"rjj_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_rjj(rapport, societe):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    periode = rapport.get('periode')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RJJ"

    COLS   = ['Date', 'Type', 'Cuve', 'Produit', 'Creux', 'Creux corr.',
              'H.Produit', 'T1', 'T2', 'T3', 'T°Moy', 'T°Obs',
              'D.Moy', 'D15°C', 'VCF',
              'V Amb Bac', 'V Amb Dépôt', 'V @15°C', 'V Dispo']
    WIDTHS = [12, 8, 8, 12, 9, 10, 10, 7, 7, 7, 8, 8, 8, 8, 8, 13, 14, 12, 12]

    ws.merge_cells(f'A1:{get_column_letter(len(COLS))}1')
    ws['A1'] = societe.raison_sociale if societe else 'SGDS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{get_column_letter(len(COLS))}2')
    ws['A2'] = f"RELEVÉ JOURNALIER DE JAUGEAGE (RJJ) — {periode}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    BLEU = PatternFill(fill_type='solid', fgColor='1F4E79')
    row = 4
    for ci, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = BLEU
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 28
    row += 1

    def _f(v):
        return float(v) if v is not None else ''

    for l in rapport.get('lignes', []):
        ws.cell(row=row, column=1,  value=str(l['date']))
        ws.cell(row=row, column=2,  value=l['type_jaugeage'])
        ws.cell(row=row, column=3,  value=l['cuve_numero'])
        ws.cell(row=row, column=4,  value=l['produit_nom'])
        ws.cell(row=row, column=5,  value=l['creux'])
        ws.cell(row=row, column=6,  value=l['creux_corrige'])
        ws.cell(row=row, column=7,  value=l['hauteur_produit'])
        ws.cell(row=row, column=8,  value=_f(l['t1']))
        ws.cell(row=row, column=9,  value=_f(l['t2']))
        ws.cell(row=row, column=10, value=_f(l['t3']))
        ws.cell(row=row, column=11, value=_f(l['temp_moy']))
        ws.cell(row=row, column=12, value=_f(l['temp_obs']))
        ws.cell(row=row, column=13, value=_f(l['densite_moy']))
        ws.cell(row=row, column=14, value=_f(l['densite_15c']))
        ws.cell(row=row, column=15, value=_f(l['facteur_vcf']))
        for ci, key in enumerate(['v_amb_bac', 'v_amb_depot', 'v_15c', 'v_disponible'], 16):
            c = ws.cell(row=row, column=ci, value=_f(l[key]))
            if isinstance(c.value, float):
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════
#  ÉTAT 4 — COULAGE RÉPARTITION  (admin + marketeur)
# ═════════════════════════════════════════════════════════════

def _rapport_coulage_pour_periode(periode):
    """Retourne (rapport_dict, source_str) — snapshot si clôturé, temps réel sinon."""
    from SGDS.services.coulage_repartition import calculer_repartition_coulage
    if periode.statut == 'CLOTUREE' and hasattr(periode, 'cloture_coulage'):
        from SGDS.views.coulage import RepartitionCoulageView
        vue = RepartitionCoulageView()
        return vue._rapport_depuis_snapshot(periode.cloture_coulage), 'SNAPSHOT'
    return calculer_repartition_coulage(periode), 'TEMPS_REEL'


@login_required
def etat_coulage_repartition(request):
    from SGDS.models import Societe, PeriodeComptable
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_dashboard')

    periodes = _periodes_disponibles()

    # Si aucune période choisie, prendre la dernière clôturée (pas forcément la plus récente)
    if request.GET.get('periode_id'):
        periode = _get_periode(request)
    else:
        periode = (
            PeriodeComptable.objects.filter(statut='CLOTUREE')
            .order_by('-annee', '-mois').first()
            or _get_periode(request)
        )

    societe  = Societe.get_instance()
    rapport  = source = erreur = None

    if periode:
        cloture_dispo = (periode.statut == 'CLOTUREE') or hasattr(periode, 'cloture_coulage')
        if cloture_dispo:
            rapport, source = _rapport_coulage_pour_periode(periode)
        else:
            erreur = "La période n'est pas encore clôturée — coulage répartition indisponible."

    return render(request, 'Etat/mensuel/coulage_repartition.html', {
        'periodes': periodes,
        'periode':  periode,
        'rapport':  rapport,
        'source':   source,
        'erreur':   erreur,
        'societe':  societe,
    })


@login_required
def etat_coulage_repartition_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    if request.user.is_marketeur_role:
        return redirect('client_dashboard')

    periode = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_coulage_repartition')

    if not (periode.statut == 'CLOTUREE' or hasattr(periode, 'cloture_coulage')):
        messages.warning(request, "Période non clôturée.")
        return redirect('etat_mensuel_coulage_repartition')

    rapport, _ = _rapport_coulage_pour_periode(periode)
    societe = Societe.get_instance()
    contenu = _xlsx_coulage_repartition(rapport, societe)

    nom = f"coulage_repartition_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


@marketeur_required
def etat_coulage_repartition_marketeur(request):
    from SGDS.models import Societe, PeriodeComptable
    periodes = _periodes_disponibles()

    if request.GET.get('periode_id'):
        periode = _get_periode(request)
    else:
        periode = (
            PeriodeComptable.objects.filter(statut='CLOTUREE')
            .order_by('-annee', '-mois').first()
            or _get_periode(request)
        )

    societe  = Societe.get_instance()
    mkt      = request.user.marketeur
    rapport  = source = erreur = None

    if periode:
        cloture_dispo = (periode.statut == 'CLOTUREE') or hasattr(periode, 'cloture_coulage')
        if cloture_dispo:
            rapport_complet, source = _rapport_coulage_pour_periode(periode)
            lignes_mkt = [
                l for l in rapport_complet.get('lignes', [])
                if l['marketeur'].pk == mkt.pk
            ]
            rapport = {**rapport_complet, 'lignes': lignes_mkt}
        else:
            erreur = "La période n'est pas encore clôturée — coulage répartition indisponible."

    return render(request, 'Espace_Marketeur/mensuel/coulage_repartition.html', {
        'periodes': periodes,
        'periode':  periode,
        'rapport':  rapport,
        'source':   source,
        'erreur':   erreur,
        'societe':  societe,
        'marketeur': mkt,
    })


@marketeur_required
def etat_coulage_repartition_marketeur_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse

    periode = _get_periode(request)
    if not periode:
        return redirect('client_mensuel_coulage_repartition')
    if not (periode.statut == 'CLOTUREE' or hasattr(periode, 'cloture_coulage')):
        messages.warning(request, "Période non clôturée.")
        return redirect('client_mensuel_coulage_repartition')

    mkt = request.user.marketeur
    rapport_complet, _ = _rapport_coulage_pour_periode(periode)
    lignes_mkt = [l for l in rapport_complet.get('lignes', []) if l['marketeur'].pk == mkt.pk]
    rapport    = {**rapport_complet, 'lignes': lignes_mkt}
    societe    = Societe.get_instance()
    contenu    = _xlsx_coulage_repartition(rapport, societe, marketeur_filtre=mkt)

    nom = f"coulage_repartition_{mkt.sigle or 'mkt'}_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_coulage_repartition(rapport, societe, marketeur_filtre=None):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    if not rapport:
        buf = io.BytesIO()
        openpyxl.Workbook().save(buf)
        buf.seek(0)
        return buf.read()

    periode  = rapport.get('periode')
    produits = rapport.get('produits', [])
    lignes   = rapport.get('lignes', [])
    totaux   = rapport.get('totaux', {})

    FIELDS  = ['brut_entree', 'coul_entree', 'entree_nette', 'sortie',
               'base_qp_coul', 'coef_qp_coul', 'qp_coul', 'volume_sorti']
    SHDR    = ['Brut Entrée', 'Coul.Entrée', 'Entrée Nette', 'Sortie',
               'Base QP', 'Coef QP', 'QP Coul.', 'Vol.Sorti']

    n_cols    = 1 + len(produits) * len(FIELDS)
    last_col  = get_column_letter(n_cols)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coulage Répartition"

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = societe.raison_sociale if societe else 'SGDS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col}2')
    suffix_mkt = (f" — {marketeur_filtre.sigle or marketeur_filtre.raison_sociale}"
                  if marketeur_filtre else "")
    ws['A2'] = f"COULAGE RÉPARTITION — {periode}{suffix_mkt}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    BLEU  = PatternFill(fill_type='solid', fgColor='1F4E79')
    BLEU2 = PatternFill(fill_type='solid', fgColor='2E75B6')
    VERT  = PatternFill(fill_type='solid', fgColor='E2EFDA')
    ZEBRA = PatternFill(fill_type='solid', fgColor='EBF3FB')

    row = 4
    # En-tête produits (rang 1)
    c = ws.cell(row=row, column=1, value='Marketeur')
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = BLEU
    ws.column_dimensions['A'].width = 22

    prod_col_start = {}
    col = 2
    for p in produits:
        prod_col_start[p.pk] = col
        end_col = col + len(FIELDS) - 1
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
        c = ws.cell(row=row, column=col, value=p.nom.upper())
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = BLEU
        c.alignment = Alignment(horizontal='center')
        col += len(FIELDS)
    row += 1

    # Sous-colonnes (rang 2)
    ws.cell(row=row, column=1).fill = BLEU2
    col = 2
    for p in produits:
        for sh in SHDR:
            c = ws.cell(row=row, column=col, value=sh)
            c.font = Font(bold=True, color='FFFFFF', size=8)
            c.fill = BLEU2
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = 13
            col += 1
    ws.row_dimensions[row].height = 28
    row += 1

    def _f(v):
        return float(v) if v is not None else 0.0

    for i, ligne in enumerate(lignes):
        mkt = ligne['marketeur']
        c = ws.cell(row=row, column=1, value=mkt.sigle or mkt.raison_sociale)
        if i % 2 == 0:
            c.fill = ZEBRA
        for p in produits:
            pp   = ligne.get('par_produit', {}).get(p.pk, {})
            col  = prod_col_start[p.pk]
            for fi, field in enumerate(FIELDS):
                cell = ws.cell(row=row, column=col+fi, value=_f(pp.get(field)))
                cell.number_format = '#,##0.00' if field != 'coef_qp_coul' else '0.000000'
                cell.alignment = Alignment(horizontal='right')
                if i % 2 == 0:
                    cell.fill = ZEBRA
        row += 1

    # Totaux
    c = ws.cell(row=row, column=1, value='TOTAUX')
    c.font = Font(bold=True)
    c.fill = VERT
    for p in produits:
        tp  = totaux.get('par_produit', {}).get(p.pk, {})
        col = prod_col_start[p.pk]
        for fi, field in enumerate(FIELDS):
            cell = ws.cell(row=row, column=col+fi, value=_f(tp.get(field)))
            cell.font = Font(bold=True)
            cell.fill = VERT
            cell.number_format = '#,##0.00' if field != 'coef_qp_coul' else '0.000000'
            cell.alignment = Alignment(horizontal='right')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════
#  ÉTAT 5 — STOCK À 15° (admin + marketeur)
#  Même format pivot que le Stock Mensuel B, avec en plus la
#  valeur @15°C affichée à côté de chaque valeur AMB.
# ═════════════════════════════════════════════════════════════

def _calculer_stock_a_15(periode, marketeur=None, date_fin_override=None):
    """
    Format pivot (lignes = statistiques, colonnes = produits actifs),
    identique à Stock Mensuel B, mais avec les volumes AMB ET @15°C.
    marketeur=None => Tous marketeurs (stock global depuis StockOuverture)
    marketeur=<obj> => filtre par ce marketeur

    date_fin_override : si fourni, calcule un instantane "fermeture du jour"
    (mêmes règles que _calculer_stock_ambiant).
    """
    from SGDS.models import Produit, StockOuverture, Mouvement, JaugeageJour, InventaireInitialMarketeur

    produits = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))

    date_fin_calc     = date_fin_override if date_fin_override is not None else periode.date_fin
    filtre_journalier = date_fin_override is not None

    # Repli 1ere periode (aucune periode precedente => pas encore de StockOuverture
    # resolue depuis un jaugeage) : on agrege les inventaires initiaux par produit,
    # comme le fait _calculer_stock_ouverture_fermeture.
    _inv_initiaux = {}
    if marketeur is None and periode.periode_precedente() is None:
        for inv in InventaireInitialMarketeur.objects.filter(date_inventaire__lte=periode.date_fin):
            pid = inv.produit_id
            if pid not in _inv_initiaux:
                _inv_initiaux[pid] = {'amb': _Z, '15c': _Z}
            _inv_initiaux[pid]['amb'] += _D(inv.volume_ambiant)
            _inv_initiaux[pid]['15c'] += _D(inv.volume_15c)

    # Marketeur specifique : Pertes/Gains AMB = quote-part de coulage du marketeur
    # (pas le jaugeage du depot, qui n'est pas attribuable a un marketeur en particulier).
    # Le coulage n'est pas suivi en @15°C dans ce systeme.
    qp_coul_par_produit = {}
    if marketeur is not None and not filtre_journalier:
        try:
            from SGDS.services.coulage_repartition import calculer_repartition_coulage
            rapport_coul = calculer_repartition_coulage(periode, marketeurs=[marketeur])
            if rapport_coul['lignes']:
                qp_coul_par_produit = rapport_coul['lignes'][0]['par_produit']
        except Exception:
            pass

    periode_suiv = periode.periode_suivante()
    sf_amb_map   = {}
    sf_15c_map   = {}
    if filtre_journalier:
        date_j = date_fin_override
        dernier_j = (
            JaugeageJour.objects
            .filter(date_jaugeage=date_j, est_valide=True)
            .order_by('-heure_jaugeage', '-date_creation')
            .first()
        )
        if not dernier_j and date_j >= periode.date_debut:
            dernier_j = (
                JaugeageJour.objects
                .filter(date_jaugeage__range=(periode.date_debut, date_j), est_valide=True)
                .order_by('-date_jaugeage', '-heure_jaugeage', '-date_creation')
                .first()
            )
        if dernier_j:
            for mesure in dernier_j.mesures.select_related('cuve__produit').all():
                cuve = mesure.cuve
                if cuve.produit is None:
                    continue
                pid = cuve.produit_id
                if mesure.volume_ambiant_depot is not None:
                    sf_amb_map[pid] = sf_amb_map.get(pid, _Z) + _D(mesure.volume_ambiant_depot)
                if mesure.volume_standard_15c_calcule is not None:
                    sf_15c_map[pid] = sf_15c_map.get(pid, _Z) + _D(mesure.volume_standard_15c_calcule)
    else:
        # Période entière : StockOuverture de la période suivante en priorité
        if periode_suiv:
            for so in StockOuverture.objects.filter(periode=periode_suiv).select_related('produit'):
                sf_amb_map[so.produit_id] = sf_amb_map.get(so.produit_id, _Z) + _D(so.volume_ambiant)
                sf_15c_map[so.produit_id] = sf_15c_map.get(so.produit_id, _Z) + _D(so.volume_15c)
        # Repli : dernier jaugeage validé de la période (StockOuverture de la
        # période suivante pas encore résolue / période en cours)
        if not sf_amb_map:
            dernier_j = (
                JaugeageJour.objects
                .filter(date_jaugeage__range=(periode.date_debut, periode.date_fin), est_valide=True)
                .order_by('-date_jaugeage', '-heure_jaugeage', '-date_creation')
                .first()
            )
            if dernier_j:
                for mesure in dernier_j.mesures.select_related('cuve__produit').all():
                    cuve = mesure.cuve
                    if cuve.produit is None:
                        continue
                    pid = cuve.produit_id
                    if mesure.volume_ambiant_depot is not None:
                        sf_amb_map[pid] = sf_amb_map.get(pid, _Z) + _D(mesure.volume_ambiant_depot)
                    if mesure.volume_standard_15c_calcule is not None:
                        sf_15c_map[pid] = sf_15c_map.get(pid, _Z) + _D(mesure.volume_standard_15c_calcule)

    stk_ouv_amb   = []
    cumul_ent_amb = []
    cumul_sor_amb = []
    stk_cpt_amb   = []
    pertes_gains_vals = []
    ratios_vals       = []
    stk_ferm_amb  = []

    for produit in produits:
        if marketeur is None:
            # TM : stock global depuis StockOuverture (somme SD + Acquittée),
            # avec repli sur les inventaires initiaux si rien n'a encore été résolu
            # (1ère période, avant tout jaugeage de la période précédente).
            from django.db.models import Sum as _Sum
            agg_ouv = (
                StockOuverture.objects
                .filter(periode=periode, produit=produit)
                .aggregate(v_amb=_Sum('volume_ambiant'), v_15c=_Sum('volume_15c'))
            )
            if agg_ouv['v_amb'] is not None:
                ouv_amb = _D(agg_ouv['v_amb'] or 0)
                ouv_15c = _D(agg_ouv['v_15c'] or 0)
            else:
                _inv = _inv_initiaux.get(produit.pk, {})
                ouv_amb = _inv.get('amb', _Z)
                ouv_15c = _inv.get('15c', _Z)
            mvts_qs = Mouvement.objects.filter(
                produit=produit,
                date_mouvement__range=(periode.date_debut, date_fin_calc),
            )
        else:
            # Marketeur spécifique : stock d'ouverture reporté depuis la
            # fermeture du mois précédent (StockOuvertureMarketeur), au lieu
            # de rejouer tout l'historique des mouvements.
            from django.db.models import Sum as _Sum
            from SGDS.models import StockOuvertureMarketeur
            agg_ouv_mkt = (
                StockOuvertureMarketeur.objects
                .filter(periode=periode, marketeur=marketeur, produit=produit)
                .aggregate(v_amb=_Sum('volume_ambiant'), v_15c=_Sum('volume_15c'))
            )
            if agg_ouv_mkt['v_amb'] is not None:
                ouv_amb = _D(agg_ouv_mkt['v_amb'] or 0)
                ouv_15c = _D(agg_ouv_mkt['v_15c'] or 0)
            else:
                inv_agg = (
                    InventaireInitialMarketeur.objects
                    .filter(marketeur=marketeur, produit=produit)
                    .aggregate(v_amb=_Sum('volume_ambiant'), v_15c=_Sum('volume_15c'))
                )
                ouv_amb = _D(inv_agg['v_amb'] or 0)
                ouv_15c = _D(inv_agg['v_15c'] or 0)
            mvts_qs = Mouvement.objects.filter(
                produit=produit, marketeur=marketeur,
                date_mouvement__range=(periode.date_debut, date_fin_calc),
            )

        mvts = list(mvts_qs)
        ent_amb = sum(_D(m.volume_ambiant_recu)   for m in mvts if m.type_mouvement == 'ENTREE')
        ent_15c = sum(_D(m.volume_15c_recu)       for m in mvts if m.type_mouvement == 'ENTREE')
        sor_amb = sum(_D(m.volume_ambiant_sortie) for m in mvts if m.type_mouvement == 'SORTIE')
        sor_15c = sum(_D(m.volume_15c_sortie)     for m in mvts if m.type_mouvement == 'SORTIE')

        if marketeur is not None:
            # Cessions émises par ce marketeur durant la période -> sortie
            sor_amb += sum(_D(m.cession_volume_ambiant) for m in mvts if m.type_mouvement == 'CESSION')
            sor_15c += sum(_D(m.cession_volume_15c)     for m in mvts if m.type_mouvement == 'CESSION')
            # Cessions reçues par ce marketeur durant la période -> entrée
            # (Mouvement.marketeur = l'émetteur de la cession, pas le destinataire,
            # donc requête séparée filtrée sur cession_marketeur_destinataire)
            cess_recues = Mouvement.objects.filter(
                produit=produit, type_mouvement='CESSION',
                cession_marketeur_destinataire=marketeur,
                date_mouvement__range=(periode.date_debut, date_fin_calc),
            )
            for m in cess_recues:
                ent_amb += _D(m.cession_volume_ambiant)
                ent_15c += _D(m.cession_volume_15c)

        cpt_amb = ouv_amb + ent_amb - sor_amb
        cpt_15c = ouv_15c + ent_15c - sor_15c

        if marketeur is None:
            # TM : Pertes/Gains = jaugeage physique du dépôt - stock comptable
            sf_amb = sf_amb_map.get(produit.pk)
            sf_15c = sf_15c_map.get(produit.pk)
            pg_amb = (sf_amb - cpt_amb) if sf_amb is not None else None
            pg_15c = (sf_15c - cpt_15c) if sf_15c is not None else None
        else:
            # Marketeur : Pertes/Gains AMB = sa quote-part de coulage (pas le jaugeage
            # du dépôt). Le coulage n'est pas suivi en @15°C : le stock fermeture @15°C
            # reste donc égal au stock comptable @15°C, sans ajustement.
            qp = qp_coul_par_produit.get(produit.pk)
            if qp is not None:
                pg_amb = qp.get('qp_coul', _Z)
                sf_amb = cpt_amb + pg_amb
            else:
                pg_amb = None
                sf_amb = None
            pg_15c = None
            sf_15c = cpt_15c
        ratio_amb = (float(pg_amb) / float(sor_amb) * 100) if (pg_amb is not None and sor_amb) else None
        ratio_15c = (float(pg_15c) / float(sor_15c) * 100) if (pg_15c is not None and sor_15c) else None

        stk_ouv_amb.append({'amb': ouv_amb, 'c15': ouv_15c})
        cumul_ent_amb.append({'amb': ent_amb, 'c15': ent_15c})
        cumul_sor_amb.append({'amb': sor_amb, 'c15': sor_15c})
        stk_cpt_amb.append({'amb': cpt_amb, 'c15': cpt_15c})
        pertes_gains_vals.append({'amb': pg_amb, 'c15': pg_15c})
        ratios_vals.append({'amb': ratio_amb, 'c15': ratio_15c})
        stk_ferm_amb.append({'amb': sf_amb, 'c15': sf_15c})

    return {
        'periode':       periode,
        'produits':      produits,
        'date_ouv':      periode.date_debut,
        'date_ferm':     date_fin_calc,
        'date_filtre':   date_fin_override,
        'stk_ouverture': stk_ouv_amb,
        'cumul_entrees': cumul_ent_amb,
        'cumul_sorties': cumul_sor_amb,
        'stk_comptable': stk_cpt_amb,
        'pertes_gains':  pertes_gains_vals,
        'ratios':        ratios_vals,
        'stk_fermeture': stk_ferm_amb,
    }


@login_required
def etat_stock_mensuel_15(request):
    from SGDS.models import Societe, Marketeur as MktModel
    from datetime import datetime
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_dashboard')

    periodes    = _periodes_disponibles()
    periode     = _get_periode(request)
    marketeurs  = list(MktModel.objects.filter(statut='ACTIF').order_by('raison_sociale'))
    marketeur   = _get_marketeur_optionnel(request)
    date_str    = request.GET.get('date', '').strip()
    date_filtre = None
    date_fin_override = None

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                date_fin_override = d
        except ValueError:
            pass

    rapport    = _calculer_stock_a_15(periode, marketeur, date_fin_override) if periode else {}
    societe    = Societe.get_instance()

    return render(request, 'Etat/mensuel/stock_a_15.html', {
        'periodes':      periodes,
        'periode':       periode,
        'rapport':       rapport,
        'marketeurs':    marketeurs,
        'marketeur_sel': marketeur,
        'societe':       societe,
        'date_filtre':   date_filtre,
    })


@login_required
def etat_stock_mensuel_15_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    if request.user.is_marketeur_role:
        return redirect('client_dashboard')

    periode   = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_stock_15')
    marketeur = _get_marketeur_optionnel(request)
    rapport   = _calculer_stock_a_15(periode, marketeur)
    societe   = Societe.get_instance()
    contenu   = _xlsx_stock_a_15(rapport, societe)

    suffix = f"_{marketeur.sigle or 'mkt'}" if marketeur else "_TM"
    nom = f"stock_a_15_{periode.mois:02d}_{periode.annee}{suffix}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


@marketeur_required
def etat_stock_mensuel_15_marketeur(request):
    from SGDS.models import Societe
    from datetime import datetime
    periodes    = _periodes_disponibles()
    periode     = _get_periode(request)
    mkt         = request.user.marketeur
    date_str    = request.GET.get('date', '').strip()
    date_filtre = None
    date_fin_override = None

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                date_fin_override = d
        except ValueError:
            pass

    rapport  = _calculer_stock_a_15(periode, mkt, date_fin_override) if periode else {}
    societe  = Societe.get_instance()

    return render(request, 'Espace_Marketeur/mensuel/stock_a_15.html', {
        'periodes':    periodes,
        'periode':     periode,
        'rapport':     rapport,
        'societe':     societe,
        'marketeur':   mkt,
        'date_filtre': date_filtre,
    })


@marketeur_required
def etat_stock_mensuel_15_marketeur_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    periode = _get_periode(request)
    if not periode:
        return redirect('client_mensuel_stock_15')
    mkt     = request.user.marketeur
    rapport = _calculer_stock_a_15(periode, mkt)
    societe = Societe.get_instance()
    contenu = _xlsx_stock_a_15(rapport, societe)

    nom = f"stock_a_15_{mkt.sigle or 'mkt'}_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_stock_a_15(rapport, societe):
    """
    Export pivot Stock à 15° : lignes = statistiques, colonnes = produits
    (2 sous-colonnes AMB / @15°C par produit). Même format que l'export B.
    """
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    NAVY   = PatternFill(fill_type='solid', fgColor='1B3A6B')
    OUV    = PatternFill(fill_type='solid', fgColor='EEF2FF')
    ENT    = PatternFill(fill_type='solid', fgColor='F0FDF4')
    SOR    = PatternFill(fill_type='solid', fgColor='FFF7ED')
    CPT    = PatternFill(fill_type='solid', fgColor='EFF6FF')
    PG     = PatternFill(fill_type='solid', fgColor='FEF2F2')
    RAT    = PatternFill(fill_type='solid', fgColor='FAFAFA')
    FERM   = PatternFill(fill_type='solid', fgColor='F0FDF4')
    WHITE  = Font(bold=True, color='FFFFFF')
    NAVY_F = Font(bold=True, color='1B3A6B')
    RED_F  = Font(bold=True, color='DC2626')
    GRN_F  = Font(bold=True, color='16A34A')
    BOLD   = Font(bold=True)
    CENTER = Alignment(horizontal='center', vertical='center')
    RIGHT  = Alignment(horizontal='right', vertical='center')
    LEFT   = Alignment(horizontal='left', vertical='center')

    periode  = rapport.get('periode')
    produits = rapport.get('produits', [])
    n_cols   = 1 + len(produits) * 2

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock a 15"

    last_col = get_column_letter(n_cols)
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = societe.raison_sociale if societe else 'SGDS'
    c.font = Font(bold=True, size=14, color='1B3A6B')
    c.alignment = CENTER

    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']
    c.value = f"ETAT STOCK A 15 DEGRES - {periode}"
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    row = 4
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
    c = ws.cell(row=row, column=1, value='DESIGNATION')
    c.font = WHITE; c.fill = NAVY; c.alignment = CENTER
    ws.column_dimensions['A'].width = 28

    col = 2
    for prod in produits:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col, value=prod.nom.upper())
        c.font = WHITE; c.fill = NAVY; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions[get_column_letter(col + 1)].width = 15
        for ci, label in ((col, 'AMB (L)'), (col + 1, '@15°C (L)')):
            sub = ws.cell(row=row + 1, column=ci, value=label)
            sub.font = WHITE; sub.fill = NAVY; sub.alignment = CENTER
        col += 2
    ws.row_dimensions[row].height = 22
    row += 2

    def _f(v):
        return float(v) if v is not None else None

    STAT_ROWS = [
        ('STOCKS OUVERTURE (L)',  'stk_ouverture', OUV,  NAVY_F, False),
        ('CUMULS ENTREES (L)',    'cumul_entrees', ENT,  BOLD,   False),
        ('CUMULS SORTIES (L)',    'cumul_sorties', SOR,  BOLD,   False),
        ('STOCKS COMPTABLE (L)',  'stk_comptable', CPT,  NAVY_F, False),
        ('PERTES OU GAINS (L)',   'pertes_gains',  PG,   BOLD,   True),
        ('RATIO (%)',             'ratios',        RAT,  BOLD,   True),
        ('STOCKS FERMETURE (L)',  'stk_fermeture', FERM, GRN_F,  False),
    ]

    for label, key, fill, label_font, signe in STAT_ROWS:
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill; c.alignment = LEFT; c.font = label_font

        col = 2
        for entry in rapport.get(key, []):
            for v in (entry.get('amb'), entry.get('c15')):
                num = _f(v)
                cc = ws.cell(row=row, column=col)
                cc.fill = fill; cc.alignment = RIGHT
                if num is None:
                    cc.value = ''
                else:
                    cc.value = num
                    cc.number_format = '0.0000' if key == 'ratios' else '#,##0.00'
                    if signe:
                        cc.font = RED_F if num < 0 else (GRN_F if num > 0 else BOLD)
                col += 1
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════
#  ÉTAT 6 — STOCK AMBIANT  (simplifié AMB — admin + mkt)
# ═════════════════════════════════════════════════════════════

def _calculer_stock_ambiant(periode, marketeur=None, date_fin_override=None):
    """
    Format pivot exact du fichier Excel "STOCK AMB" :
      Lignes   = statistiques (Ouverture, Entrees, Sorties, Comptable, P/G, Ratio, Fermeture)
      Colonnes = produits actifs
    V AMB uniquement.
    marketeur=None => Tous marketeurs (stock global depuis StockOuverture)
    marketeur=<obj> => filtre par ce marketeur

    date_fin_override : si fourni, calcule un instantane "fermeture du jour" -
    les cumuls Entrees/Sorties s'arretent a cette date (incluse) et le Stock
    Fermeture/P-G/Ratio utilisent le jaugeage de ce jour au lieu du
    StockOuverture de la periode suivante (meme logique que la page Fermeture).
    """
    from SGDS.models import Produit, StockOuverture, Mouvement, JaugeageJour, InventaireInitialMarketeur

    produits = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))

    date_fin_calc     = date_fin_override if date_fin_override is not None else periode.date_fin
    filtre_journalier = date_fin_override is not None

    # Repli 1ere periode (aucune periode precedente => pas encore de StockOuverture
    # resolue depuis un jaugeage) : on agrege les inventaires initiaux par produit,
    # comme le fait _calculer_stock_ouverture_fermeture.
    _inv_initiaux = {}
    if marketeur is None and periode.periode_precedente() is None:
        for inv in InventaireInitialMarketeur.objects.filter(date_inventaire__lte=periode.date_fin):
            pid = inv.produit_id
            _inv_initiaux[pid] = _inv_initiaux.get(pid, _Z) + _D(inv.volume_ambiant)

    # Marketeur specifique : Pertes/Gains = quote-part de coulage du marketeur
    # (pas le jaugeage du depot, qui n'est pas attribuable a un marketeur en particulier)
    qp_coul_par_produit = {}
    if marketeur is not None and not filtre_journalier:
        try:
            from SGDS.services.coulage_repartition import calculer_repartition_coulage
            rapport_coul = calculer_repartition_coulage(periode, marketeurs=[marketeur])
            if rapport_coul['lignes']:
                qp_coul_par_produit = rapport_coul['lignes'][0]['par_produit']
        except Exception:
            pass

    # Stock fermeture physique = jaugeage du jour choisi (vue journaliere)
    # ou StockOuverture de la periode suivante (vue periode entiere)
    periode_suiv = periode.periode_suivante()
    sf_physique  = {}
    if filtre_journalier:
        date_j = date_fin_override
        dernier_j = (
            JaugeageJour.objects
            .filter(date_jaugeage=date_j, est_valide=True)
            .order_by('-heure_jaugeage', '-date_creation')
            .first()
        )
        if not dernier_j and date_j >= periode.date_debut:
            dernier_j = (
                JaugeageJour.objects
                .filter(date_jaugeage__range=(periode.date_debut, date_j), est_valide=True)
                .order_by('-date_jaugeage', '-heure_jaugeage', '-date_creation')
                .first()
            )
        if dernier_j:
            for mesure in dernier_j.mesures.select_related('cuve__produit').all():
                cuve = mesure.cuve
                if cuve.produit is None or mesure.volume_ambiant_depot is None:
                    continue
                pid = cuve.produit_id
                sf_physique[pid] = sf_physique.get(pid, _Z) + _D(mesure.volume_ambiant_depot)
    else:
        # Période entière : StockOuverture de la période suivante en priorité
        if periode_suiv:
            for so in StockOuverture.objects.filter(periode=periode_suiv).select_related('produit'):
                sf_physique[so.produit_id] = sf_physique.get(so.produit_id, _Z) + _D(so.volume_ambiant)
        # Repli : dernier jaugeage validé de la période (StockOuverture de la
        # période suivante pas encore résolue / période en cours)
        if not sf_physique:
            dernier_j = (
                JaugeageJour.objects
                .filter(date_jaugeage__range=(periode.date_debut, periode.date_fin), est_valide=True)
                .order_by('-date_jaugeage', '-heure_jaugeage', '-date_creation')
                .first()
            )
            if dernier_j:
                for mesure in dernier_j.mesures.select_related('cuve__produit').all():
                    cuve = mesure.cuve
                    if cuve.produit is None or mesure.volume_ambiant_depot is None:
                        continue
                    pid = cuve.produit_id
                    sf_physique[pid] = sf_physique.get(pid, _Z) + _D(mesure.volume_ambiant_depot)

    stk_ouv_vals      = []
    cumul_ent_vals    = []
    cumul_sor_vals    = []
    stk_cpt_vals      = []
    pertes_gains_vals = []
    ratios_vals       = []
    stk_ferm_vals     = []

    for produit in produits:
        if marketeur is None:
            # TM : stock global depuis StockOuverture (somme SD + Acquittée),
            # avec repli sur les inventaires initiaux si rien n'a encore été résolu
            # (1ère période, avant tout jaugeage de la période précédente).
            from django.db.models import Sum as _Sum
            agg_ouv = (
                StockOuverture.objects
                .filter(periode=periode, produit=produit)
                .aggregate(v_amb=_Sum('volume_ambiant'))
            )
            if agg_ouv['v_amb'] is not None:
                ouv = _D(agg_ouv['v_amb'])
            else:
                ouv = _inv_initiaux.get(produit.pk, _Z)
            mvts_qs = Mouvement.objects.filter(
                produit=produit,
                date_mouvement__range=(periode.date_debut, date_fin_calc),
            )
        else:
            # Marketeur spécifique : stock d'ouverture reporté depuis la
            # fermeture du mois précédent (StockOuvertureMarketeur), au lieu
            # de rejouer tout l'historique des mouvements.
            from django.db.models import Sum as _Sum
            from SGDS.models import StockOuvertureMarketeur
            agg_ouv_mkt = (
                StockOuvertureMarketeur.objects
                .filter(periode=periode, marketeur=marketeur, produit=produit)
                .aggregate(v_amb=_Sum('volume_ambiant'))
            )
            if agg_ouv_mkt['v_amb'] is not None:
                ouv = _D(agg_ouv_mkt['v_amb'] or 0)
            else:
                inv_agg = (
                    InventaireInitialMarketeur.objects
                    .filter(marketeur=marketeur, produit=produit)
                    .aggregate(v_amb=_Sum('volume_ambiant'))
                )
                ouv = _D(inv_agg['v_amb'] or 0)
            mvts_qs = Mouvement.objects.filter(
                produit=produit, marketeur=marketeur,
                date_mouvement__range=(periode.date_debut, date_fin_calc),
            )

        mvts = list(mvts_qs)
        ent  = sum(_D(m.volume_ambiant_recu)   for m in mvts if m.type_mouvement == 'ENTREE')
        sor  = sum(_D(m.volume_ambiant_sortie) for m in mvts if m.type_mouvement == 'SORTIE')

        if marketeur is not None:
            # Cessions émises par ce marketeur durant la période -> sortie
            sor += sum(_D(m.cession_volume_ambiant) for m in mvts if m.type_mouvement == 'CESSION')
            # Cessions reçues par ce marketeur durant la période -> entrée
            # (Mouvement.marketeur = l'émetteur de la cession, pas le destinataire,
            # donc requête séparée filtrée sur cession_marketeur_destinataire)
            cess_recues = Mouvement.objects.filter(
                produit=produit, type_mouvement='CESSION',
                cession_marketeur_destinataire=marketeur,
                date_mouvement__range=(periode.date_debut, date_fin_calc),
            )
            ent += sum(_D(m.cession_volume_ambiant) for m in cess_recues)

        cpt  = ouv + ent - sor

        if marketeur is None:
            # TM : Pertes/Gains = jaugeage physique du dépôt - stock comptable
            sf = sf_physique.get(produit.pk)
            pg = (sf - cpt) if sf is not None else None
        else:
            # Marketeur : Pertes/Gains = sa quote-part de coulage (pas le jaugeage du dépôt,
            # qui n'est pas attribuable à un marketeur en particulier)
            qp = qp_coul_par_produit.get(produit.pk)
            if qp is not None:
                pg = qp.get('qp_coul', _Z)
                sf = cpt + pg
            else:
                pg = None
                sf = None
        ratio = (float(pg) / float(sor) * 100) if (pg is not None and sor) else None

        stk_ouv_vals.append(ouv)
        cumul_ent_vals.append(ent)
        cumul_sor_vals.append(sor)
        stk_cpt_vals.append(cpt)
        pertes_gains_vals.append(pg)
        ratios_vals.append(ratio)
        stk_ferm_vals.append(sf)

    return {
        'periode':       periode,
        'produits':      produits,
        'date_ouv':      periode.date_debut,
        'date_ferm':     date_fin_calc,
        'date_filtre':   date_fin_override,
        'stk_ouverture': stk_ouv_vals,
        'cumul_entrees': cumul_ent_vals,
        'cumul_sorties': cumul_sor_vals,
        'stk_comptable': stk_cpt_vals,
        'pertes_gains':  pertes_gains_vals,
        'ratios':        ratios_vals,
        'stk_fermeture': stk_ferm_vals,
    }


@login_required
def etat_stock_ambiant(request):
    from SGDS.models import Societe, Marketeur as MktModel
    from datetime import datetime
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_dashboard')

    periodes    = _periodes_disponibles()
    periode     = _get_periode(request)
    marketeurs  = list(MktModel.objects.filter(statut='ACTIF').order_by('raison_sociale'))
    marketeur   = _get_marketeur_optionnel(request)
    date_str    = request.GET.get('date', '').strip()
    date_filtre = None
    date_fin_override = None

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                # Fermeture du jour J = mouvements jusqu'à J inclus
                date_fin_override = d
        except ValueError:
            pass

    rapport    = _calculer_stock_ambiant(periode, marketeur, date_fin_override) if periode else {}
    societe    = Societe.get_instance()

    return render(request, 'Etat/mensuel/stock_ambiant.html', {
        'periodes':      periodes,
        'periode':       periode,
        'rapport':       rapport,
        'marketeurs':    marketeurs,
        'marketeur_sel': marketeur,
        'societe':       societe,
        'date_filtre':   date_filtre,
    })


@login_required
def etat_stock_ambiant_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    if request.user.is_marketeur_role:
        return redirect('client_dashboard')

    periode   = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_stock_ambiant')
    marketeur = _get_marketeur_optionnel(request)
    rapport   = _calculer_stock_ambiant(periode, marketeur)
    societe   = Societe.get_instance()
    contenu   = _xlsx_stock_ambiant(rapport, societe)

    suffix = f"_{marketeur.sigle or 'mkt'}" if marketeur else "_TM"
    nom = f"stock_ambiant_{periode.mois:02d}_{periode.annee}{suffix}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


@marketeur_required
def etat_stock_ambiant_marketeur(request):
    from SGDS.models import Societe
    from datetime import datetime
    periodes    = _periodes_disponibles()
    periode     = _get_periode(request)
    mkt         = request.user.marketeur
    date_str    = request.GET.get('date', '').strip()
    date_filtre = None
    date_fin_override = None

    if date_str and periode:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if periode.date_debut <= d <= periode.date_fin:
                date_filtre = d
                date_fin_override = d
        except ValueError:
            pass

    rapport  = _calculer_stock_ambiant(periode, mkt, date_fin_override) if periode else {}
    societe  = Societe.get_instance()

    return render(request, 'Espace_Marketeur/mensuel/stock_ambiant.html', {
        'periodes':    periodes,
        'periode':     periode,
        'rapport':     rapport,
        'societe':     societe,
        'marketeur':   mkt,
        'date_filtre': date_filtre,
    })


@marketeur_required
def etat_stock_ambiant_marketeur_export(request):
    from SGDS.models import Societe
    from django.http import HttpResponse
    periode = _get_periode(request)
    if not periode:
        return redirect('client_mensuel_stock_ambiant')
    mkt     = request.user.marketeur
    rapport = _calculer_stock_ambiant(periode, mkt)
    societe = Societe.get_instance()
    contenu = _xlsx_stock_ambiant(rapport, societe)

    nom = f"stock_ambiant_{mkt.sigle or 'mkt'}_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response


def _xlsx_stock_ambiant(rapport, societe):
    """
    Export pivot Stock Ambiant : lignes = statistiques, colonnes = produits.
    Correspond au format exact du classeur Excel STOCK AMB.
    """
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY   = PatternFill(fill_type='solid', fgColor='1B3A6B')
    BLUE2  = PatternFill(fill_type='solid', fgColor='BDD7EE')
    OUV    = PatternFill(fill_type='solid', fgColor='EEF2FF')
    ENT    = PatternFill(fill_type='solid', fgColor='F0FDF4')
    SOR    = PatternFill(fill_type='solid', fgColor='FFF7ED')
    CPT    = PatternFill(fill_type='solid', fgColor='EFF6FF')
    PG     = PatternFill(fill_type='solid', fgColor='FEF2F2')
    RAT    = PatternFill(fill_type='solid', fgColor='FAFAFA')
    FERM   = PatternFill(fill_type='solid', fgColor='F0FDF4')
    WHITE  = Font(bold=True, color='FFFFFF')
    NAVY_F = Font(bold=True, color='1B3A6B')
    RED_F  = Font(bold=True, color='DC2626')
    GRN_F  = Font(bold=True, color='16A34A')
    BOLD   = Font(bold=True)
    CENTER = Alignment(horizontal='center', vertical='center')
    RIGHT  = Alignment(horizontal='right', vertical='center')
    LEFT   = Alignment(horizontal='left', vertical='center')

    periode  = rapport.get('periode')
    produits = rapport.get('produits', [])
    n_prod   = len(produits)
    n_cols   = 1 + n_prod  # label col + one col per product

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Ambiant"

    last_col = get_column_letter(n_cols)

    # Row 1: company name
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = societe.raison_sociale if societe else 'SGDS'
    c.font = Font(bold=True, size=14, color='1B3A6B')
    c.alignment = CENTER

    # Row 2: title
    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']
    c.value = f"ETAT STOCK AMBIANT - {periode}"
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    # Row 3: blank
    row = 4

    # Header row
    c = ws.cell(row=row, column=1, value='DESIGNATION')
    c.font = WHITE; c.fill = NAVY; c.alignment = CENTER
    ws.column_dimensions['A'].width = 28

    for i, prod in enumerate(produits, 2):
        c = ws.cell(row=row, column=i, value=prod.nom.upper())
        c.font = WHITE; c.fill = NAVY; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(i)].width = 18

    row += 1

    def _f(v):
        return float(v) if v is not None else None

    STAT_ROWS = [
        ('STOCKS OUVERTURE (L)',  'stk_ouverture', OUV,  NAVY_F),
        ('CUMULS ENTREES (L)',    'cumul_entrees',  ENT,  BOLD),
        ('CUMULS SORTIES (L)',    'cumul_sorties',  SOR,  BOLD),
        ('STOCKS COMPTABLE (L)',  'stk_comptable',  CPT,  NAVY_F),
        ('PERTES OU GAINS (L)',   'pertes_gains',   PG,   None),
        ('RATIO (%)',             'ratios',         RAT,  None),
        ('STOCKS FERMETURE (L)',  'stk_fermeture',  FERM, Font(bold=True, color='16A34A')),
    ]

    for label, key, fill, label_font in STAT_ROWS:
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill
        c.alignment = LEFT
        if label_font:
            c.font = label_font
        else:
            c.font = BOLD

        vals = rapport.get(key, [])
        for i, v in enumerate(vals, 2):
            num = _f(v)
            c2 = ws.cell(row=row, column=i)
            c2.fill = fill
            c2.alignment = RIGHT
            if num is None:
                c2.value = ''
            elif key == 'ratios':
                c2.value = num
                c2.number_format = '0.0000'
                if num < 0:
                    c2.font = RED_F
                elif num > 0:
                    c2.font = Font(bold=True, color='16A34A')
            elif key == 'pertes_gains':
                c2.value = num
                c2.number_format = '#,##0.00'
                if num < 0:
                    c2.font = RED_F
                elif num > 0:
                    c2.font = Font(bold=True, color='16A34A')
            else:
                c2.value = num
                c2.number_format = '#,##0.00'
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═════════════════════════════════════════════════════════════
#  ÉTAT 7 — FRAIS DE PASSAGE MENSUEL  (admin + marketeur)
# ═════════════════════════════════════════════════════════════

def _filtrer_rapport_frais_marketeur(rapport_complet, mkt):
    """Filtre un rapport frais_passage pour ne garder que les lignes d'un marketeur."""
    modes_filtres = []
    for mode_data in rapport_complet.get('modes', []):
        lignes_mkt = [l for l in mode_data['lignes'] if l['marketeur'].pk == mkt.pk]
        if lignes_mkt:
            modes_filtres.append({**mode_data, 'lignes': lignes_mkt})

    produits  = rapport_complet.get('produits', [])
    tot_vols  = {p.id: _Z for p in produits}
    tot_vol_g = _Z
    tot_mont  = _Z
    for md in modes_filtres:
        for l in md['lignes']:
            for pid, v in l['volumes_par_produit'].items():
                tot_vols[pid] = tot_vols.get(pid, _Z) + _D(v)
            tot_vol_g += _D(l['volume_global'])
            tot_mont  += _D(l['montant'])

    return {
        **rapport_complet,
        'modes': modes_filtres,
        'total_general': {
            'volumes_par_produit': tot_vols,
            'volume_global':       tot_vol_g,
            'montant':             tot_mont,
        },
    }


@login_required
def etat_frais_passage_mensuel(request):
    from SGDS.models import Societe
    from SGDS.services.frais_passage import calculer_frais_passage
    if request.user.is_marketeur_role:
        messages.error(request, "Accès non autorisé.")
        return redirect('client_dashboard')

    periodes = _periodes_disponibles()
    periode  = _get_periode(request)
    societe  = Societe.get_instance()
    rapport  = calculer_frais_passage(periode) if periode else None

    return render(request, 'Etat/mensuel/frais_passage.html', {
        'periodes': periodes,
        'periode':  periode,
        'rapport':  rapport,
        'societe':  societe,
    })


@login_required
def etat_frais_passage_mensuel_export(request):
    from SGDS.models import Societe
    from SGDS.services.frais_passage import calculer_frais_passage
    from SGDS.services.export_excel import exporter_frais_passage_xlsx
    from django.http import HttpResponse
    if request.user.is_marketeur_role:
        return redirect('client_dashboard')

    periode = _get_periode(request)
    if not periode:
        messages.warning(request, "Aucune période disponible.")
        return redirect('etat_mensuel_frais_passage')

    rapport = calculer_frais_passage(periode)
    contenu = exporter_frais_passage_xlsx(rapport)

    nom = f"frais_passage_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response



@marketeur_required
def etat_frais_passage_mensuel_marketeur(request):
    from SGDS.models import Societe
    from SGDS.services.frais_passage import calculer_frais_passage
    periodes        = _periodes_disponibles()
    periode         = _get_periode(request)
    societe         = Societe.get_instance()
    mkt             = request.user.marketeur
    rapport         = None

    if periode:
        rapport_complet = calculer_frais_passage(periode)
        rapport         = _filtrer_rapport_frais_marketeur(rapport_complet, mkt)

    return render(request, 'Espace_Marketeur/mensuel/frais_passage.html', {
        'periodes':  periodes,
        'periode':   periode,
        'rapport':   rapport,
        'societe':   societe,
        'marketeur': mkt,
    })


@marketeur_required
def etat_frais_passage_mensuel_marketeur_export(request):
    from SGDS.models import Societe
    from SGDS.services.frais_passage import calculer_frais_passage
    from SGDS.services.export_excel import exporter_frais_passage_xlsx
    from django.http import HttpResponse

    periode = _get_periode(request)
    if not periode:
        return redirect('client_mensuel_frais_passage')

    mkt             = request.user.marketeur
    rapport_complet = calculer_frais_passage(periode)
    rapport         = _filtrer_rapport_frais_marketeur(rapport_complet, mkt)
    contenu         = exporter_frais_passage_xlsx(rapport)

    nom = f"frais_passage_{mkt.sigle or 'mkt'}_{periode.mois:02d}_{periode.annee}.xlsx"
    response = HttpResponse(
        content=contenu,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nom}"'
    return response
