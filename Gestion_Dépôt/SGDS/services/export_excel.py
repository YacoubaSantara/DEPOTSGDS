"""
Exports Excel fidèles aux templates SGDS.
Utilise openpyxl. Entièrement générique sur les produits.
"""
from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles communs ────────────────────────────────────────────
_THIN   = Side(border_style='thin', color='A0A0A0')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENT   = Alignment(horizontal='center', vertical='center', wrap_text=True)
_RIGHT  = Alignment(horizontal='right',  vertical='center')
_LEFT   = Alignment(horizontal='left',   vertical='center')

def _fh(bold=True, size=10, color='FFFFFF'):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def _fb(bold=False, size=9):
    return Font(name='Calibri', bold=bold, size=size)

def _fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

_FILL_NAVY  = _fill('1E3A5F')
_FILL_NAVY2 = _fill('2d4e7a')
_FILL_TOTAL = _fill('dde6f0')
_FILL_JAU   = _fill('FFF4CE')  # fond jaugeage (suivi évolution)
_FILL_GROUP = _fill('1E3A5F')  # sous-titre mode (frais passage)

FMT_VOL  = '#,##0;[Red](#,##0);"-"'
FMT_COEF = '0.00000000'
FMT_PU   = '0.0000'
FMT_MNT  = '#,##0.00;[Red](#,##0.00);"-"'


def _style(cell, font=None, fill=None, align=None, border=True, fmt=None):
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = _BORDER
    if fmt:    cell.number_format = fmt


# ─────────────────────────────────────────────────────────────
#  1. COULAGE RÉPARTITION
# ─────────────────────────────────────────────────────────────
def exporter_coulage_xlsx(rapport) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'COULA REPAR'

    periode  = rapport['periode']
    produits = rapport['produits']
    nb_prod  = len(produits)
    # Colonnes : A code, B marketeur, puis par produit 7 cols, puis 4 cols (vol, motif, pu, mnt)
    nb_cols = 2 + nb_prod * 7 + 4

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    tc = ws['A1']
    tc.value = f"TABLEAU RÉPARTITION COULAGE — {periode.libelle.upper()}"
    tc.font  = Font(name='Calibri', bold=True, size=14, color='1E3A5F')
    tc.alignment = _CENT
    ws.row_dimensions[1].height = 28

    # Ligne 3 : groupes
    for c_idx in [1, 2]:
        ws.merge_cells(start_row=3, start_column=c_idx, end_row=4, end_column=c_idx)
    ws.cell(row=3, column=1, value='CODE')
    ws.cell(row=3, column=2, value='MARKETEUR')

    col = 3
    for p in produits:
        ws.merge_cells(start_row=3, start_column=col,   end_row=3, end_column=col+2)
        ws.merge_cells(start_row=3, start_column=col+3, end_row=4, end_column=col+3)
        ws.merge_cells(start_row=3, start_column=col+4, end_row=3, end_column=col+6)
        ws.cell(row=3, column=col,   value=f"ENTRÉE {p.code}")
        ws.cell(row=3, column=col+3, value=f"SORTIE {p.code}")
        ws.cell(row=3, column=col+4, value=f"RÉPART. COULAGE {p.code}")
        for i, txt in enumerate(['BRUT', 'COUL', 'NET']):
            ws.cell(row=4, column=col+i, value=txt)
        for i, txt in enumerate(['BASE QP', 'COEF', 'QP COUL']):
            ws.cell(row=4, column=col+4+i, value=txt)
        col += 7

    for lbl, span in [('VOL. GLOBAL', 1), ('MOTIF', 1), ('PU/L', 1), ('MONTANT', 1)]:
        ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        ws.cell(row=3, column=col, value=lbl)
        col += 1

    # Styles en-têtes
    for r in [3, 4]:
        for c_idx in range(1, nb_cols + 1):
            cell = ws.cell(row=r, column=c_idx)
            _style(cell, font=_fh(), fill=_FILL_NAVY, align=_CENT)
    ws.row_dimensions[3].height = 28
    ws.row_dimensions[4].height = 22

    # Données
    row = 5
    for ligne in rapport['lignes']:
        mkt = ligne['marketeur']
        ws.cell(row=row, column=1, value=f"MARK{mkt.pk:03d}")
        ws.cell(row=row, column=2, value=mkt.raison_sociale)
        _style(ws.cell(row=row, column=1), font=_fb(), align=_LEFT)
        _style(ws.cell(row=row, column=2), font=_fb(), align=_LEFT)

        col = 3
        for p in produits:
            pp = ligne['par_produit'].get(p.pk, {})
            vals = [
                float(pp.get('brut_entree', 0)),
                float(pp.get('coul_entree', 0)),
                float(pp.get('entree_nette', 0)),
                float(pp.get('sortie', 0)),
                float(pp.get('base_qp_coul', 0)),
                float(pp.get('coef_qp_coul', 0)),
                float(pp.get('qp_coul', 0)),
            ]
            fmts = [FMT_VOL]*5 + [FMT_COEF, FMT_VOL]
            for i, (v, fmt) in enumerate(zip(vals, fmts)):
                _style(ws.cell(row=row, column=col+i, value=v),
                       font=_fb(), align=_RIGHT, fmt=fmt)
            col += 7

        vol_g  = float(ligne['volume_global_sorti'])
        montant = float(ligne['montant'])
        for v, fmt, al in [
            (vol_g,    FMT_VOL, _RIGHT),
            (ligne['motif'], None, _LEFT),
            (float(ligne['prix_unitaire']), FMT_PU, _RIGHT),
            (montant,  FMT_MNT, _RIGHT),
        ]:
            c = ws.cell(row=row, column=col, value=v)
            _style(c, font=_fb(), align=al, fmt=fmt)
            col += 1
        row += 1

    # Ligne totaux
    totaux = rapport['totaux']
    ws.cell(row=row, column=1, value='TOTAUX')
    ws.cell(row=row, column=2, value='ENTRÉE / SORTIE')
    col = 3
    for p in produits:
        tp  = totaux['par_produit'].get(p.pk, {})
        coef = rapport['coefficients'].get(p.pk, Decimal('0'))
        vals = [
            float(tp.get('brut_entree', 0)),
            float(tp.get('coul_entree', 0)),
            float(tp.get('entree_nette', 0)),
            float(tp.get('sortie', 0)),
            float(tp.get('base_qp_coul', 0)),
            float(coef),
            float(tp.get('qp_coul', 0)),
        ]
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=col+i, value=v)
            _style(c, font=_fh(color='1E3A5F'), fill=_FILL_TOTAL, align=_RIGHT,
                   fmt=FMT_COEF if i == 5 else FMT_VOL)
        col += 7

    for v, fmt in [
        (float(totaux['volume_global_sorti']), FMT_VOL),
        (totaux['motif'], None),
        (float(totaux['prix_unitaire']), FMT_PU),
        (float(totaux['montant']), FMT_MNT),
    ]:
        c = ws.cell(row=row, column=col, value=v)
        _style(c, font=_fh(color='1E3A5F'), fill=_FILL_TOTAL, align=_RIGHT, fmt=fmt)
        col += 1

    # Largeurs
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    for c_idx in range(3, nb_cols + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 13
    ws.freeze_panes = 'C5'

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────
#  2. SUIVI ÉVOLUTION
# ─────────────────────────────────────────────────────────────
def exporter_suivi_xlsx(rapport) -> bytes:
    wb  = Workbook()
    ws  = wb.active
    ws.title = f"SUIVI {rapport['produit'].code}"[:31]

    produit = rapport['produit']
    cuves   = rapport['cuves']
    nc      = len(cuves)

    GROUPES = [
        ('STOCK INITIAL',   nc),
        ('ENTRÉE BRUTE',    nc),
        ('COULAGE RÉCEPT.', nc),
        ('SORTIE',          nc),
        ('STOCK COMPTABLE', nc),
        ('STOCK PHYSIQUE',  nc),
    ]
    nb_cols = 2 + sum(n for _, n in GROUPES) + 1  # date, jour, groupes, P/G cumul

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    tc = ws['A1']
    tc.value = (f"SUIVI ÉVOLUTION COULAGE {produit.nom.upper()} "
                f"— {rapport['periode'].libelle.upper()}")
    tc.font  = Font(name='Calibri', bold=True, size=13, color='1E3A5F')
    tc.alignment = _CENT

    # En-têtes groupe (ligne 3) + cuve (ligne 4)
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    ws.cell(row=3, column=1, value='DATE')
    ws.cell(row=3, column=2, value='JOUR')

    col = 3
    for label, n in GROUPES:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+n-1)
        ws.cell(row=3, column=col, value=label)
        for i, cuve in enumerate(cuves):
            ws.cell(row=4, column=col+i, value=cuve.numero)
        col += n

    ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
    ws.cell(row=3, column=col, value='P/G CUMULÉ')

    for r in [3, 4]:
        for c_idx in range(1, nb_cols+1):
            _style(ws.cell(row=r, column=c_idx), font=_fh(), fill=_FILL_NAVY, align=_CENT)
    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 20

    # Données
    row = 5
    for jour in rapport['jours']:
        fill_row = _FILL_JAU if jour['a_jaugeage'] else None
        ws.cell(row=row, column=1, value=jour['date'].strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=jour['jour_semaine'])
        _style(ws.cell(row=row, column=1), font=_fb(), fill=fill_row, align=_CENT)
        _style(ws.cell(row=row, column=2), font=_fb(), fill=fill_row, align=_CENT)

        col = 3
        srcs = [
            jour['stock_initial'],
            jour['entree_brute'],
            jour['coulage_reception'],
            jour['sortie'],
            jour['stock_comptable'],
            jour['stock_physique'],
        ]
        for src in srcs:
            for cuve in cuves:
                v = src.get(cuve.id)
                c = ws.cell(row=row, column=col,
                             value=float(v) if v is not None else None)
                _style(c, font=_fb(), fill=fill_row, align=_RIGHT, fmt=FMT_VOL)
                col += 1

        pg = jour['pg_cumul']
        c = ws.cell(row=row, column=col, value=float(pg))
        c.font = Font(name='Calibri', size=9, bold=True,
                      color='C00000' if pg < 0 else '16a34a')
        _style(c, fill=fill_row, align=_RIGHT, fmt=FMT_VOL)
        row += 1

    # Totaux
    t = rapport['totaux']
    ws.cell(row=row, column=1, value='TOTAUX')
    col = 3
    for src_key, n in zip(
        ['entree_brute', 'coulage_reception', 'sortie'],
        [nc, nc, nc]
    ):
        pass  # Totaux par groupe pas nécessaires dans xlsx simplifié
    c = ws.cell(row=row, column=col + nc*4, value=float(t['pg_total']))
    _style(c, font=_fh(color='1E3A5F'), fill=_FILL_TOTAL, align=_RIGHT, fmt=FMT_VOL)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    for c_idx in range(3, nb_cols+1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 11
    ws.freeze_panes = 'C5'

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ─────────────────────────────────────────────────────────────
#  3. FRAIS DE PASSAGE
# ─────────────────────────────────────────────────────────────
def exporter_frais_passage_xlsx(rapport) -> bytes:
    wb  = Workbook()
    ws  = wb.active
    ws.title = 'FRAIS PASSAGE'

    produits = rapport['produits']
    np       = len(produits)
    # cols: marketeur, mode, [np produits], vol global, motif, PU, montant
    nb_cols  = 2 + np + 4

    # Titres
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
    ws['A1'].value = 'DÉPÔT SANKÉ SGDS — DÉPÔT DE DROIT'
    ws['A1'].font  = Font(name='Calibri', bold=True, size=14, color='1E3A5F')
    ws['A1'].alignment = _CENT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nb_cols)
    p = rapport['periode']
    ws['A2'].value = (f"FRAIS DE PASSAGE SGDS "
                      f"DU {p.date_debut.strftime('%d-%m-%Y')} "
                      f"AU {p.date_fin.strftime('%d-%m-%Y')}")
    ws['A2'].font  = Font(name='Calibri', bold=True, size=12)
    ws['A2'].alignment = _CENT
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20

    row = 4
    for mode_data in rapport['modes']:
        # Sous-titre mode
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nb_cols)
        sc = ws.cell(row=row, column=1,
                     value=f"FRAIS DE PASSAGE — {mode_data['mode_libelle'].upper()}")
        _style(sc, font=_fh(size=11), fill=_FILL_NAVY, align=_CENT)
        ws.row_dimensions[row].height = 22
        row += 1

        # En-têtes colonnes
        ws.cell(row=row, column=1, value='MARKETEUR')
        ws.cell(row=row, column=2, value='MODE')
        for i, p in enumerate(produits):
            ws.cell(row=row, column=3+i, value=p.code)
        ws.cell(row=row, column=3+np,   value='VOL. GLOBAL')
        ws.cell(row=row, column=4+np,   value='MOTIF')
        ws.cell(row=row, column=5+np,   value='PU/L')
        ws.cell(row=row, column=6+np,   value='MONTANT')
        for c_idx in range(1, nb_cols+1):
            _style(ws.cell(row=row, column=c_idx), font=_fh(), fill=_FILL_NAVY2, align=_CENT)
        row += 1

        for ligne in mode_data['lignes']:
            ws.cell(row=row, column=1, value=ligne['marketeur'].raison_sociale)
            ws.cell(row=row, column=2, value=mode_data['mode_libelle'])
            for i, pr in enumerate(produits):
                v = ligne['volumes_par_produit'].get(pr.id, 0)
                c = ws.cell(row=row, column=3+i, value=float(v))
                _style(c, font=_fb(), align=_RIGHT, fmt=FMT_VOL)
            for v, col_off, fmt in [
                (float(ligne['volume_global']), 3+np, FMT_VOL),
                (ligne['motif'],               4+np, None),
                (float(ligne['pu']),            5+np, FMT_PU),
                (float(ligne['montant']),       6+np, FMT_MNT),
            ]:
                c = ws.cell(row=row, column=col_off, value=v)
                _style(c, font=_fb(), align=_RIGHT if fmt else _LEFT, fmt=fmt)
            _style(ws.cell(row=row, column=1), font=_fb(), align=_LEFT)
            _style(ws.cell(row=row, column=2), font=_fb(), align=_LEFT)
            row += 1

        # Sous-total
        st = mode_data['sous_totaux']
        ws.cell(row=row, column=1, value=f"TOTAL {mode_data['mode_libelle'].upper()}")
        for i, pr in enumerate(produits):
            c = ws.cell(row=row, column=3+i,
                        value=float(st['volumes_par_produit'].get(pr.id, 0)))
            _style(c, font=_fh(color='1E3A5F'), fill=_FILL_TOTAL, align=_RIGHT, fmt=FMT_VOL)
        for v, col_off, fmt in [
            (float(st['volume_global']), 3+np, FMT_VOL),
            (float(st['montant']),       6+np, FMT_MNT),
        ]:
            c = ws.cell(row=row, column=col_off, value=v)
            _style(c, font=_fh(color='1E3A5F'), fill=_FILL_TOTAL, align=_RIGHT, fmt=fmt)
        for c_idx in [1, 2, 4+np, 5+np]:
            _style(ws.cell(row=row, column=c_idx), font=_fh(color='1E3A5F'), fill=_FILL_TOTAL)
        row += 2

    # Total général
    tg = rapport['total_general']
    ws.cell(row=row, column=1, value=f"TOTAL PASSAGE {rapport['periode'].libelle.upper()}")
    for i, pr in enumerate(produits):
        c = ws.cell(row=row, column=3+i,
                    value=float(tg['volumes_par_produit'].get(pr.id, 0)))
        _style(c, font=_fh(), fill=_FILL_NAVY, align=_RIGHT, fmt=FMT_VOL)
    for v, col_off, fmt in [
        (float(tg['volume_global']), 3+np, FMT_VOL),
        (float(tg['montant']),       6+np, FMT_MNT),
    ]:
        c = ws.cell(row=row, column=col_off, value=v)
        _style(c, font=_fh(), fill=_FILL_NAVY, align=_RIGHT, fmt=fmt)
    for c_idx in [1, 2, 4+np, 5+np]:
        _style(ws.cell(row=row, column=c_idx), font=_fh(), fill=_FILL_NAVY)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    for c_idx in range(3, nb_cols+1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 14

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
