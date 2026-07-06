"""
Génération et envoi des états mensuels marketeur à la clôture d'une période —
Stock Ouverture, Stock Fermeture, Coulage Répartition, Stock à 15°, Stock
Ambiant, Frais de passage. Chaque état est envoyé en pièce jointe séparée
(1 PDF + 1 Excel par état, donc 12 fichiers), pas fusionné dans un classeur
ou un PDF unique.

Les mouvements ne font pas partie de cet envoi : ils sont envoyés à l'unité,
immédiatement, via SGDS/signals.py::_envoyer_email_mouvement.
"""
from datetime import timedelta
from io import BytesIO


# ─────────────────────────────────────────────────────────────
#  1. STOCK OUVERTURE
# ─────────────────────────────────────────────────────────────
def _xlsx_stock_ouverture_marketeur(periode, marketeur, lignes, societe) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    NAVY = PatternFill(fill_type='solid', fgColor='1E3A5F')
    HDR  = PatternFill(fill_type='solid', fgColor='F0F4F8')
    SD   = PatternFill(fill_type='solid', fgColor='FEF9EE')
    AC   = PatternFill(fill_type='solid', fgColor='F0F9F4')
    SUBT = PatternFill(fill_type='solid', fgColor='EEF2FF')
    TOT  = PatternFill(fill_type='solid', fgColor='E8F0FE')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Ouverture'

    ws.merge_cells('A1:C1')
    ws['A1'] = societe.raison_sociale if societe else 'SGDS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = f"STOCK D'OUVERTURE — {periode} — {marketeur.sigle or marketeur.raison_sociale}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    for ci, (h, w) in enumerate([('Désignation', 38), ('V AMB (L)', 16), ('V @15°C (L)', 16)], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = NAVY
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    row += 1

    def _r(label, amb, c15, fill=None, bold=False, indent=0):
        nonlocal row
        c = ws.cell(row=row, column=1, value=('  ' * indent) + label)
        if bold: c.font = Font(bold=True)
        if fill: c.fill = fill
        for ci, v in [(2, amb), (3, c15)]:
            cell = ws.cell(row=row, column=ci, value=float(v) if v is not None else None)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            if fill: cell.fill = fill
            if bold: cell.font = Font(bold=True)
        row += 1

    for l in lignes:
        c = ws.cell(row=row, column=1, value=f"═══ {l['produit'].nom.upper()} ═══")
        c.font = Font(bold=True, color='1E3A5F')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c.fill = HDR
        row += 1

        _r('STOCK OUVERTURE', None, None, HDR, True)
        _r('Sous douane (SD)', l['ouv_sd_amb'], l['ouv_sd_15c'], SD, indent=1)
        _r('Acquitté (AC)', l['ouv_ac_amb'], l['ouv_ac_15c'], AC, indent=1)
        _r('Total Stock Ouverture', l['stock_debut_amb'], l['stock_debut_15c'], TOT, True)

        _r('ENTRÉES', None, None, HDR, True)
        _r('Réception CC (SD)', l['rec_sd_amb'], l['rec_sd_15c'], indent=1)
        _r('Réception CC (AC)', l['rec_ac_amb'], l['rec_ac_15c'], indent=1)
        _r('Cession reçue (SD)', l['cess_recues_sd_amb'], l['cess_recues_sd_15c'], indent=1)
        _r('Cession reçue (AC)', l['cess_recues_ac_amb'], l['cess_recues_ac_15c'], indent=1)
        _r('Reclassement (SD)', l['recl_sd_entree_amb'], l['recl_sd_entree_15c'], indent=1)
        _r('Reclassement (AC)', l['recl_ac_entree_amb'], l['recl_ac_entree_15c'], indent=1)
        _r('Total Entrées (SD)', l['total_entrees_sd_amb'], l['total_entrees_sd_15c'], SUBT, True, indent=1)
        _r('Total Entrées (AC)', l['total_entrees_ac_amb'], l['total_entrees_ac_15c'], SUBT, True, indent=1)
        _r('Total Entrées', l['total_entrees_amb'], l['total_entrees_15c'], TOT, True)

        _r('SORTIES', None, None, HDR, True)
        _r('CC Acquitté (AC)', l['livr_ac_amb'], l['livr_ac_15c'], indent=1)
        _r('Livraison (SD)', l['livr_sd_amb'], l['livr_sd_15c'], indent=1)
        _r('Cession émise (SD)', l['cess_emises_sd_amb'], l['cess_emises_sd_15c'], indent=1)
        _r('Cession émise (AC)', l['cess_emises_ac_amb'], l['cess_emises_ac_15c'], indent=1)
        _r('Reclassement (SD)', l['recl_sd_sortie_amb'], l['recl_sd_sortie_15c'], indent=1)
        _r('Reclassement (AC)', l['recl_ac_sortie_amb'], l['recl_ac_sortie_15c'], indent=1)
        _r('Total Sorties (SD)', l['total_sorties_sd_amb'], l['total_sorties_sd_15c'], SUBT, True, indent=1)
        _r('Total Sorties (AC)', l['total_sorties_ac_amb'], l['total_sorties_ac_15c'], SUBT, True, indent=1)
        _r('Total Sorties', l['total_sorties_amb'], l['total_sorties_15c'], TOT, True)

        _r('STOCK COMPTABLE', None, None, HDR, True)
        _r('Stock Comptable (SD)', l['stk_c_sd_amb'], l['stk_c_sd_15c'], SD, indent=1)
        _r('Stock Comptable (AC)', l['stk_c_ac_amb'], l['stk_c_ac_15c'], AC, indent=1)
        _r('Total Stock Comptable', l['stock_fin_comptable_amb'], l['stock_fin_comptable_15c'], TOT, True)
        row += 1

    if not lignes:
        ws.cell(row=row, column=1, value="Aucune activité sur la période pour ce marketeur.")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def generer_pdf_stock_ouverture(periode, marketeur, lignes, societe) -> bytes:
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf_bytes
    return render_to_pdf_bytes('Email/etat_stock_ouverture_pdf.html', {
        'societe': societe, 'marketeur': marketeur, 'periode': periode,
        'lignes': lignes, 'generated_at': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    })


# ─────────────────────────────────────────────────────────────
#  2. STOCK FERMETURE
# ─────────────────────────────────────────────────────────────
def _xlsx_stock_fermeture_marketeur(periode, marketeur, lignes, societe) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    NAVY = PatternFill(fill_type='solid', fgColor='1E3A5F')
    HDR  = PatternFill(fill_type='solid', fgColor='F0F4F8')
    SD   = PatternFill(fill_type='solid', fgColor='FEF9EE')
    AC   = PatternFill(fill_type='solid', fgColor='F0F9F4')
    SUBT = PatternFill(fill_type='solid', fgColor='EEF2FF')
    TOT  = PatternFill(fill_type='solid', fgColor='E8F0FE')
    FIN  = PatternFill(fill_type='solid', fgColor='FFF9C4')
    PERTE = PatternFill(fill_type='solid', fgColor='FCE8E6')
    GAIN  = PatternFill(fill_type='solid', fgColor='E8F5E9')
    CLOT  = PatternFill(fill_type='solid', fgColor='E0ECFF')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Stock Fermeture'

    ws.merge_cells('A1:C1')
    ws['A1'] = societe.raison_sociale if societe else 'SGDS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:C2')
    ws['A2'] = f"STOCK DE FERMETURE — {periode} — {marketeur.sigle or marketeur.raison_sociale}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    for ci, (h, w) in enumerate([('Désignation', 38), ('V AMB (L)', 16), ('V @15°C (L)', 16)], 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = NAVY
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(ci)].width = w
    row += 1

    def _r(label, amb, c15, fill=None, bold=False, indent=0):
        nonlocal row
        c = ws.cell(row=row, column=1, value=('  ' * indent) + label)
        if bold: c.font = Font(bold=True)
        if fill: c.fill = fill
        for ci, v in [(2, amb), (3, c15)]:
            cell = ws.cell(row=row, column=ci, value=float(v) if v is not None else None)
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            if fill: cell.fill = fill
            if bold: cell.font = Font(bold=True)
        row += 1

    for l in lignes:
        c = ws.cell(row=row, column=1, value=f"═══ {l['produit'].nom.upper()} ═══")
        c.font = Font(bold=True, color='1E3A5F')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c.fill = HDR
        row += 1

        _r('STOCK OUVERTURE', None, None, HDR, True)
        _r('Sous douane (SD)', l['ouv_sd_amb'], l['ouv_sd_15c'], SD, indent=1)
        _r('Acquitté (AC)', l['ouv_ac_amb'], l['ouv_ac_15c'], AC, indent=1)
        _r('Total Stock Ouverture', l['stock_debut_amb'], l['stock_debut_15c'], TOT, True)

        _r('ENTRÉES', None, None, HDR, True)
        _r('Total Entrées (SD)', l['total_entrees_sd_amb'], l['total_entrees_sd_15c'], SUBT, True, indent=1)
        _r('Total Entrées (AC)', l['total_entrees_ac_amb'], l['total_entrees_ac_15c'], SUBT, True, indent=1)
        _r('Total Entrées', l['total_entrees_amb'], l['total_entrees_15c'], TOT, True)

        _r('SORTIES', None, None, HDR, True)
        _r('Total Sorties (SD)', l['total_sorties_sd_amb'], l['total_sorties_sd_15c'], SUBT, True, indent=1)
        _r('Total Sorties (AC)', l['total_sorties_ac_amb'], l['total_sorties_ac_15c'], SUBT, True, indent=1)
        _r('Total Sorties', l['total_sorties_amb'], l['total_sorties_15c'], TOT, True)

        _r('STOCK COMPTABLE', None, None, HDR, True)
        _r('Stock Comptable (SD)', l['stk_c_sd_amb'], l['stk_c_sd_15c'], SD, indent=1)
        _r('Stock Comptable (AC)', l['stk_c_ac_amb'], l['stk_c_ac_15c'], AC, indent=1)
        _r('Total Stock Comptable (fermeture)', l['stock_fin_comptable_amb'], l['stock_fin_comptable_15c'], FIN, True)

        _r('QUOTE-PART P/G INSTALLATION', None, None, HDR, True)
        pg_fill = GAIN if (l['pg_inst_amb'] or 0) >= 0 else PERTE
        _r('Quote-part P/G Installation (AC)', l['pg_inst_amb'], None, pg_fill, indent=1)

        _r('RATIO', None, None, HDR, True)
        _r('Ratio P/G / Sorties Acquittées (dépôt)', l['ratio'], None, indent=1)

        _r('STOCKS CLÔTURE', None, None, HDR, True)
        _r('Stock Clôture (SD)', l['cloture_sd_amb'], l['cloture_sd_15c'], SD, indent=1)
        _r('Stock Clôture (AC)', l['cloture_ac_amb'], l['cloture_ac_15c'], AC, indent=1)
        _r('Total Stock Clôture', l['cloture_total_amb'], l['cloture_total_15c'], CLOT, True)
        row += 1

    if not lignes:
        ws.cell(row=row, column=1, value="Aucune activité sur la période pour ce marketeur.")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()


def generer_pdf_stock_fermeture(periode, marketeur, lignes, societe) -> bytes:
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf_bytes
    return render_to_pdf_bytes('Email/etat_stock_fermeture_pdf.html', {
        'societe': societe, 'marketeur': marketeur, 'periode': periode,
        'lignes': lignes, 'generated_at': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    })


# ─────────────────────────────────────────────────────────────
#  3. COULAGE RÉPARTITION
# ─────────────────────────────────────────────────────────────
def generer_pdf_coulage_repartition(periode, marketeur, rapport, societe) -> bytes:
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf_bytes

    produits = rapport.get('produits', [])
    ligne = next((l for l in rapport.get('lignes', []) if l['marketeur'].pk == marketeur.pk), None)
    par_produit = ligne.get('par_produit', {}) if ligne else {}
    lignes_par_produit = [
        {'produit': p, **par_produit.get(p.pk, {})}
        for p in produits
    ]
    return render_to_pdf_bytes('Email/etat_coulage_repartition_pdf.html', {
        'societe': societe, 'marketeur': marketeur, 'periode': periode,
        'lignes_par_produit': lignes_par_produit, 'ligne': ligne,
        'generated_at': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    })


# ─────────────────────────────────────────────────────────────
#  4. STOCK À 15°
# ─────────────────────────────────────────────────────────────
def generer_pdf_stock_15(periode, marketeur, rapport, societe) -> bytes:
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf_bytes
    return render_to_pdf_bytes('Email/etat_stock_15_pdf.html', {
        'societe': societe, 'marketeur': marketeur, 'periode': periode,
        'rapport': rapport, 'generated_at': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    })


# ─────────────────────────────────────────────────────────────
#  5. STOCK AMBIANT
# ─────────────────────────────────────────────────────────────
def generer_pdf_stock_ambiant(periode, marketeur, rapport, societe) -> bytes:
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf_bytes
    return render_to_pdf_bytes('Email/etat_stock_ambiant_pdf.html', {
        'societe': societe, 'marketeur': marketeur, 'periode': periode,
        'rapport': rapport, 'generated_at': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    })


# ─────────────────────────────────────────────────────────────
#  6. FRAIS DE PASSAGE
# ─────────────────────────────────────────────────────────────
def generer_pdf_frais_passage(periode, marketeur, rapport, societe) -> bytes:
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf_bytes
    return render_to_pdf_bytes('Email/etat_frais_passage_pdf.html', {
        'societe': societe, 'marketeur': marketeur, 'periode': periode,
        'rapport': rapport, 'generated_at': timezone.now().strftime('%d/%m/%Y à %H:%M'),
    })


# ─────────────────────────────────────────────────────────────
#  ENVOI — 6 états × (1 PDF + 1 Excel) = 12 pièces jointes, 1 seul email
# ─────────────────────────────────────────────────────────────
def _generer_pieces_jointes(periode, marketeur):
    """Retourne la liste [(nom_fichier, bytes, mimetype), ...] des 12 fichiers."""
    from SGDS.models import Societe
    from SGDS.services.frais_passage import calculer_frais_passage
    from SGDS.services.export_excel import exporter_frais_passage_xlsx
    from SGDS.views.mensuel import (
        _calculer_stock_ouverture_fermeture_marketeur,
        _rapport_coulage_pour_periode,
        _xlsx_coulage_repartition,
        _filtrer_rapport_frais_marketeur,
        _calculer_stock_a_15,
        _xlsx_stock_a_15,
        _calculer_stock_ambiant,
        _xlsx_stock_ambiant,
    )

    societe = Societe.get_instance()
    suffixe = f"{marketeur.sigle or marketeur.pk}_{periode.mois:02d}_{periode.annee}"
    XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    pieces = []

    # 1. Stock Ouverture — instantané PUR avant tout mouvement de la période.
    # date_fin_override antérieur à date_debut → plage de filtre des
    # mouvements invalide → 0 mouvement compté, Stock Comptable = Stock
    # Ouverture. Évite que des mouvements datés du 1er (même si réellement
    # exécutés plus tard dans le mois) ne contaminent l'état d'ouverture.
    veille_periode = periode.date_debut - timedelta(days=1)
    lignes_ouverture = _calculer_stock_ouverture_fermeture_marketeur(periode, marketeur, date_fin_override=veille_periode)
    pieces.append((f"Stock_Ouverture_{suffixe}.xlsx", _xlsx_stock_ouverture_marketeur(periode, marketeur, lignes_ouverture, societe), XLSX_MIME))
    pieces.append((f"Stock_Ouverture_{suffixe}.pdf", generer_pdf_stock_ouverture(periode, marketeur, lignes_ouverture, societe), 'application/pdf'))

    # 2. Stock Fermeture — mois complet (entrées/sorties/clôture réelles)
    lignes_stock = _calculer_stock_ouverture_fermeture_marketeur(periode, marketeur)
    pieces.append((f"Stock_Fermeture_{suffixe}.xlsx", _xlsx_stock_fermeture_marketeur(periode, marketeur, lignes_stock, societe), XLSX_MIME))
    pieces.append((f"Stock_Fermeture_{suffixe}.pdf", generer_pdf_stock_fermeture(periode, marketeur, lignes_stock, societe), 'application/pdf'))

    # 3. Coulage Répartition
    rapport_coulage_complet, _source = _rapport_coulage_pour_periode(periode)
    lignes_mkt = [l for l in rapport_coulage_complet.get('lignes', []) if l['marketeur'].pk == marketeur.pk]
    rapport_coulage = {**rapport_coulage_complet, 'lignes': lignes_mkt}
    pieces.append((f"Coulage_Repartition_{suffixe}.xlsx", _xlsx_coulage_repartition(rapport_coulage, societe, marketeur_filtre=marketeur), XLSX_MIME))
    pieces.append((f"Coulage_Repartition_{suffixe}.pdf", generer_pdf_coulage_repartition(periode, marketeur, rapport_coulage, societe), 'application/pdf'))

    # 4. Stock à 15°
    rapport_15 = _calculer_stock_a_15(periode, marketeur)
    pieces.append((f"Stock_15_{suffixe}.xlsx", _xlsx_stock_a_15(rapport_15, societe), XLSX_MIME))
    pieces.append((f"Stock_15_{suffixe}.pdf", generer_pdf_stock_15(periode, marketeur, rapport_15, societe), 'application/pdf'))

    # 5. Stock Ambiant
    rapport_ambiant = _calculer_stock_ambiant(periode, marketeur)
    pieces.append((f"Stock_Ambiant_{suffixe}.xlsx", _xlsx_stock_ambiant(rapport_ambiant, societe), XLSX_MIME))
    pieces.append((f"Stock_Ambiant_{suffixe}.pdf", generer_pdf_stock_ambiant(periode, marketeur, rapport_ambiant, societe), 'application/pdf'))

    # 6. Frais de passage
    rapport_frais_complet = calculer_frais_passage(periode)
    rapport_frais = _filtrer_rapport_frais_marketeur(rapport_frais_complet, marketeur)
    pieces.append((f"Frais_Passage_{suffixe}.xlsx", exporter_frais_passage_xlsx(rapport_frais), XLSX_MIME))
    pieces.append((f"Frais_Passage_{suffixe}.pdf", generer_pdf_frais_passage(periode, marketeur, rapport_frais, societe), 'application/pdf'))

    # Cachet virtuel de conformité + protection sur tous les PDF envoyés
    from SGDS.services.export_pdf import appliquer_cachet_et_protection
    pieces = [
        (nom, appliquer_cachet_et_protection(contenu, societe) if mime == 'application/pdf' else contenu, mime)
        for nom, contenu, mime in pieces
    ]

    return pieces


def envoyer_etat_mensuel_marketeur(periode, marketeur, config, *, declenche_par=None, email_override=None):
    """Génère les 6 états (1 PDF + 1 Excel chacun, 12 fichiers) et les envoie
    en pièces jointes dans un seul email au marketeur. Ne lève jamais —
    retourne toujours un EnvoiEtatMensuel (SUCCES ou ECHEC), pour que l'échec
    d'un marketeur n'empêche jamais l'envoi aux autres."""
    from django.core.mail import EmailMessage
    from SGDS.models import EnvoiEtatMensuel, ModeleEmailEtatMensuel, Societe

    email_dest = email_override or marketeur.email or marketeur.email_representant or ''
    try:
        if not email_dest:
            raise ValueError("Aucune adresse email renseignée pour ce marketeur.")

        societe = Societe.get_instance()
        gabarit = ModeleEmailEtatMensuel.get_instance()
        sujet, corps = gabarit.rendre(marketeur=marketeur, periode=periode, societe=societe)

        pieces = _generer_pieces_jointes(periode, marketeur)

        message = EmailMessage(sujet, corps, config.from_email, [email_dest], connection=config.get_connection())
        for nom_fichier, contenu, mimetype in pieces:
            message.attach(nom_fichier, contenu, mimetype)
        message.send(fail_silently=False)

        return EnvoiEtatMensuel.objects.create(
            periode=periode, marketeur=marketeur, email_destinataire=email_dest,
            statut='SUCCES', declenche_par=declenche_par,
        )
    except Exception as exc:
        return EnvoiEtatMensuel.objects.create(
            periode=periode, marketeur=marketeur, email_destinataire=email_dest,
            statut='ECHEC', message_erreur=str(exc), declenche_par=declenche_par,
        )
