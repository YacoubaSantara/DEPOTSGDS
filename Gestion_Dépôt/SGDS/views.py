from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.forms import modelformset_factory
from django.db import IntegrityError
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from .models import Marketeur, Camion, Chauffeur, Famille, Produit, Cuve, ParametreJaugeageCuve, JaugeageJour, MesureCuve, Mouvement, LigneMouvement
from .forms import (
    MarketeurForm, CamionForm, ChauffeurForm, FamilleForm, ProduitForm, CuveForm,
    ParametreJaugeageCuveForm, JaugeageJourForm, MesureCuveForm, MouvementForm,
    LigneMouvementFormSet, CompartimentCamionFormSet,
)
import qrcode
import base64
import io


# Helpers
def _deny_marketeur(request):
    """Refuse l'accès aux utilisateurs avec rôle MARKETEUR (lecture seule)."""
    if request.user.is_marketeur_role:
        messages.error(request, "Action non autorisée pour votre rôle.")
        return True
    return False


def _check_mouvement_acces(request, mouvement):
    """
    Vérifie qu'un marketeur connecté a le droit de voir ce mouvement.
    Retourne une réponse de redirection si accès refusé, sinon None.
    Un marketeur ne peut accéder qu'aux mouvements dont il est le marketeur
    ou le destinataire (cession).
    """
    if not request.user.is_marketeur_role:
        return None  # Admin / staff : accès libre
    mkt = getattr(request.user, 'marketeur', None)
    if mkt is None:
        messages.error(request, "Votre compte n'est lié à aucun marketeur.")
        return redirect('connexion')
    est_concerne = (
        mouvement.marketeur_id == mkt.pk or
        (mouvement.cession_marketeur_destinataire_id is not None and
         mouvement.cession_marketeur_destinataire_id == mkt.pk)
    )
    if not est_concerne:
        messages.error(request, "Accès refusé — ce mouvement ne vous appartient pas.")
        return redirect('client_mouvements')
    return None


#
#  MARKETEUR
# 

@login_required
def marketeur_list(request):
    # Un marketeur est redirigé vers son propre profil
    if request.user.is_marketeur_role:
        if request.user.marketeur:
            return redirect('marketeur_detail', uuid=request.user.marketeur.uuid, slug=request.user.marketeur.slug)
        messages.error(request, "Votre compte n'est lié Ã  aucun marketeur.")
        return redirect('connexion')

    qs     = Marketeur.objects.all()
    q      = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')
    ville  = request.GET.get('ville', '')

    if q:
        qs = qs.filter(
            Q(raison_sociale__icontains=q) |
            Q(sigle__icontains=q) |
            Q(ville__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    if ville:
        qs = qs.filter(ville__icontains=ville)

    villes = Marketeur.objects.values_list('ville', flat=True).distinct().order_by('ville')

    paginator  = Paginator(qs, 10)
    page_num   = request.GET.get('page', 1)
    marketeurs = paginator.get_page(page_num)

    ctx = {
        'marketeurs': marketeurs,
        'count':      qs.count(),
        'total':      Marketeur.objects.count(),
        'nb_actif':   Marketeur.objects.filter(statut='ACTIF').count(),
        'nb_suspendu':Marketeur.objects.filter(statut='SUSPENDU').count(),
        'nb_black':   Marketeur.objects.filter(statut='BLACKLIST').count(),
        'villes':     villes,
        'q':          q,
        'statut':     statut,
        'ville':      ville,
        'filtres':    {'q': q, 'statut': statut, 'ville': ville},
    }
    return render(request, 'Marketeur/marketeur_list.html', ctx)


@login_required
def marketeur_detail(request, uuid, slug):
    mkt = get_object_or_404(Marketeur, uuid=uuid)
    # Un marketeur ne peut voir que son propre profil
    if request.user.is_marketeur_role:
        if not request.user.marketeur or str(request.user.marketeur.uuid) != str(uuid):
            messages.error(request, "Accès refusé.")
            if request.user.marketeur:
                return redirect('marketeur_detail', uuid=request.user.marketeur.uuid, slug=request.user.marketeur.slug)
            return redirect('connexion')
    return render(request, 'Marketeur/marketeur_detail.html', {'mkt': mkt})


@login_required
def marketeur_create(request):
    if _deny_marketeur(request):
        return redirect('marketeur_list')
    if request.method == 'POST':
        form = MarketeurForm(request.POST, request.FILES)
        if form.is_valid():
            mkt = form.save()
            messages.success(request, f'Marketeur « {mkt.raison_sociale} » créé avec succès.')
            return redirect('marketeur_list')
    else:
        form = MarketeurForm()
    return render(request, 'Marketeur/marketeur_form.html', {'form': form, 'action': 'Nouveau'})


@login_required
def marketeur_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('marketeur_detail', uuid=mkt.uuid, slug=mkt.slug)
    mkt = get_object_or_404(Marketeur, uuid=uuid)
    if request.method == 'POST':
        form = MarketeurForm(request.POST, request.FILES, instance=mkt)
        if form.is_valid():
            form.save()
            messages.success(request, f'Marketeur « {mkt.raison_sociale} » modifié avec succès.')
            return redirect('marketeur_detail', uuid=mkt.uuid, slug=mkt.slug)
    else:
        form = MarketeurForm(instance=mkt)
    return render(request, 'Marketeur/marketeur_form.html', {'form': form, 'action': 'Modifier', 'mkt': mkt})


@login_required
def marketeur_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('marketeur_detail', uuid=mkt.uuid, slug=mkt.slug)
    mkt = get_object_or_404(Marketeur, uuid=uuid)
    if request.method == 'POST':
        nom = mkt.raison_sociale
        mkt.delete()
        messages.success(request, f'Marketeur « {nom} » supprimé.')
        return redirect('marketeur_list')
    return render(request, 'Marketeur/marketeur_confirm_delete.html', {'mkt': mkt})


#
#  CAMION
#

@login_required
def camion_list(request):
    qs = Camion.objects.select_related('marketeur').all()

    # Filtrer par marketeur si rôle MARKETEUR
    if request.user.is_marketeur_role and request.user.marketeur:
        qs = qs.filter(marketeur=request.user.marketeur)

    q      = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')
    type_p = request.GET.get('type_produit', '')

    if q:
        qs = qs.filter(
            Q(immatriculation__icontains=q) |
            Q(marque__icontains=q) |
            Q(modele__icontains=q) |
            Q(marketeur__raison_sociale__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    if type_p:
        qs = qs.filter(type_produit=type_p)

    paginator = Paginator(qs, 10)
    page_num  = request.GET.get('page', 1)
    camions   = paginator.get_page(page_num)

    ctx = {
        'camions':          camions,
        'total':            Camion.objects.count(),
        'nb_service':       Camion.objects.filter(statut='EN_SERVICE').count(),
        'nb_maintenance':   Camion.objects.filter(statut='EN_MAINTENANCE').count(),
        'nb_hors_service':  Camion.objects.filter(statut='HORS_SERVICE').count(),
        'type_choices':     Camion.TYPE_PRODUIT_CHOICES,
        'statut_choices':   Camion.STATUT_CHOICES,
        'q':        q,
        'statut':   statut,
        'type_produit': type_p,
        'filtres':  {'q': q, 'statut': statut, 'type_produit': type_p},
    }
    return render(request, 'Camion/camion_list.html', ctx)


@login_required
def camion_detail(request, uuid, slug):
    camion = get_object_or_404(Camion, uuid=uuid)
    # Vérification accès marketeur
    if request.user.is_marketeur_role and request.user.marketeur:
        if camion.marketeur_id != request.user.marketeur.pk:
            messages.error(request, "Accès refusé.")
            return redirect('camion_list')
    return render(request, 'Camion/camion_detail.html', {'camion': camion})


@login_required
def camion_create(request):
    if _deny_marketeur(request):
        return redirect('camion_list')
    if request.method == 'POST':
        form    = CamionForm(request.POST, request.FILES)
        formset = CompartimentCamionFormSet(request.POST, prefix='compartiments')
        if form.is_valid() and formset.is_valid():
            cam = form.save()
            formset.instance = cam
            formset.save()
            messages.success(request, f'Camion « {cam.immatriculation} » enregistré avec succès.')
            return redirect('camion_list')
    else:
        form    = CamionForm()
        formset = CompartimentCamionFormSet(prefix='compartiments')
    return render(request, 'Camion/camion_form.html', {
        'form': form, 'formset': formset, 'action': 'Nouveau'
    })


@login_required
def camion_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('camion_detail', uuid=camion.uuid, slug=camion.slug)
    camion = get_object_or_404(Camion, uuid=uuid)
    if request.method == 'POST':
        form    = CamionForm(request.POST, request.FILES, instance=camion)
        formset = CompartimentCamionFormSet(request.POST, instance=camion, prefix='compartiments')
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, f'Camion « {camion.immatriculation} » modifié avec succès.')
            return redirect('camion_detail', uuid=camion.uuid, slug=camion.slug)
    else:
        form    = CamionForm(instance=camion)
        formset = CompartimentCamionFormSet(instance=camion, prefix='compartiments')
    return render(request, 'Camion/camion_form.html', {
        'form': form, 'formset': formset, 'action': 'Modifier', 'camion': camion
    })


@login_required
def camion_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('camion_detail', uuid=camion.uuid, slug=camion.slug)
    camion = get_object_or_404(Camion, uuid=uuid)
    if request.method == 'POST':
        immat = camion.immatriculation
        camion.delete()
        messages.success(request, f'Camion a « {immat} » été supprimé.')
        return redirect('camion_list')
    return render(request, 'Camion/camion_confirm_delete.html', {'camion': camion})


#
#  CHAUFFEUR
#

@login_required
def chauffeur_list(request):
    qs = Chauffeur.objects.select_related('marketeur', 'camion').all()

    # Filtrer par marketeur si rôle MARKETEUR
    if request.user.is_marketeur_role and request.user.marketeur:
        qs = qs.filter(marketeur=request.user.marketeur)

    q      = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) |
            Q(prenom__icontains=q) |
            Q(telephone__icontains=q) |
            Q(numero_permis__icontains=q) |
            Q(marketeur__raison_sociale__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)

    paginator  = Paginator(qs, 10)
    page_num   = request.GET.get('page', 1)
    chauffeurs = paginator.get_page(page_num)

    ctx = {
        'chauffeurs':    chauffeurs,
        'total':         Chauffeur.objects.count(),
        'nb_actif':      Chauffeur.objects.filter(statut='ACTIF').count(),
        'nb_inactif':    Chauffeur.objects.filter(statut='INACTIF').count(),
        'nb_suspendu':   Chauffeur.objects.filter(statut='SUSPENDU').count(),
        'q':      q,
        'statut': statut,
        'filtres': {'q': q, 'statut': statut},
    }
    return render(request, 'Chauffeur/chauffeur_list.html', ctx)


@login_required
def chauffeur_detail(request, uuid, slug):
    chauffeur = get_object_or_404(Chauffeur, uuid=uuid)
    if request.user.is_marketeur_role and request.user.marketeur:
        if chauffeur.marketeur_id != request.user.marketeur.pk:
            messages.error(request, "Accès refusé.")
            return redirect('chauffeur_list')
    return render(request, 'Chauffeur/chauffeur_detail.html', {'chauffeur': chauffeur})


@login_required
def chauffeur_create(request):
    if _deny_marketeur(request):
        return redirect('chauffeur_list')
    if request.method == 'POST':
        form = ChauffeurForm(request.POST, request.FILES)
        if form.is_valid():
            chf = form.save(commit=False)
            chf.numero_employe = Chauffeur.get_next_numero()
            chf.save()
            messages.success(request, f'Chauffeur « {chf.nom_complet} » enregistré avec succès.')
            return redirect('chauffeur_list')
    else:
        form = ChauffeurForm()
    return render(request, 'Chauffeur/chauffeur_form.html', {'form': form, 'action': 'Nouveau'})


@login_required
def chauffeur_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('chauffeur_detail', uuid=chauffeur.uuid, slug=chauffeur.slug)
    chauffeur = get_object_or_404(Chauffeur, uuid=uuid)
    if request.method == 'POST':
        form = ChauffeurForm(request.POST, request.FILES, instance=chauffeur)
        if form.is_valid():
            form.save()
            messages.success(request, f'Chauffeur « {chauffeur.nom_complet} » modifié avec succès.')
            return redirect('chauffeur_detail', uuid=chauffeur.uuid, slug=chauffeur.slug)
    else:
        form = ChauffeurForm(instance=chauffeur)
    return render(request, 'Chauffeur/chauffeur_form.html', {'form': form, 'action': 'Modifier', 'chauffeur': chauffeur})


@login_required
def chauffeur_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('chauffeur_detail', uuid=chauffeur.uuid, slug=chauffeur.slug)
    chauffeur = get_object_or_404(Chauffeur, uuid=uuid)
    if request.method == 'POST':
        nom = chauffeur.nom_complet
        chauffeur.delete()
        messages.success(request, f'Chauffeur « {nom} » supprimé.')
        return redirect('chauffeur_list')
    return render(request, 'Chauffeur/chauffeur_confirm_delete.html', {'chauffeur': chauffeur})


@login_required
def chauffeur_badge(request, uuid, slug):
    chauffeur = get_object_or_404(Chauffeur.objects.select_related('marketeur', 'camion'), uuid=uuid)
    if request.user.is_marketeur_role and request.user.marketeur:
        if chauffeur.marketeur_id != request.user.marketeur.pk:
            messages.error(request, "Accès refusé.")
            return redirect('chauffeur_list')

    badge_url = request.build_absolute_uri(f'/chauffeurs/{pk}/')
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=2,
    )
    qr.add_data(badge_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1E3A5F", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    return render(request, 'Chauffeur/chauffeur_badge.html', {
        'chauffeur': chauffeur,
        'qr_b64': qr_b64,
    })


# 
#  FAMILLE
# 

@login_required
def famille_list(request):
    qs = Famille.objects.all()
    q      = request.GET.get('q', '').strip()
    statut = request.GET.get('statut', '')

    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(code__icontains=q) | Q(description__icontains=q))
    if statut:
        qs = qs.filter(statut=statut)

    paginator = Paginator(qs, 10)
    page_num  = request.GET.get('page', 1)
    familles  = paginator.get_page(page_num)

    ctx = {
        'familles':  familles,
        'total':     Famille.objects.count(),
        'nb_actif':  Famille.objects.filter(statut='ACTIF').count(),
        'nb_inactif':Famille.objects.filter(statut='INACTIF').count(),
        'q':     q,
        'statut':statut,
        'filtres': {'q': q, 'statut': statut},
    }
    return render(request, 'Famille/famille_list.html', ctx)


@login_required
def famille_detail(request, uuid, slug):
    famille = get_object_or_404(Famille, uuid=uuid)
    produits = famille.produits.all()
    return render(request, 'Famille/famille_detail.html', {'famille': famille, 'produits': produits})


@login_required
def famille_create(request):
    if _deny_marketeur(request):
        return redirect('famille_list')
    if request.method == 'POST':
        form = FamilleForm(request.POST)
        if form.is_valid():
            fam = form.save()
            messages.success(request, f'Famille « {fam.nom} » créée avec succès.')
            return redirect('famille_list')
    else:
        form = FamilleForm()
    return render(request, 'Famille/famille_form.html', {'form': form, 'action': 'Nouvelle'})


@login_required
def famille_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('famille_detail', uuid=famille.uuid, slug=famille.slug)
    famille = get_object_or_404(Famille, uuid=uuid)
    if request.method == 'POST':
        form = FamilleForm(request.POST, instance=famille)
        if form.is_valid():
            form.save()
            messages.success(request, f'Famille « {famille.nom} » modifiée avec succès.')
            return redirect('famille_detail', uuid=famille.uuid, slug=famille.slug)
    else:
        form = FamilleForm(instance=famille)
    return render(request, 'Famille/famille_form.html', {'form': form, 'action': 'Modifier', 'famille': famille})


@login_required
def famille_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('famille_detail', uuid=famille.uuid, slug=famille.slug)
    famille = get_object_or_404(Famille, uuid=uuid)
    if request.method == 'POST':
        nom = famille.nom
        famille.delete()
        messages.success(request, f'Famille « {nom} » supprimée.')
        return redirect('famille_list')
    return render(request, 'Famille/famille_confirm_delete.html', {'famille': famille})


# 
#  PRODUIT
# 

@login_required
def produit_list(request):
    qs = Produit.objects.select_related('famille').all()
    q          = request.GET.get('q', '').strip()
    statut     = request.GET.get('statut', '')
    famille_id = request.GET.get('famille', '')

    if q:
        qs = qs.filter(
            Q(nom__icontains=q) | Q(code__icontains=q) |
            Q(description__icontains=q) | Q(famille__nom__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    if famille_id:
        qs = qs.filter(famille_id=famille_id)

    paginator = Paginator(qs, 10)
    page_num  = request.GET.get('page', 1)
    produits  = paginator.get_page(page_num)

    ctx = {
        'produits':      produits,
        'total':         Produit.objects.count(),
        'nb_actif':      Produit.objects.filter(statut='ACTIF').count(),
        'nb_inactif':    Produit.objects.filter(statut='INACTIF').count(),
        'nb_discontinue':Produit.objects.filter(statut='DISCONTINUE').count(),
        'familles':      Famille.objects.filter(statut='ACTIF'),
        'q':         q,
        'statut':    statut,
        'famille_id':famille_id,
        'filtres':   {'q': q, 'statut': statut, 'famille': famille_id},
    }
    return render(request, 'Produit/produit_list.html', ctx)


@login_required
def produit_detail(request, uuid, slug):
    produit = get_object_or_404(Produit.objects.select_related('famille'), uuid=uuid)
    cuves = produit.cuves.all()
    return render(request, 'Produit/produit_detail.html', {'produit': produit, 'cuves': cuves})


@login_required
def produit_create(request):
    if _deny_marketeur(request):
        return redirect('produit_list')
    if request.method == 'POST':
        form = ProduitForm(request.POST)
        if form.is_valid():
            prod = form.save()
            messages.success(request, f'Produit « {prod.nom} » créé avec succès.')
            return redirect('produit_list')
    else:
        form = ProduitForm()
    return render(request, 'Produit/produit_form.html', {'form': form, 'action': 'Nouveau'})


@login_required
def produit_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('produit_detail', uuid=produit.uuid, slug=produit.slug)
    produit = get_object_or_404(Produit, uuid=uuid)
    if request.method == 'POST':
        form = ProduitForm(request.POST, instance=produit)
        if form.is_valid():
            form.save()
            messages.success(request, f'Produit « {produit.nom} » modifié avec succès.')
            return redirect('produit_detail', uuid=produit.uuid, slug=produit.slug)
    else:
        form = ProduitForm(instance=produit)
    return render(request, 'Produit/produit_form.html', {'form': form, 'action': 'Modifier', 'produit': produit})


@login_required
def produit_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('produit_detail', uuid=produit.uuid, slug=produit.slug)
    produit = get_object_or_404(Produit, uuid=uuid)
    if request.method == 'POST':
        nom = produit.nom
        produit.delete()
        messages.success(request, f'Produit « {nom} » supprimé.')
        return redirect('produit_list')
    return render(request, 'Produit/produit_confirm_delete.html', {'produit': produit})


# 
#  CUVE
# 

@login_required
def cuve_list(request):
    qs = Cuve.objects.select_related('produit', 'produit__famille').all()
    q          = request.GET.get('q', '').strip()
    statut     = request.GET.get('statut', '')
    produit_id = request.GET.get('produit', '')

    if q:
        qs = qs.filter(
            Q(numero__icontains=q) | Q(designation__icontains=q) |
            Q(localisation__icontains=q) | Q(produit__nom__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    if produit_id:
        qs = qs.filter(produit_id=produit_id)

    paginator = Paginator(qs, 10)
    page_num  = request.GET.get('page', 1)
    cuves     = paginator.get_page(page_num)

    ctx = {
        'cuves':           cuves,
        'total':           Cuve.objects.count(),
        'nb_active':       Cuve.objects.filter(statut='ACTIVE').count(),
        'nb_maintenance':  Cuve.objects.filter(statut='EN_MAINTENANCE').count(),
        'nb_hors_service': Cuve.objects.filter(statut='HORS_SERVICE').count(),
        'nb_inactive':     Cuve.objects.filter(statut='INACTIVE').count(),
        'produits':        Produit.objects.filter(statut='ACTIF'),
        'statut_choices':  Cuve.STATUT_CHOICES,
        'q':          q,
        'statut':     statut,
        'produit_id': produit_id,
        'filtres':    {'q': q, 'statut': statut, 'produit': produit_id},
    }
    return render(request, 'Cuve/cuve_list.html', ctx)


@login_required
def cuve_detail(request, uuid, slug):
    cuve = get_object_or_404(Cuve.objects.select_related('produit', 'produit__famille'), uuid=uuid)
    return render(request, 'Cuve/cuve_detail.html', {'cuve': cuve})


@login_required
def cuve_create(request):
    if _deny_marketeur(request):
        return redirect('cuve_list')
    if request.method == 'POST':
        form = CuveForm(request.POST)
        if form.is_valid():
            cuve = form.save()
            messages.success(request, f'Cuve « {cuve.numero} » enregistrée avec succès.')
            return redirect('cuve_list')
    else:
        form = CuveForm()
    return render(request, 'Cuve/cuve_form.html', {'form': form, 'action': 'Nouvelle'})


@login_required
def cuve_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('cuve_detail', uuid=cuve.uuid, slug=cuve.slug)
    cuve = get_object_or_404(Cuve, uuid=uuid)
    if request.method == 'POST':
        form = CuveForm(request.POST, instance=cuve)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cuve « {cuve.numero} » modifiée avec succès.')
            return redirect('cuve_detail', uuid=cuve.uuid, slug=cuve.slug)
    else:
        form = CuveForm(instance=cuve)
    return render(request, 'Cuve/cuve_form.html', {'form': form, 'action': 'Modifier', 'cuve': cuve})


@login_required
def cuve_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('cuve_detail', uuid=cuve.uuid, slug=cuve.slug)
    cuve = get_object_or_404(Cuve, uuid=uuid)
    if request.method == 'POST':
        num = cuve.numero
        cuve.delete()
        messages.success(request, f'Cuve « {num} » supprimée.')
        return redirect('cuve_list')
    return render(request, 'Cuve/cuve_confirm_delete.html', {'cuve': cuve})


# 
#  PARAMÈTRES DE JAUGEAGE
# 

@login_required
def parametre_list(request):
    cuves = Cuve.objects.select_related('parametre_jaugeage').order_by('numero')
    return render(request, 'ParametreJaugeage/parametre_list.html', {
        'cuves':         cuves,
        'total_cuves':   cuves.count(),
        'nb_configures': sum(1 for c in cuves if hasattr(c, 'parametre_jaugeage')),
    })


@login_required
def parametre_detail(request, uuid, slug):
    parametre = get_object_or_404(ParametreJaugeageCuve.objects.select_related('cuve'), uuid=uuid)
    return render(request, 'ParametreJaugeage/parametre_detail.html', {'parametre': parametre})


@login_required
def parametre_create_update(request, cuve_uuid, cuve_slug):
    if _deny_marketeur(request):
        return redirect('parametre_list')
    cuve = get_object_or_404(Cuve, uuid=cuve_uuid)
    instance = getattr(cuve, 'parametre_jaugeage', None)
    action = 'Modifier' if instance else 'Configurer'

    if request.method == 'POST':
        form = ParametreJaugeageCuveForm(request.POST, instance=instance)
        if form.is_valid():
            parametre = form.save(commit=False)
            parametre.cuve = cuve
            parametre.save()
            messages.success(request, f'Paramètres de jaugeage de la cuve {cuve.numero} enregistrés.')
            return redirect('parametre_detail', uuid=parametre.uuid, slug=parametre.slug)
    else:
        form = ParametreJaugeageCuveForm(instance=instance)

    return render(request, 'ParametreJaugeage/parametre_form.html', {
        'form': form, 'cuve': cuve, 'action': action, 'instance': instance,
    })


@login_required
def parametre_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('parametre_list')
    parametre = get_object_or_404(ParametreJaugeageCuve.objects.select_related('cuve'), uuid=uuid)
    if request.method == 'POST':
        num = parametre.cuve.numero
        parametre.delete()
        messages.success(request, f'Paramètres de jaugeage de la cuve {num} supprimés.')
        return redirect('parametre_list')
    return render(request, 'ParametreJaugeage/parametre_confirm_delete.html', {'parametre': parametre})


# 
#  JAUGEAGE DU JOUR
# 

@login_required
def jaugeage_list(request):
    qs = JaugeageJour.objects.prefetch_related(
        'mesures__cuve__parametre_jaugeage'
    ).all()
    q      = request.GET.get('q', '').strip()
    type_j = request.GET.get('type_jaugeage', '')
    date_d = request.GET.get('date_debut', '')
    date_f = request.GET.get('date_fin', '')

    if q:
        qs = qs.filter(Q(operateur__icontains=q) | Q(depot__icontains=q))
    if type_j:
        qs = qs.filter(type_jaugeage=type_j)
    if date_d:
        qs = qs.filter(date_jaugeage__gte=date_d)
    if date_f:
        qs = qs.filter(date_jaugeage__lte=date_f)

    # Pagination avant le calcul Python (pour ne traiter que la page courante)
    paginator     = Paginator(qs, 10)
    page_num      = request.GET.get('page', 1)
    jaugeages_page = paginator.get_page(page_num)

    # Calcul du volume total @15°C par jaugeage (propriétés Python, non annotables SQL)
    jaugeages_with_totals = []
    for j in jaugeages_page:
        mesures = list(j.mesures.all())
        total = sum(float(m.volume_standard_15c_calcule or 0) for m in mesures)
        total_vad = sum(float(m.volume_ambiant_depot or 0) for m in mesures)
        jaugeages_with_totals.append((j, total if total > 0 else None, total_vad if total_vad > 0 else None))

    # Stocks produits pour le bandeau récapitulatif
    from django.db.models import Max
    stocks_produits = Produit.objects.filter(statut='ACTIF').order_by('famille', 'nom')
    date_derniere_maj = stocks_produits.aggregate(Max('date_maj_stock'))['date_maj_stock__max']

    ctx = {
        'jaugeages':              jaugeages_page,
        'jaugeages_with_totals':  jaugeages_with_totals,
        'total':          JaugeageJour.objects.count(),
        'nb_avr':         JaugeageJour.objects.filter(type_jaugeage='AVR').count(),
        'nb_apr':         JaugeageJour.objects.filter(type_jaugeage='APR').count(),
        'nb_j':           JaugeageJour.objects.filter(type_jaugeage='J').count(),
        'type_choices':   JaugeageJour.TYPE_CHOICES,
        'q':              q,
        'type_jaugeage':  type_j,
        'date_debut':     date_d,
        'date_fin':       date_f,
        'filtres':        {'q': q, 'type_jaugeage': type_j, 'date_debut': date_d, 'date_fin': date_f},
        'stocks_produits':    stocks_produits,
        'date_derniere_maj':  date_derniere_maj,
    }
    return render(request, 'Jaugeage/jaugeage_list.html', ctx)


@login_required
def jaugeage_detail(request, uuid, slug):
    jaugeage = get_object_or_404(
        JaugeageJour.objects.prefetch_related('mesures__cuve__parametre_jaugeage'),
        uuid=uuid
    )
    # â"€â"€ Totaux volumes pour le tfoot â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    mesures = list(jaugeage.mesures.all())
    def _sum(attr):
        vals = [float(getattr(m, attr)) for m in mesures if getattr(m, attr) is not None]
        return sum(vals) if vals else None

    return render(request, 'Jaugeage/jaugeage_detail.html', {
        'jaugeage':     jaugeage,
        'total_vab':    _sum('volume_ambiant_bac'),
        'total_vad':    _sum('volume_ambiant_depot'),
        'total_v15c':   _sum('volume_standard_15c'),
        'total_vdispo': _sum('volume_disponible'),
    })


@login_required
def jaugeage_create(request):
    if _deny_marketeur(request):
        return redirect('jaugeage_list')
    if request.method == 'POST':
        form = JaugeageJourForm(request.POST)
        if form.is_valid():
            try:
                from datetime import date as _date
                jaugeage = JaugeageJour.creer_nouveau_jaugeage(
                    date_jaugeage=form.cleaned_data['date_jaugeage'],
                    type_jaugeage=form.cleaned_data['type_jaugeage'],
                    heure_jaugeage=form.cleaned_data.get('heure_jaugeage'),
                    operateur=form.cleaned_data.get('operateur'),
                    notes=form.cleaned_data.get('notes'),
                )
                jaugeage.depot               = form.cleaned_data.get('depot') or 'SGDS SANKE'
                jaugeage.type_depot          = form.cleaned_data.get('type_depot') or 'Dépôt de droit'
                jaugeage.temperature_reference = form.cleaned_data.get('temperature_reference') or 15.0
                jaugeage.save()
                messages.success(request, f'Jaugeage du {jaugeage.date_jaugeage} créé. Saisissez les mesures.')
                return redirect('jaugeage_saisie', uuid=jaugeage.uuid, slug=jaugeage.slug)
            except IntegrityError:
                form.add_error(None, 'Un jaugeage avec cette date, ce type et cette heure existe déjÃ .')
    else:
        from datetime import date as _date
        form = JaugeageJourForm(initial={
            'date_jaugeage': _date.today(),
            'operateur':     request.user.nom_complet,
            'depot':         'SGDS SANKE',
            'type_depot':    'Dépôt de droit',
        })
    return render(request, 'Jaugeage/jaugeage_form.html', {'form': form, 'action': 'Nouveau'})


@login_required
def jaugeage_update(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)
    jaugeage = get_object_or_404(JaugeageJour, uuid=uuid)
    if request.method == 'POST':
        form = JaugeageJourForm(request.POST, instance=jaugeage)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Jaugeage modifié avec succès.')
                return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)
            except IntegrityError:
                form.add_error(None, 'Un jaugeage avec cette date, ce type et cette heure existe déjÃ .')
    else:
        form = JaugeageJourForm(instance=jaugeage)
    return render(request, 'Jaugeage/jaugeage_form.html', {'form': form, 'action': 'Modifier', 'jaugeage': jaugeage})


@login_required
def jaugeage_delete(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)
    jaugeage = get_object_or_404(JaugeageJour, uuid=uuid)
    if request.method == 'POST':
        label = str(jaugeage)
        jaugeage.delete()
        # Recalcul explicite après suppression (ceinture + bretelles en plus du signal)
        from .services.recalcul_stock import recalculer_tous_stocks
        recalculer_tous_stocks()
        messages.success(request, f'Jaugeage « {label} » supprimé. Stocks recalculés.')
        return redirect('jaugeage_list')
    return render(request, 'Jaugeage/jaugeage_confirm_delete.html', {'jaugeage': jaugeage})


@login_required
def jaugeage_saisie(request, uuid, slug):
    if _deny_marketeur(request):
        return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)
    jaugeage = get_object_or_404(JaugeageJour, uuid=uuid)
    MesureCuveFormSet = modelformset_factory(MesureCuve, form=MesureCuveForm, extra=0)
    qs = jaugeage.mesures.select_related(
        'cuve', 'cuve__produit', 'cuve__produit__famille', 'cuve__parametre_jaugeage'
    ).order_by('cuve__numero')

    if request.method == 'POST':
        # Refus de modification si le jaugeage est validé
        if jaugeage.est_valide:
            messages.error(
                request,
                "Ce jaugeage est validé et ne peut plus être modifié. "
                "Demandez au chef de dépôt de le déverrouiller."
            )
            return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)
        formset = MesureCuveFormSet(request.POST, queryset=qs)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Mesures enregistrées avec succès.')
            return redirect('jaugeage_saisie', uuid=jaugeage.uuid, slug=jaugeage.slug)   # rester sur la saisie pour voir les calculs
        else:
            messages.error(request, 'Des erreurs ont été détectées. Vérifiez les valeurs saisies.')
    else:
        formset = MesureCuveFormSet(queryset=qs)

    # Associer chaque form Ã  sa mesure
    forms_with_mesures = list(zip(formset.forms, qs))

    # â"€â"€ Grouper par produit pour le layout colonne-par-cuve â"€â"€â"€â"€â"€â"€
    from collections import OrderedDict
    pg_dict = OrderedDict()
    for form, mesure in forms_with_mesures:
        produit = mesure.cuve.produit
        key = produit.pk if produit else 0
        if key not in pg_dict:
            pg_dict[key] = {'produit': produit, 'cuves': [], 'count': 0}
        pg_dict[key]['cuves'].append((form, mesure))
        pg_dict[key]['count'] += 1

    product_groups = list(pg_dict.values())

    # Liste plate dans le même ordre (pour forloop.counter0 global en template)
    all_cuves = [(form, mesure) for pg in product_groups for form, mesure in pg['cuves']]

    # â"€â"€ Totaux volumes ambiant dépôt (calculés après le dernier save) â"€â"€â"€â"€
    totaux_groupes = []
    total_depot_val = 0.0
    total_v15c_val = 0.0
    has_any = False
    has_any_v15c = False
    for pg in product_groups:
        t = 0.0
        has_t = False
        t15 = 0.0
        has_t15 = False
        for _, m in pg['cuves']:
            v = m.volume_ambiant_depot
            if v is not None:
                t += float(v)
                has_t = True
                has_any = True
            v15 = m.volume_standard_15c_calcule
            if v15 is not None:
                t15 += float(v15)
                has_t15 = True
                has_any_v15c = True
        pg['total_vad'] = t if has_t else None
        pg['total_v15c'] = t15 if has_t15 else None
        totaux_groupes.append({'produit': pg['produit'], 'total': t if has_t else None, 'total_v15c': t15 if has_t15 else None})
        total_depot_val += t
        total_v15c_val += t15

    return render(request, 'Jaugeage/jaugeage_saisie.html', {
        'jaugeage':          jaugeage,
        'formset':           formset,
        'forms_with_mesures': forms_with_mesures,
        'product_groups':    product_groups,
        'all_cuves':         all_cuves,
        'totaux_groupes':    totaux_groupes,
        'total_depot':       total_depot_val if has_any else None,
        'total_v15c_depot':  total_v15c_val if has_any_v15c else None,
        'nb_cuves':          len(all_cuves),
    })


# 
#  VALIDATION / DÉVALIDATION D'UN JAUGEAGE
# 

@login_required
def valider_jaugeage(request, uuid, slug):
    """Valide un jaugeage (POST, staff uniquement)."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Réservé aux responsables dépôt.")
    if request.method != 'POST':
        return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)

    jaugeage = get_object_or_404(JaugeageJour, uuid=uuid)

    if jaugeage.est_valide:
        messages.warning(request, "Ce jaugeage est déjÃ  validé.")
        return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)

    # Vérifier que toutes les mesures sont complètes
    mesures_incompletes = []
    for mesure in jaugeage.mesures.select_related('cuve').all():
        champs_requis = [
            mesure.creux_mesure, mesure.v_a_saisi,
            mesure.t1, mesure.t2, mesure.t3,
            mesure.temperature_obs, mesure.densite_moyenne,
        ]
        if any(v is None for v in champs_requis):
            mesures_incompletes.append(mesure.cuve.numero)

    if mesures_incompletes:
        cuves_str = ', '.join(mesures_incompletes)
        messages.error(
            request,
            f"Impossible de valider : les mesures des cuves {cuves_str} sont incomplètes."
        )
        return redirect('jaugeage_saisie', uuid=jaugeage.uuid, slug=jaugeage.slug)

    from django.utils import timezone
    jaugeage.est_valide = True
    jaugeage.date_validation = timezone.now()
    jaugeage.valide_par = request.user
    jaugeage.save(update_fields=['est_valide', 'date_validation', 'valide_par'])

    # Mise Ã  jour automatique des stocks produits
    Produit.mettre_a_jour_stocks(jaugeage)

    messages.success(
        request,
        f"Jaugeage du {jaugeage.date_jaugeage.strftime('%d/%m/%Y')} validé avec succès."
    )
    return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)


@login_required
def devalider_jaugeage(request, uuid, slug):
    """Déverrouille un jaugeage validé (POST, staff uniquement)."""
    if not request.user.is_staff:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Réservé aux responsables dépôt.")
    if request.method != 'POST':
        return redirect('jaugeage_detail', uuid=jaugeage.uuid, slug=jaugeage.slug)

    jaugeage = get_object_or_404(JaugeageJour, uuid=uuid)
    jaugeage.est_valide = False
    jaugeage.date_validation = None
    jaugeage.valide_par = None
    jaugeage.save(update_fields=['est_valide', 'date_validation', 'valide_par'])

    # Recalcul des stocks : reprend le jaugeage précédent s'il existe
    from .services.recalcul_stock import recalculer_tous_stocks
    recalculer_tous_stocks()

    messages.info(
        request,
        f"Jaugeage du {jaugeage.date_jaugeage.strftime('%d/%m/%Y')} déverrouillé. Stocks recalculés."
    )
    return redirect('jaugeage_saisie', uuid=jaugeage.uuid, slug=jaugeage.slug)


# 
#  RAPPORT D'IMPRESSION — Vue A4 du jaugeage
# 

@login_required
def jaugeage_rapport(request, uuid, slug):
    """Vue d'impression A4 du Rapport de Jaugeage Journalier (RJJ)."""
    jaugeage = get_object_or_404(
        JaugeageJour.objects.prefetch_related(
            'mesures__cuve__parametre_jaugeage',
            'mesures__cuve__produit__famille',
        ),
        uuid=uuid,
    )

    # Grouper les mesures par produit (même ordre que la saisie)
    from collections import OrderedDict
    groups_dict = OrderedDict()
    for m in jaugeage.mesures.select_related(
        'cuve__produit__famille', 'cuve__parametre_jaugeage'
    ).order_by('cuve__numero'):
        produit = m.cuve.produit
        key = produit.pk if produit else 0
        if key not in groups_dict:
            groups_dict[key] = {'produit': produit, 'mesures': []}
        groups_dict[key]['mesures'].append(m)

    product_groups = list(groups_dict.values())

    # Totaux par groupe : volume_ambiant_depot (primaire, comme saisie) + @15°C (secondaire)
    totaux_groupes = []
    total_depot_vad = 0.0
    total_depot_v15c = 0.0
    has_any_vad = False
    has_any_v15c = False
    for pg in product_groups:
        t_vad = 0.0
        t_v15c = 0.0
        has_vad = False
        has_v15c = False
        for m in pg['mesures']:
            v_vad = m.volume_ambiant_depot
            v_v15c = m.volume_standard_15c_calcule
            if v_vad is not None:
                t_vad += float(v_vad)
                has_vad = True
                has_any_vad = True
            if v_v15c is not None:
                t_v15c += float(v_v15c)
                has_v15c = True
                has_any_v15c = True
        pg['total_vad'] = t_vad if has_vad else None
        pg['total_v15c'] = t_v15c if has_v15c else None
        totaux_groupes.append({
            'produit':    pg['produit'],
            'total_vad':  t_vad if has_vad else None,
            'total_v15c': t_v15c if has_v15c else None,
        })
        total_depot_vad += t_vad
        total_depot_v15c += t_v15c

    # Liste plate des mesures (pour colonnes du tableau rapport)
    all_mesures = [m for pg in product_groups for m in pg['mesures']]

    return render(request, 'Jaugeage/jaugeage_rapport.html', {
        'jaugeage':       jaugeage,
        'product_groups': product_groups,
        'all_mesures':    all_mesures,
        'totaux_groupes': totaux_groupes,
        'total_depot':    total_depot_vad if has_any_vad else None,
        'total_v15c':     total_depot_v15c if has_any_v15c else None,
        'nb_cuves':       len(all_mesures),
    })


# 
#  PARAMÈTRES MÉTROLOGIQUES — Fiche normative API MPMS
# 

@login_required
def parametres_metrologiques(request):
    """
    Écran informatif en lecture seule : présente les constantes et l'algorithme
    API MPMS Chapter 11.1 / ASTM D1250 (Octobre 1995) utilisés dans petroleum_calc.py.
    """
    from .petroleum_calc import K_SUPER, K_MIDDLE, K_HEAVY, A_AMB, B_AMB

    plages = [
        {
            'id': 'super',
            'label': 'Produits légers',
            'categorie': 'Super, essence sans plomb',
            'seuil': 'Ï â‰¤ 770 kg/mÂ³',
            'type': 'standard',
            'k0': K_SUPER[0],
            'k1': K_SUPER[1],
            'icone_couleur': '#E8760A',
            'bg_couleur': '#fff7ed',
            'border_couleur': '#fed7aa',
        },
        {
            'id': 'ambigue',
            'label': 'Zone ambiguÃ«',
            'categorie': 'Algorithme itératif spécial (point fixe, démarrage Ã  778,84 kg/mÂ³)',
            'seuil': '770 < Ï < 788 kg/mÂ³',
            'type': 'ambiguous',
            'a': A_AMB,
            'b': B_AMB,
            'icone_couleur': '#3b82f6',
            'bg_couleur': '#eff6ff',
            'border_couleur': '#bfdbfe',
        },
        {
            'id': 'middle',
            'label': 'Produits moyens',
            'categorie': 'Gasoil, kérosène, jet-A1',
            'seuil': '788 â‰¤ Ï < 839 kg/mÂ³',
            'type': 'standard',
            'k0': K_MIDDLE[0],
            'k1': K_MIDDLE[1],
            'icone_couleur': '#16a34a',
            'bg_couleur': '#f0fdf4',
            'border_couleur': '#bbf7d0',
        },
        {
            'id': 'heavy',
            'label': 'Produits lourds',
            'categorie': 'Fuel-oil, résidus lourds',
            'seuil': 'Ï â‰¥ 839 kg/mÂ³',
            'type': 'standard',
            'k0': K_HEAVY[0],
            'k1': K_HEAVY[1],
            'icone_couleur': '#64748b',
            'bg_couleur': '#f8fafc',
            'border_couleur': '#e2e8f0',
        },
    ]

    norme = {
        'code': 'API MPMS Chapter 11.1 / ASTM D1250',
        'titre': (
            'Generalized Crude Oils, Refined Products and Lubricating Oils — '
            'Correction of observed density to density at 15°C'
        ),
        'date': 'Octobre 1995',
        'source_historique': r'J:\OPS\TRH_15.XLS / TVCF_15.XLS',
        'statut': "En vigueur dans l'industrie pétrolière de l'Afrique de l'Ouest",
        'iterations': 7,
    }

    return render(request, 'Jaugeage/parametres_metrologiques.html', {
        'plages': plages,
        'norme': norme,
    })


# 
#  MOUVEMENTS (entrée / sortie / cession / acquittement)
# 

@login_required
def mouvement_liste(request):
    """Liste paginée de tous les mouvements avec filtres GET."""
    from django.core.paginator import Paginator
    from django.db.models import Sum, Count

    qs = Mouvement.objects.select_related('produit', 'marketeur', 'camion').prefetch_related('lignes__cuve').order_by('-date_mouvement', '-date_saisie')

    # â"€â"€ Filtres â"€â"€
    type_m      = request.GET.get('type', '').strip()
    regime      = request.GET.get('regime', '').strip()
    mkt_pk      = request.GET.get('marketeur', '').strip()
    produit_pk  = request.GET.get('produit', '').strip()
    date_debut  = request.GET.get('date_debut', '').strip()
    date_fin    = request.GET.get('date_fin', '').strip()
    q           = request.GET.get('q', '').strip()

    if type_m:
        qs = qs.filter(type_mouvement=type_m)
    if regime:
        qs = qs.filter(regime_douanier=regime)
    if mkt_pk:
        qs = qs.filter(marketeur_id=mkt_pk)
    if produit_pk:
        qs = qs.filter(produit_id=produit_pk)
    if date_debut:
        qs = qs.filter(date_mouvement__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_mouvement__lte=date_fin)
    if q:
        qs = qs.filter(
            Q(numero_enregistrement__icontains=q) |
            Q(camion__immatriculation__icontains=q) |
            Q(bl_expediteur__icontains=q) |
            Q(marketeur__raison_sociale__icontains=q)
        )

    # â"€â"€ Totaux sur le queryset filtré (avant pagination) â"€â"€
    nb_total   = qs.count()

    # Pagination
    paginator  = Paginator(qs, 50)
    page_num   = request.GET.get('page', 1)
    mouvements = paginator.get_page(page_num)

    marketeurs = Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale')
    produits   = Produit.objects.filter(statut='ACTIF').order_by('nom')

    return render(request, 'mouvements/liste.html', {
        'mouvements':  mouvements,
        'marketeurs':  marketeurs,
        'produits':    produits,
        'nb_total':    nb_total,
        'filtres': {
            'type':       type_m,
            'regime':     regime,
            'marketeur':  mkt_pk,
            'produit':    produit_pk,
            'date_debut': date_debut,
            'q':          q,
            'date_fin':   date_fin,
        },
        'type_choices':   Mouvement.TYPE_CHOICES,
        'regime_choices': Mouvement.REGIME_CHOICES,
    })


@login_required
def mouvement_creer(request):
    """Formulaire de création d'un nouveau mouvement."""
    if _deny_marketeur(request):
        return redirect('mouvement_liste')

    if request.method == 'POST':
        form = MouvementForm(request.POST)
        lignes_formset = LigneMouvementFormSet(request.POST)
        if form.is_valid():
            mouvement = form.save(commit=False)
            mouvement.collaborateur = request.user.get_full_name() or request.user.username
            mouvement.save()
            lignes_formset.instance = mouvement
            if lignes_formset.is_valid():
                for ligne_form in lignes_formset:
                    if ligne_form.cleaned_data and not ligne_form.cleaned_data.get('DELETE'):
                        if ligne_form.cleaned_data.get('cuve') or ligne_form.cleaned_data.get('volume_ambiant'):
                            ligne = ligne_form.save(commit=False)
                            ligne.produit = mouvement.produit
                            ligne.save()
                messages.success(request, f"Mouvement {mouvement.numero_enregistrement} enregistré avec succès.")
                return redirect('mouvement_liste')
    else:
        form = MouvementForm()
        lignes_formset = LigneMouvementFormSet()

    camions    = Camion.objects.select_related('marketeur').filter(statut='EN_SERVICE').order_by('immatriculation')
    cuves      = Cuve.objects.select_related('produit').filter(statut='ACTIVE').order_by('numero')
    marketeurs = Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale')

    return render(request, 'mouvements/saisie.html', {
        'form':            form,
        'lignes_formset':  lignes_formset,
        'titre':           'Nouveau mouvement',
        'camions':         camions,
        'cuves':           cuves,
        'marketeurs':      marketeurs,
        'mode':            'creer',
    })


@login_required
def mouvement_modifier(request, uuid, slug):
    """Formulaire d'édition d'un mouvement existant."""
    if _deny_marketeur(request):
        return redirect('mouvement_liste')

    mouvement = get_object_or_404(Mouvement, uuid=uuid)

    if request.method == 'POST':
        form = MouvementForm(request.POST, instance=mouvement)
        lignes_formset = LigneMouvementFormSet(request.POST, instance=mouvement)
        if form.is_valid() and lignes_formset.is_valid():
            mouvement = form.save()
            for ligne_form in lignes_formset:
                if ligne_form.cleaned_data:
                    if ligne_form.cleaned_data.get('DELETE'):
                        if ligne_form.instance.pk:
                            ligne_form.instance.delete()
                    elif ligne_form.cleaned_data.get('cuve') or ligne_form.cleaned_data.get('volume_ambiant'):
                        ligne = ligne_form.save(commit=False)
                        ligne.produit = mouvement.produit
                        ligne.save()
            # Synchroniser le produit dénormalisé sur toutes les lignes restantes
            LigneMouvement.objects.filter(mouvement=mouvement).update(produit=mouvement.produit)
            messages.success(
                request,
                f"Mouvement N° {mouvement.numero_enregistrement} modifié avec succès. "
                f"Les calculs automatiques ont été actualisés."
            )
            return redirect('mouvement_modifier', pk=mouvement.pk)
    else:
        form = MouvementForm(instance=mouvement)
        lignes_formset = LigneMouvementFormSet(instance=mouvement)

    camions    = Camion.objects.select_related('marketeur').filter(statut='EN_SERVICE').order_by('immatriculation')
    cuves      = Cuve.objects.select_related('produit').filter(statut='ACTIVE').order_by('numero')
    marketeurs = Marketeur.objects.filter(statut='ACTIF').order_by('raison_sociale')

    return render(request, 'mouvements/saisie.html', {
        'form':            form,
        'lignes_formset':  lignes_formset,
        'mouvement':       mouvement,
        'titre':           f'Modifier — {mouvement.numero_enregistrement}',
        'camions':         camions,
        'cuves':           cuves,
        'marketeurs':      marketeurs,
        'mode':            'modification',
    })


@login_required
@require_POST
def mouvement_calcul_preview(request):
    """
    Endpoint AJAX : retourne les calculs petroleum_calc pour la section
    "Contrôle final avant déchargement" sans enregistrer de mouvement.
    Entrée : JSON (type_mouvement='ENTREE' ou 'SORTIE' + champs numériques).
    Sortie : JSON avec les valeurs calculées.
    """
    import json
    from . import petroleum_calc as pc
    try:
        data = json.loads(request.body)
    except (ValueError, KeyError):
        return JsonResponse({'error': 'JSON invalide'}, status=400)

    def _f(key):
        v = data.get(key)
        try:
            return float(v) if v not in (None, '', 'null') else None
        except (ValueError, TypeError):
            return None

    result = {}
    type_mvt = data.get('type_mouvement', '')

    if type_mvt == 'ENTREE':
        v_amb_recu   = _f('volume_ambiant_recu')
        v_amb_expe   = _f('volume_ambiant_expediteur')
        temp_labo    = _f('temperature_labo')
        dens_obs     = _f('densite_observee_labo')
        temp_recept  = _f('temperature_reception')
        dens_expe    = _f('densite_15c_expediteur')

        # P/G réception (col. AC) — simple soustraction
        if v_amb_recu is not None and v_amb_expe is not None:
            result['perte_gain_reception'] = round(v_amb_recu - v_amb_expe, 2)

        # D@15°C calculée (col. AG) via API MPMS TRH_15
        d15 = None
        if dens_obs is not None and temp_labo is not None:
            try:
                d15 = round(pc.density_at_15c(dens_obs, temp_labo), 2)
                result['densite_15c_calculee'] = d15
            except Exception:
                pass

        # Écart densité (col. AH)
        if d15 is not None and dens_expe is not None:
            result['ecart_densite_15c'] = round((d15 - dens_expe) / 1000, 4)

        # Vcf @15°C (col. AI) via API MPMS TVCF_15
        vcf = None
        if d15 is not None and temp_recept is not None:
            try:
                vcf = round(pc.vcf_to_15c(d15, temp_recept), 4)
                result['coefficient_conversion_15c'] = vcf
            except Exception:
                pass

        # Volume @15°C reçu (col. AJ)
        if v_amb_recu is not None and vcf is not None:
            result['volume_15c_recu'] = round(v_amb_recu * vcf, 2)

        # P/G @15°C (col. AK)
        pg_recept = result.get('perte_gain_reception')
        if pg_recept is not None and vcf is not None:
            result['perte_gain_15c'] = round(pg_recept * vcf, 2)

        # Poids (col. AL)
        v15 = result.get('volume_15c_recu')
        if v15 is not None and d15 is not None:
            result['poids_kg'] = round(v15 * d15 / 1000, 2)

    elif type_mvt == 'SORTIE':
        v_amb_sortie = _f('volume_ambiant_sortie')
        dens_15_sortie = _f('densite_15c_sortie')
        temp_sortie  = _f('temperature_sortie')

        vcf = None
        if dens_15_sortie is not None and temp_sortie is not None:
            try:
                vcf = round(pc.vcf_to_15c(dens_15_sortie, temp_sortie), 4)
                result['coefficient_conversion_sortie'] = vcf
            except Exception:
                pass

        if v_amb_sortie is not None and vcf is not None:
            result['volume_15c_sortie'] = round(v_amb_sortie * vcf, 2)
            if dens_15_sortie is not None:
                result['poids_sortie_kg'] = round(result['volume_15c_sortie'] * dens_15_sortie / 1000, 2)

    elif type_mvt == 'CESSION':
        v_amb_cession  = _f('cession_volume_ambiant')
        dens_obs       = _f('cession_densite_observee')
        temp_cession   = _f('cession_temperature')

        # D@15°C calculée via TRH_15
        d15 = None
        if dens_obs is not None and temp_cession is not None:
            try:
                d15 = round(pc.density_at_15c(dens_obs, temp_cession), 2)
                result['cession_densite_15c'] = d15
            except Exception:
                pass

        # Vcf @15°C via TVCF_15
        vcf = None
        if d15 is not None and temp_cession is not None:
            try:
                vcf = round(pc.vcf_to_15c(d15, temp_cession), 4)
                result['cession_coefficient_vcf'] = vcf
            except Exception:
                pass

        # Vol. @15°C cédé
        if v_amb_cession is not None and vcf is not None:
            result['cession_volume_15c'] = round(v_amb_cession * vcf, 2)

    return JsonResponse(result)


@login_required
def mouvement_detail(request, uuid, slug):
    """Fiche de détail d'un mouvement — lecture seule."""
    mouvement = get_object_or_404(
        Mouvement.objects.select_related(
            'marketeur', 'produit', 'camion', 'camion__marketeur',
            'chauffeur', 'cession_marketeur_destinataire',
        ).prefetch_related('lignes__cuve__produit'),
        uuid=uuid
    )
    acces = _check_mouvement_acces(request, mouvement)
    if acces:
        return acces
    return render(request, 'mouvements/detail.html', {'mouvement': mouvement})


@login_required
def mouvement_supprimer(request, uuid, slug):
    """Suppression d'un mouvement — réservée au staff."""
    if not request.user.is_staff:
        messages.error(request, "La suppression de mouvements est réservée au staff.")
        return redirect('mouvement_liste')

    mouvement = get_object_or_404(Mouvement, uuid=uuid)

    if request.method == 'POST':
        mouvement.delete()
        messages.success(request, "Mouvement supprimé.")
        return redirect('mouvement_liste')

    return render(request, 'mouvements/confirmer_suppression.html', {'mouvement': mouvement})


# 
#  EXPORT PDF — MOUVEMENTS
# 

@login_required
def mouvement_detail_pdf(request, uuid, slug):
    """Télécharge la fiche détaillée d'un mouvement en PDF."""
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf

    mouvement = get_object_or_404(
        Mouvement.objects.select_related(
            'marketeur', 'produit', 'camion', 'camion__marketeur',
            'chauffeur', 'cession_marketeur_destinataire',
        ).prefetch_related('lignes__cuve__produit'),
        uuid=uuid,
    )
    acces = _check_mouvement_acces(request, mouvement)
    if acces:
        return acces
    filename = f"MVT_{mouvement.numero_enregistrement}.pdf"
    return render_to_pdf(
        'mouvements/detail_pdf.html',
        {
            'mouvement':    mouvement,
            'generated_at': timezone.now().strftime('%d/%m/%Y Ã  %H:%M'),
        },
        filename=filename,
    )


@login_required
def mouvement_bordereau(request, uuid, slug):
    """Bordereau de mouvement A4 imprimable (Ctrl+P → PDF navigateur)."""
    from django.utils import timezone as tz
    from SGDS.models import Societe

    mouvement = get_object_or_404(
        Mouvement.objects.select_related(
            "produit", "marketeur", "cuve", "cuve__produit",
            "camion", "chauffeur",
        ).prefetch_related(
            "lignes__cuve__produit",
        ),
        uuid=uuid,
    )
    acces = _check_mouvement_acces(request, mouvement)
    if acces:
        return acces
    societe = Societe.get_instance()

    def flag(name, default=True):
        v = request.GET.get(name)
        if v is None:
            return default
        return v not in ("0", "false", "no", "off", "")

    # Période comptable label
    try:
        from SGDS.services.periode_comptable import periode_pour_date
        periode = periode_pour_date(mouvement.date_mouvement)
        periode_label = str(periode) if periode else mouvement.date_mouvement.strftime("%B %Y")
    except Exception:
        periode_label = mouvement.date_mouvement.strftime("%B %Y")

    # Volumes calculés (entrée)
    vol_exp = float(mouvement.volume_15c_expediteur or 0)
    vol_recu = float(mouvement.volume_15c_recu or 0)
    ecart = vol_recu - vol_exp
    if vol_exp:
        ecart_signe = f"{'+' if ecart >= 0 else ''}{ecart:,.0f}".replace(",", "â€¯")
        ecart_pct = f"{'+' if ecart >= 0 else ''}{(ecart / vol_exp * 100):.2f}"
        tolerance_status = "OK" if abs(ecart / vol_exp) <= 0.005 else "HORS TOLÉRANCE"
    else:
        ecart_signe = "—"
        ecart_pct = "—"
        tolerance_status = "—"

    pg_amb = float(mouvement.perte_gain_reception or 0)
    pg_15c = float(mouvement.perte_gain_15c or 0)
    perte_gain_ambiant_signe = f"{'+' if pg_amb >= 0 else ''}{pg_amb:,.0f}".replace(",", "â€¯") if mouvement.perte_gain_reception is not None else "—"
    perte_gain_15c_signe = f"{'+' if pg_15c >= 0 else ''}{pg_15c:,.0f}".replace(",", "â€¯") if mouvement.perte_gain_15c is not None else "—"
    poids_volumique = float(mouvement.densite_15c_calculee) if mouvement.densite_15c_calculee else None

    ctx = {
        "mouvement": mouvement,
        "societe": societe,
        "now": tz.now(),
        "periode_label": periode_label,
        "show_calc": flag("calc", True),
        "show_sigs": flag("sigs", True),
        "show_stamp": flag("stamp", False),
        "compact": flag("compact", False),
        "bw": flag("bw", False),
        "auto_print": flag("auto", False),
        "ecart_signe": ecart_signe,
        "ecart_pct": ecart_pct,
        "tolerance_status": tolerance_status,
        "perte_gain_ambiant_signe": perte_gain_ambiant_signe,
        "perte_gain_15c_signe": perte_gain_15c_signe,
        "poids_volumique": poids_volumique,
        "statut_acquittement": mouvement.statut_acquittement,
    }
    return render(request, "mouvements/bordereau.html", ctx)


@login_required
def mouvement_bordereau_pdf(request, uuid, slug):
    """PDF server-side du bordereau (nécessite WeasyPrint — pip install weasyprint)."""
    try:
        from weasyprint import HTML, CSS
        from django.template.loader import render_to_string
        from django.contrib.staticfiles import finders
    except ImportError:
        from django.http import HttpResponse
        return HttpResponse(
            "WeasyPrint non installé. Installez-le avec : pip install weasyprint",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    # Réutilise la même vue pour construire le contexte
    fake_request = request
    from django.utils import timezone as tz
    from SGDS.models import Societe

    mouvement = get_object_or_404(
        Mouvement.objects.select_related(
            "produit", "marketeur", "cuve", "cuve__produit",
            "camion", "chauffeur",
        ).prefetch_related(
            "lignes__cuve__produit",
        ),
        uuid=uuid,
    )
    acces = _check_mouvement_acces(request, mouvement)
    if acces:
        return acces
    societe = Societe.get_instance()

    try:
        from SGDS.services.periode_comptable import periode_pour_date
        periode = periode_pour_date(mouvement.date_mouvement)
        periode_label = str(periode) if periode else mouvement.date_mouvement.strftime("%B %Y")
    except Exception:
        periode_label = mouvement.date_mouvement.strftime("%B %Y")

    vol_exp = float(mouvement.volume_15c_expediteur or 0)
    vol_recu = float(mouvement.volume_15c_recu or 0)
    ecart = vol_recu - vol_exp
    if vol_exp:
        ecart_signe = f"{'+' if ecart >= 0 else ''}{ecart:,.0f}".replace(",", "â€¯")
        ecart_pct = f"{'+' if ecart >= 0 else ''}{(ecart / vol_exp * 100):.2f}"
        tolerance_status = "OK" if abs(ecart / vol_exp) <= 0.005 else "HORS TOLÉRANCE"
    else:
        ecart_signe = "—"
        ecart_pct = "—"
        tolerance_status = "—"

    pg_amb = float(mouvement.perte_gain_reception or 0)
    pg_15c_val = float(mouvement.perte_gain_15c or 0)
    perte_gain_ambiant_signe = f"{'+' if pg_amb >= 0 else ''}{pg_amb:,.0f}".replace(",", "â€¯") if mouvement.perte_gain_reception is not None else "—"
    perte_gain_15c_signe = f"{'+' if pg_15c_val >= 0 else ''}{pg_15c_val:,.0f}".replace(",", "â€¯") if mouvement.perte_gain_15c is not None else "—"
    poids_volumique = float(mouvement.densite_15c_calculee) if mouvement.densite_15c_calculee else None

    ctx = {
        "mouvement": mouvement,
        "societe": societe,
        "now": tz.now(),
        "periode_label": periode_label,
        "show_calc": True,
        "show_sigs": True,
        "show_stamp": False,
        "compact": False,
        "bw": False,
        "auto_print": False,
        "ecart_signe": ecart_signe,
        "ecart_pct": ecart_pct,
        "tolerance_status": tolerance_status,
        "perte_gain_ambiant_signe": perte_gain_ambiant_signe,
        "perte_gain_15c_signe": perte_gain_15c_signe,
        "poids_volumique": poids_volumique,
        "statut_acquittement": mouvement.statut_acquittement,
    }
    html_string = render_to_string("mouvements/bordereau.html", ctx, request=request)
    css_path = finders.find("css/bordereau.css")
    pdf_bytes = HTML(
        string=html_string,
        base_url=request.build_absolute_uri("/"),
    ).write_pdf(
        stylesheets=[CSS(filename=css_path)] if css_path else None,
        presentational_hints=True,
    )
    from django.http import HttpResponse
    filename = f"bordereau_{mouvement.numero_enregistrement}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response


@login_required
def mouvements_liste_pdf(request):
    """Télécharge la liste filtrée des mouvements en PDF (max 500 lignes)."""
    from django.utils import timezone
    from SGDS.services.export_pdf import render_to_pdf

    qs = Mouvement.objects.select_related(
        'produit', 'marketeur', 'camion', 'cession_marketeur_destinataire',
    ).order_by('-date_mouvement', '-date_saisie')

    type_m     = request.GET.get('type', '').strip()
    regime     = request.GET.get('regime', '').strip()
    mkt_pk     = request.GET.get('marketeur', '').strip()
    produit_pk = request.GET.get('produit', '').strip()
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin   = request.GET.get('date_fin', '').strip()
    q          = request.GET.get('q', '').strip()

    if type_m:      qs = qs.filter(type_mouvement=type_m)
    if regime:      qs = qs.filter(regime_douanier=regime)
    if mkt_pk:      qs = qs.filter(marketeur_id=mkt_pk)
    if produit_pk:  qs = qs.filter(produit_id=produit_pk)
    if date_debut:  qs = qs.filter(date_mouvement__gte=date_debut)
    if date_fin:    qs = qs.filter(date_mouvement__lte=date_fin)
    if q:
        qs = qs.filter(
            Q(numero_enregistrement__icontains=q) |
            Q(camion__immatriculation__icontains=q) |
            Q(bl_expediteur__icontains=q) |
            Q(marketeur__raison_sociale__icontains=q)
        )

    nb_total   = qs.count()
    mouvements = list(qs[:500])
    today      = timezone.now().strftime('%Y%m%d')

    return render_to_pdf(
        'mouvements/liste_pdf.html',
        {
            'mouvements':   mouvements,
            'nb_total':     nb_total,
            'generated_at': timezone.now().strftime('%d/%m/%Y Ã  %H:%M'),
            'filtres': {
                'type': type_m, 'regime': regime, 'marketeur': mkt_pk,
                'produit': produit_pk, 'date_debut': date_debut,
                'date_fin': date_fin, 'q': q,
            },
        },
        filename=f"Mouvements_{today}.pdf",
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  PÉRIODES COMPTABLES
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views import View
from django.views.generic import ListView


class ListePeriodesView(LoginRequiredMixin, ListView):
    """Liste chronologique des périodes comptables + bouton d'ouverture."""
    template_name   = 'periode/liste.html'
    context_object_name = 'periodes'
    paginate_by     = 24
    ordering        = ['-annee', '-mois']

    def get_queryset(self):
        from .models import PeriodeComptable
        return PeriodeComptable.objects.all().order_by('-annee', '-mois')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from .models import PeriodeComptable
        from .services.periode_comptable import mois_suivant

        derniere = PeriodeComptable.objects.order_by('-annee', '-mois').first()

        if derniere is None:
            # Aucune période : proposer le mois courant comme première
            from django.utils import timezone
            aujourd_hui = timezone.now().date()
            ctx['peut_ouvrir_suivante'] = True
            ctx['premiere_periode']     = True
            ctx['mois_a_ouvrir']        = aujourd_hui.month
            ctx['annee_a_ouvrir']       = aujourd_hui.year
        elif derniere.statut == 'CLOTUREE':
            m, a = mois_suivant(derniere.mois, derniere.annee)
            ctx['peut_ouvrir_suivante'] = True
            ctx['premiere_periode']     = False
            ctx['mois_a_ouvrir']        = m
            ctx['annee_a_ouvrir']       = a
        else:
            ctx['peut_ouvrir_suivante'] = False
            ctx['premiere_periode']     = False
            ctx['mois_a_ouvrir']        = None
            ctx['annee_a_ouvrir']       = None

        return ctx


class OuvrirPeriodeView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Confirmation et exécution de l'ouverture d'une période."""

    def test_func(self):
        return self.request.user.is_staff

    def _get_mois_annee(self, request):
        from django.utils import timezone
        from .models import PeriodeComptable
        from .services.periode_comptable import mois_suivant

        # Paramètres GET ou POST
        source = request.GET if request.method == 'GET' else request.POST
        try:
            mois  = int(source.get('mois', 0))
            annee = int(source.get('annee', 0))
            if not (1 <= mois <= 12 and 2020 <= annee <= 2100):
                raise ValueError
        except (ValueError, TypeError):
            # Fallback : calculer depuis la dernière période
            derniere = PeriodeComptable.objects.order_by('-annee', '-mois').first()
            if derniere is None:
                aujourd_hui = timezone.now().date()
                mois, annee = aujourd_hui.month, aujourd_hui.year
            else:
                mois, annee = mois_suivant(derniere.mois, derniere.annee)
        return mois, annee

    def get(self, request):
        from .models import PeriodeComptable
        mois, annee   = self._get_mois_annee(request)
        premiere      = not PeriodeComptable.objects.exists()
        return render(request, 'periode/ouvrir_confirm.html', {
            'mois_propose':  mois,
            'annee_propose': annee,
            'premiere':      premiere,
        })

    def post(self, request):
        from django.core.exceptions import ValidationError
        from .services.periode_comptable import ouvrir_periode
        mois, annee = self._get_mois_annee(request)
        try:
            ouvrir_periode(mois, annee, user=request.user)
            messages.success(request, f"Période {mois}/{annee} ouverte avec succès.")
            return redirect('periode_liste')
        except ValidationError as exc:
            messages.error(request, exc.message)
            from .models import PeriodeComptable
            premiere = not PeriodeComptable.objects.exists()
            return render(request, 'periode/ouvrir_confirm.html', {
                'mois_propose':  mois,
                'annee_propose': annee,
                'premiere':      premiere,
            })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  COULAGE — RÉPARTITION MENSUELLE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _rapport_depuis_snapshot(cloture):
    """
    Reconstruit le dict rapport compatible avec repartition.html
    depuis la snapshot DB (ClotureCoulageMensuel).
    """
    from decimal import Decimal
    from collections import defaultdict

    D0 = Decimal('0')

    produits_list = [
        cp.produit
        for cp in cloture.produits_coulage
            .select_related('produit')
            .order_by('produit__nom')
    ]

    coefficients = {}
    pertes_gains = {}
    cumuls       = {}
    for cp in cloture.produits_coulage.all():
        pk = cp.produit_id
        coefficients[pk] = cp.coefficient
        pertes_gains[pk] = cp.pertes_gains
        cumuls[pk] = {
            'brut_entree': cp.cumul_entree,
            'coul_entree': D0,
            'entree':      cp.cumul_entree,
            'sortie':      cp.cumul_sortie,
        }

    # Grouper les lignes par marketeur
    mkt_map = defaultdict(dict)
    mkt_meta = {}
    for ligne in cloture.lignes.select_related('marketeur', 'produit').all():
        mkt = ligne.marketeur
        mkt_meta[mkt.pk] = {
            'obj':          mkt,
            'motif':        ligne.motif,
            'prix_unitaire': ligne.prix_unitaire,
        }
        if ligne.produit_id is not None:
            mkt_map[mkt.pk][ligne.produit_id] = {
                'brut_entree':  ligne.brut_entree,
                'coul_entree':  ligne.coul_entree,
                'entree_nette': ligne.entree_nette,
                'sortie':       ligne.sortie,
                'base_qp_coul': ligne.base_qp_coul,
                'coef_qp_coul': ligne.coef_qp_coul,
                'qp_coul':      ligne.qp_coul,
                'volume_sorti': ligne.volume_sorti,
            }

    lignes = []
    for mkt_pk, par_produit in mkt_map.items():
        meta  = mkt_meta[mkt_pk]
        v_tot = sum((pp.get('volume_sorti', D0) for pp in par_produit.values()), D0)
        m_tot = sum(
            (pp.get('volume_sorti', D0) * meta['prix_unitaire']
             for pp in par_produit.values()), D0
        )
        lignes.append({
            'marketeur':          meta['obj'],
            'par_produit':        par_produit,
            'volume_global_sorti': v_tot,
            'motif':              meta['motif'],
            'prix_unitaire':      meta['prix_unitaire'],
            'montant':            m_tot,
        })
    # Trier par raison sociale pour un affichage stable
    lignes.sort(key=lambda x: x['marketeur'].raison_sociale)

    # Totaux agrégés
    totaux_par_produit = {}
    for produit in produits_list:
        pk = produit.pk
        def _s(field):
            return sum((l['par_produit'].get(pk, {}).get(field, D0) for l in lignes), D0)
        totaux_par_produit[pk] = {
            'brut_entree':  _s('brut_entree'),
            'coul_entree':  _s('coul_entree'),
            'entree_nette': _s('entree_nette'),
            'sortie':       _s('sortie'),
            'base_qp_coul': _s('base_qp_coul'),
            'coef_qp_coul': coefficients.get(pk, D0),
            'qp_coul':      _s('qp_coul'),
            'volume_sorti': _s('volume_sorti'),
        }

    totaux = {
        'par_produit':         totaux_par_produit,
        'volume_global_sorti': sum((l['volume_global_sorti'] for l in lignes), D0),
        'motif':               cloture.motif,
        'prix_unitaire':       cloture.prix_unitaire_passage,
        'montant':             cloture.total_montant,
    }

    return {
        'periode':      cloture.periode,
        'produits':     produits_list,
        'coefficients': coefficients,
        'pertes_gains': pertes_gains,
        'cumuls':       cumuls,
        'lignes':       lignes,
        'totaux':       totaux,
        'parametres': {
            'prix_unitaire_passage': cloture.prix_unitaire_passage,
            'motif':                 cloture.motif,
        },
    }


class ListePeriodesCoulageView(LoginRequiredMixin, ListView):
    """Liste des périodes avec leur statut de coulage."""
    template_name        = 'coulage/liste_periodes.html'
    context_object_name  = 'object_list'
    paginate_by          = 24

    def get_queryset(self):
        from .models import PeriodeComptable
        return (
            PeriodeComptable.objects
            .select_related('cloture_coulage')
            .order_by('-annee', '-mois')
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['produits_actifs_list'] = list(
            Produit.objects.filter(statut='ACTIF').order_by('nom')
        )
        return ctx


class RepartitionCoulageView(LoginRequiredMixin, View):
    """Affiche la répartition du coulage d'une période (live ou snapshot)."""

    def get(self, request, periode_id):
        from .models import PeriodeComptable
        from .services.coulage_repartition import calculer_repartition_coulage

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)

        if periode.statut == 'CLOTUREE':
            try:
                cloture = periode.cloture_coulage
                rapport = _rapport_depuis_snapshot(cloture)
                source  = 'SNAPSHOT'
            except Exception:
                rapport = calculer_repartition_coulage(periode)
                source  = 'LIVE'
        else:
            rapport = calculer_repartition_coulage(periode)
            source  = 'LIVE'

        peut_cloturer = (
            periode.statut == 'OUVERTE'
            and request.user.is_staff
            and not hasattr(periode, 'cloture_coulage')
        )
        # Vérification plus fiable :
        try:
            periode.cloture_coulage
            deja_cloture = True
        except Exception:
            deja_cloture = False
        peut_cloturer = periode.statut == 'OUVERTE' and request.user.is_staff

        return render(request, 'coulage/repartition.html', {
            'periode':       periode,
            'rapport':       rapport,
            'source':        source,
            'peut_cloturer': peut_cloturer,
        })


class ClotureCoulageView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Clôture une période comptable (POST)."""

    def test_func(self):
        return self.request.user.is_staff

    def post(self, request, periode_id):
        from .models import PeriodeComptable
        from .services.periode_comptable import cloturer_periode
        from django.core.exceptions import ValidationError

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)
        notes   = request.POST.get('notes', '').strip() or None

        try:
            cloturer_periode(periode, user=request.user, notes=notes)
            from .services.periode_comptable import mois_suivant
            m, a = mois_suivant(periode.mois, periode.annee)
            messages.success(
                request,
                f"Période {periode.libelle} clôturée. "
                f"Pour commencer le mois suivant, ouvrez la période {m}/{a} "
                f"depuis la liste des périodes."
            )
        except ValidationError as exc:
            messages.error(request, exc.message)

        return redirect('coulage_detail', periode_id=periode_id)

    def get(self, request, periode_id):
        return redirect('coulage_detail', periode_id=periode_id)


class ExportCoulageExcelView(LoginRequiredMixin, View):
    """Export Excel de la répartition du coulage."""

    def get(self, request, periode_id):
        from .models import PeriodeComptable
        from .services.coulage_repartition import calculer_repartition_coulage
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)

        if periode.statut == 'CLOTUREE':
            try:
                rapport = _rapport_depuis_snapshot(periode.cloture_coulage)
            except Exception:
                rapport = calculer_repartition_coulage(periode)
        else:
            rapport = calculer_repartition_coulage(periode)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Coulage {periode.libelle}"

        navy_fill = PatternFill('solid', fgColor='1E3A5F')
        navy_font = Font(color='FFFFFF', bold=True, size=10)
        header_font = Font(bold=True, size=10)
        normal_font = Font(size=9)
        center     = Alignment(horizontal='center', vertical='center')
        right_al   = Alignment(horizontal='right')
        thin_side  = Side(style='thin')
        thin_border = Border(left=thin_side, right=thin_side,
                             bottom=thin_side, top=thin_side)
        num_fmt = '#,##0'
        coef_fmt = '0.00000000'

        produits = rapport['produits']

        # â"€â"€ Titre â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        ws.merge_cells(f'A1:{get_column_letter(6 + len(produits) * 7)}1')
        titre_cell = ws['A1']
        titre_cell.value = f"Répartition du coulage — {periode.libelle}"
        titre_cell.font  = Font(bold=True, size=14, color='1E3A5F')
        ws.row_dimensions[1].height = 24

        # â"€â"€ En-têtes â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        row = 3
        ws.cell(row=row, column=1, value='Code').font       = navy_font
        ws.cell(row=row, column=1).fill                     = navy_fill
        ws.cell(row=row, column=1).alignment                = center
        ws.cell(row=row, column=2, value='Marketeur').font  = navy_font
        ws.cell(row=row, column=2).fill                     = navy_fill

        col = 3
        for produit in produits:
            for sub in [f'Brut Entrée {produit.code}',
                        f'Coul. Entrée {produit.code}',
                        f'Nette {produit.code}']:
                c = ws.cell(row=row, column=col, value=sub)
                c.font = navy_font; c.fill = navy_fill; c.alignment = center
                col += 1

        for produit in produits:
            c = ws.cell(row=row, column=col, value=f'Sortie {produit.code}')
            c.font = navy_font; c.fill = navy_fill; c.alignment = center
            col += 1

        for produit in produits:
            for sub in [f'Base QP {produit.code}',
                        f'Coef. {produit.code}',
                        f'QP Coul. {produit.code}']:
                c = ws.cell(row=row, column=col, value=sub)
                c.font = navy_font; c.fill = navy_fill; c.alignment = center
                col += 1

        for hdr in ['Vol. global sorti', 'PU/L', 'Montant FCFA']:
            c = ws.cell(row=row, column=col, value=hdr)
            c.font = navy_font; c.fill = navy_fill; c.alignment = center
            col += 1

        # â"€â"€ Données â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        for ligne in rapport['lignes']:
            row += 1
            ws.cell(row=row, column=1,
                    value=f"MK{ligne['marketeur'].pk:03d}").font = normal_font
            ws.cell(row=row, column=2,
                    value=ligne['marketeur'].sigle or ligne['marketeur'].raison_sociale)

            col = 3
            for produit in produits:
                pp = ligne['par_produit'].get(produit.pk, {})
                for field in ['brut_entree', 'coul_entree', 'entree_nette']:
                    c = ws.cell(row=row, column=col, value=float(pp.get(field, 0)))
                    c.number_format = num_fmt; c.alignment = right_al
                    col += 1

            for produit in produits:
                pp = ligne['par_produit'].get(produit.pk, {})
                c = ws.cell(row=row, column=col, value=float(pp.get('sortie', 0)))
                c.number_format = num_fmt; c.alignment = right_al
                col += 1

            for produit in produits:
                pp = ligne['par_produit'].get(produit.pk, {})
                c1 = ws.cell(row=row, column=col, value=float(pp.get('base_qp_coul', 0)))
                c1.number_format = num_fmt; c1.alignment = right_al; col += 1
                c2 = ws.cell(row=row, column=col, value=float(pp.get('coef_qp_coul', 0)))
                c2.number_format = coef_fmt; c2.alignment = right_al; col += 1
                c3 = ws.cell(row=row, column=col, value=float(pp.get('qp_coul', 0)))
                c3.number_format = num_fmt; c3.alignment = right_al; col += 1

            c = ws.cell(row=row, column=col, value=float(ligne['volume_global_sorti']))
            c.number_format = num_fmt; c.alignment = right_al; col += 1
            ws.cell(row=row, column=col, value=float(ligne['prix_unitaire'])); col += 1
            c = ws.cell(row=row, column=col, value=float(ligne['montant']))
            c.number_format = num_fmt; c.alignment = right_al; col += 1

        # â"€â"€ Ligne totaux â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
        row += 1
        c = ws.cell(row=row, column=1, value='TOTAUX')
        c.font = header_font
        c = ws.cell(row=row, column=col - 1,  # Montant
                    value=float(rapport['totaux']['montant']))
        c.number_format = num_fmt; c.alignment = right_al
        c.font = Font(bold=True)

        # Largeurs colonnes
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 22
        for i in range(3, col):
            ws.column_dimensions[get_column_letter(i)].width = 13

        # Réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="coulage_{periode.libelle.replace(" ", "_")}.xlsx"'
        )
        wb.save(response)
        return response


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  SUIVI ÉVOLUTION JOURNALIER
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

class SuiviEvolutionView(LoginRequiredMixin, View):
    """Tableau journalier stock/mouvements pour un produit sur une période."""

    def get(self, request, periode_id, produit_id):
        from .models import PeriodeComptable, Produit
        from .services.suivi_evolution import calculer_suivi_evolution

        periode  = get_object_or_404(PeriodeComptable, pk=periode_id)
        produit  = get_object_or_404(Produit, pk=produit_id)
        produits = list(Produit.objects.filter(statut='ACTIF').order_by('nom'))
        rapport  = calculer_suivi_evolution(periode, produit)

        return render(request, 'coulage/suivi_evolution.html', {
            'periode':  periode,
            'produit':  produit,
            'produits': produits,
            'rapport':  rapport,
        })


class ExportSuiviExcelView(LoginRequiredMixin, View):
    """Export Excel du suivi évolution journalier."""

    def get(self, request, periode_id, produit_id):
        from .models import PeriodeComptable, Produit
        from .services.suivi_evolution import calculer_suivi_evolution
        from .services.export_excel import exporter_suivi_xlsx
        from django.http import HttpResponse

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)
        produit = get_object_or_404(Produit, pk=produit_id)
        rapport = calculer_suivi_evolution(periode, produit)

        contenu = exporter_suivi_xlsx(rapport)
        response = HttpResponse(
            contenu,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        nom = f"suivi_{produit.code}_{periode.libelle.replace(' ', '_')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nom}"'
        return response


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  FRAIS DE PASSAGE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

class FraisPassageView(LoginRequiredMixin, View):
    """Document de facturation mensuel — frais de passage par marketeur."""

    def get(self, request, periode_id):
        from .models import PeriodeComptable
        from .services.frais_passage import calculer_frais_passage

        periode = get_object_or_404(PeriodeComptable, pk=periode_id)
        rapport = calculer_frais_passage(periode)

        return render(request, 'coulage/frais_passage.html', {
            'periode': periode,
            'rapport': rapport,
        })


class ExportFraisPassageExcelView(LoginRequiredMixin, View):
    """Export Excel des frais de passage."""

    def get(self, request, periode_id):
        from .models import PeriodeComptable
        from .services.frais_passage import calculer_frais_passage
        from .services.export_excel import exporter_frais_passage_xlsx
        from django.http import H